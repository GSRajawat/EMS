import streamlit as st
import pandas as pd
import os
import zipfile
import io
import fitz # PyMuPDF
import re
import tempfile
import ast
import requests
from datetime import date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import json
from supabase import create_client, Client
import datetime
import numpy as np
import traceback


# Initialize Supabase
try:
    SUPABASE_URL = st.secrets["supabase"]["url"]
    SUPABASE_KEY = st.secrets["supabase"]["key"]
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json"
    }
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
except KeyError:
    st.error("Supabase secrets not found. Please configure `supabase.url` and `supabase.key` in your secrets.toml file.")
    st.stop()


# --- Supabase Helper Functions (Moved to top for proper definition scope) ---

def save_prep_closing_assignments_to_supabase(data_to_save):
    """
    Saves preparation and closing day assignments to the Supabase table.
    'data_to_save' should be a list of dictionaries.
    This version correctly handles updating unique data for a given set of classes.
    """
    if not supabase:
        return False, "Supabase client not initialized."

    if not data_to_save:
        return True, "No prep/closing assignments to save." # If data_to_save is empty

    try:
        # Get the selected_classes from the first item to use for the targeted delete
        # This assumes all items in the list belong to the same class selection.
        selected_classes_list = data_to_save[0].get("selected_classes", [])
        selected_classes_json = json.dumps(selected_classes_list)

        # 1. First, delete all existing entries that match the specific selected_classes.
        # This acts as a clean "upsert" for this specific class selection.
        supabase.table("prep_closing_assignments").delete().eq("selected_classes", selected_classes_json).execute()
        
        # 2. Prepare and insert the new, cleaned data.
        cleaned_data = []
        for item in data_to_save:
            cleaned_item = {
                "name": item.get("name"),
                "role": item.get("role"),
                "prep_days": json.dumps(item.get("prep_days")) if item.get("prep_days") else None,
                "closing_days": json.dumps(item.get("closing_days")) if item.get("closing_days") else None,
                "selected_classes": json.dumps(item.get("selected_classes")) if item.get("selected_classes") else None
            }
            cleaned_data.append(cleaned_item)

        response = supabase.table("prep_closing_assignments").insert(cleaned_data).execute()
        
        # Check for errors in the Supabase response
        if response.data:
            return True, f"✅ Saved {len(cleaned_data)} prep/closing assignments for these classes to Supabase."
        else:
            return False, f"❌ Supabase error saving prep/closing assignments: {response.status_code} - {response.content}"

    except Exception as e:
        traceback.print_exc()
        return False, f"❌ Error saving prep/closing assignments to Supabase: {e}"

# New function to load prep_closing_assignments from Supabase
def load_prep_closing_assignments_from_supabase():
    """
    Loads preparation and closing day assignments from the Supabase table.
    Returns a dictionary structured as {name: {role: ..., prep_days: [...], closing_days: [...]}}
    """
    if not supabase:
        return {}

    try:
        response = supabase.table("prep_closing_assignments").select("*").execute()
        
        if not response.data:
            return {}

        loaded_data = {}
        for row in response.data:
            name = row.get('name')
            role = row.get('role')
            prep_days = json.loads(row.get('prep_days')) if row.get('prep_days') else []
            closing_days = json.loads(row.get('closing_days')) if row.get('closing_days') else []
            selected_classes = json.loads(row.get('selected_classes')) if row.get('selected_classes') else []

            # Structure the loaded data to match the expected format for 'prep_closing_assignments'
            loaded_data[name] = {
                'role': role,
                'prep_days': prep_days,
                'closing_days': closing_days,
                'selected_classes': selected_classes # Include selected_classes here
            }
        return loaded_data

    except Exception as e:
        traceback.print_exc()
        st.error(f"❌ Error loading prep/closing assignments from Supabase: {e}")
        return {}


# New function to save global settings (like holiday dates) to Supabase
def save_global_setting_to_supabase(setting_key, setting_value):
    """
    Saves a single global setting to the Supabase global_settings table.
    Overwrites if the key already exists.
    """
    if not supabase:
        return False, "Supabase client not initialized."

    try:
        # Delete existing entry for the given setting_key (to ensure a clean upsert)
        supabase.table("global_settings").delete().eq("setting_key", setting_key).execute()
        
        # Insert the new value
        response = supabase.table("global_settings").insert({
            "setting_key": setting_key,
            "setting_value": json.dumps(setting_value) if setting_value is not None else None
        }).execute()

        if response.data:
            return True, f"✅ Saved setting '{setting_key}' to Supabase."
        else:
            return False, f"❌ Supabase error saving setting '{setting_key}': {response.status_code} - {response.content}"

    except Exception as e:
        traceback.print_exc()
        return False, f"❌ Error saving setting '{setting_key}' to Supabase: {e}"

# New function to load global settings from Supabase
def load_global_setting_from_supabase(setting_key):
    """
    Loads a single global setting from the Supabase global_settings table.
    Returns the setting value, or None if not found.
    """
    if not supabase:
        return None

    try:
        response = supabase.table("global_settings").select("setting_value").eq("setting_key", setting_key).single().execute()
        
        if response.data and 'setting_value' in response.data:
            return json.loads(response.data['setting_value'])
        return None

    except Exception as e:
        # This will catch if .single() doesn't find a record, which is expected
        # traceback.print_exc() # Uncomment for debugging
        return None

# --- Configuration ---
CS_REPORTS_FILE = "cs_reports.csv"
EXAM_TEAM_MEMBERS_FILE = "exam_team_members.csv"
SHIFT_ASSIGNMENTS_FILE = "shift_assignments.csv"
ROOM_INVIGILATORS_FILE = "room_invigilator_assignments.csv"
SITTING_PLAN_FILE = "sitting_plan.csv"
TIMETABLE_FILE = "timetable.csv"
ASSIGNED_SEATS_FILE = "assigned_seats.csv"
ATTESTATION_DATA_FILE = "attestation_data_combined.csv"
COLLEGE_STATISTICS_FILE = "college_statistics_fancy.csv"

# --- NEW: Add these lines so the app knows where to store them ---
PREP_CLOSING_ASSIGNMENTS_FILE = "prep_closing_assignments.csv"
GLOBAL_SETTINGS_FILE = "global_settings.csv"


# --- Session State Initialization (MUST be at the top level of the script) ---
if 'selected_classes_for_bill_state' not in st.session_state:
    st.session_state.selected_classes_for_bill_state = []
if 'current_prep_closing_input' not in st.session_state:
    st.session_state.current_prep_closing_input = {}
if 'holiday_dates_input_state' not in st.session_state:
    st.session_state.holiday_dates_input_state = ""


# --- The rest of the file remains the same ---

def upload_csv_to_supabase(table_name, csv_path, unique_cols=None):
    try:
        df = pd.read_csv(csv_path)
        df.columns = df.columns.str.strip()

        # Column name mapping from CSV headers to database columns
        column_mappings = {
            'Roll Number': 'roll_number', 'Paper Code': 'paper_code', 'Paper Name': 'paper_name',
            'Room Number': 'room_number', 'Seat Number': 'seat_number', 'date': 'date',
            'shift': 'shift', 'SN': 'sn', 'Time': 'time', 'Class': 'class', 'Paper': 'paper', 'Name': 'name',
            'Roll Number 1': 'roll_number_1', 'Roll Number 2': 'roll_number_2',
            'Roll Number 3': 'roll_number_3', 'Roll Number 4': 'roll_number_4',
            'Roll Number 5': 'roll_number_5', 'Roll Number 6': 'roll_number_6',
            'Roll Number 7': 'roll_number_7', 'Roll Number 8': 'roll_number_8',
            'Roll Number 9': 'roll_number_9', 'Roll Number 10': 'roll_number_10',
            'Mode': 'mode', 'Type': 'type',
            'Seat Number 1': 'seat_number_1', 'Seat Number 2': 'seat_number_2',
            'Seat Number 3': 'seat_number_3', 'Seat Number 4': 'seat_number_4',
            'Seat Number 5': 'seat_number_5', 'Seat Number 6': 'seat_number_6',
            'Seat Number 7': 'seat_number_7', 'Seat Number 8': 'seat_number_8',
            'Seat Number 9': 'seat_number_9', 'Seat Number 10': 'seat_number_10',
            'Enrollment Number': 'enrollment_number', 'Session': 'session',
            'Regular/Backlog': 'regular_backlog', 'Father\'s Name': 'father_name',
            'Mother\'s Name': 'mother_name', 'Gender': 'gender',
            'Exam Name': 'exam_name', 'Exam Centre': 'exam_centre',
            'College Name': 'college_name', 'Address': 'address',
            'Paper 1': 'paper_1', 'Paper 2': 'paper_2', 'Paper 3': 'paper_3',
            'Paper 4': 'paper_4', 'Paper 5': 'paper_5', 'Paper 6': 'paper_6',
            'Paper 7': 'paper_7', 'Paper 8': 'paper_8', 'Paper 9': 'paper_9', 'Paper 10': 'paper_10',
            'report_key': 'report_key', 'room_num': 'room_num',
            'absent_roll_numbers': 'absent_roll_numbers', 'ufm_roll_numbers': 'ufm_roll_numbers',
            'invigilators': 'invigilators',
            'senior_center_superintendent': 'senior_center_superintendent',
            'center_superintendent': 'center_superintendent',
            'assistant_center_superintendent': 'assistant_center_superintendent',
            'permanent_invigilator': 'permanent_invigilator',
            'assistant_permanent_invigilator': 'assistant_permanent_invigilator',
            'class_3_worker': 'class_3_worker', 'class_4_worker': 'class_4_worker',
            # NEW MAPPINGS
            'prep_days': 'prep_days', 'closing_days': 'closing_days',
            'selected_classes': 'selected_classes', 'setting_key': 'setting_key',
            'setting_value': 'setting_value'
        }
        
        df.rename(columns=column_mappings, inplace=True)
        df = df.replace(r'^\s*$', None, regex=True)
        df = df.replace([np.inf, -np.inf], None)

        for col in df.columns:
            df[col] = df[col].apply(lambda x: None if pd.isna(x) or (isinstance(x, float) and not np.isfinite(x)) or str(x).strip() == '' else x)
        
        if 'date' in df.columns:
            def convert_date_format(date_str):
                if pd.notna(date_str) and isinstance(date_str, str) and len(date_str.split('-')) == 3:
                    try:
                        return datetime.datetime.strptime(date_str, '%d-%m-%Y').strftime('%Y-%m-%d')
                    except ValueError:
                        return date_str
                return date_str
            df['date'] = df['date'].apply(convert_date_format)

        # JSON fields handling (Added new fields)
        json_fields = [
            'absent_roll_numbers', 'ufm_roll_numbers', 'invigilators',
            'senior_center_superintendent', 'center_superintendent',
            'assistant_center_superintendent', 'permanent_invigilator',
            'assistant_permanent_invigilator', 'class_3_worker', 'class_4_worker',
            'prep_days', 'closing_days', 'selected_classes', 'setting_value'
        ]
        
        for field in json_fields:
            if field in df.columns:
                def parse_json_field(x):
                    if pd.notna(x) and isinstance(x, str) and x.strip():
                        try:
                            if x.strip().startswith('['):
                                return ast.literal_eval(x)
                            return [x.strip()]
                        except (ValueError, SyntaxError):
                            return [x.strip()]
                    return None
                df[field] = df[field].apply(parse_json_field)

        numeric_fields = ['room_number', 'seat_number', 'room_num', 'sn'] + [f'seat_number_{i}' for i in range(1, 11)]
        for field in numeric_fields:
            if field in df.columns:
                df[field] = pd.to_numeric(df[field], errors='coerce')
                df[field] = df[field].astype('Int64')

        if df.empty:
            return False, f"⚠️ `{csv_path}` is empty."

        records = df.to_dict(orient='records')
        cleaned_records = []
        for record in records:
            cleaned_record = {}
            for key, value in record.items():
                if isinstance(value, list):
                    cleaned_record[key] = value if value else None
                elif value is None or (not isinstance(value, list) and pd.isna(value)):
                    cleaned_record[key] = None
                elif isinstance(value, (np.int64, np.int32)):
                    cleaned_record[key] = int(value)
                elif isinstance(value, (np.float64, np.float32)):
                    cleaned_record[key] = float(value) if np.isfinite(value) else None
                else:
                    cleaned_record[key] = value
            cleaned_records.append(cleaned_record)

        batch_size = 100
        total_uploaded = 0
        for i in range(0, len(cleaned_records), batch_size):
            batch = cleaned_records[i:i + batch_size]
            supabase.table(table_name).insert(batch).execute()
            total_uploaded += len(batch)

        return True, f"✅ Uploaded {total_uploaded} rows to `{table_name}`."

    except Exception as e:
        return False, f"❌ Error uploading to `{table_name}`: {str(e)}"

# MODIFIED: download_supabase_to_csv (to handle API exceptions)
def download_supabase_to_csv(table_name, filename):
    all_data = []
    limit = 1000
    offset = 0

    try:
        while True:
            response = supabase.from_(table_name).select("*").limit(limit).offset(offset).execute()
            if not response.data:
                break
            all_data.extend(response.data)
            if len(response.data) < limit:
                break
            offset += limit
    
    except Exception as e:
        traceback.print_exc()
        return False, f"❌ Supabase API Error for '{table_name}': {e}"
    
    if not all_data:
        return True, f"⚠️ No data found in table `{table_name}`. An empty file has been created."

    df = pd.DataFrame(all_data)
    columns_to_drop = ['id', 'created_at']
    df = df.drop(columns=[col for col in columns_to_drop if col in df.columns])
    
    reverse_column_mappings = {
        'roll_number': 'Roll Number', 'paper_code': 'Paper Code', 'paper_name': 'Paper Name',
        'room_number': 'Room Number', 'seat_number': 'Seat Number', 'date': 'date',
        'shift': 'shift', 'sn': 'SN', 'time': 'Time', 'class': 'Class', 'paper': 'Paper', 'name': 'Name',
        'roll_number_1': 'Roll Number 1', 'roll_number_2': 'Roll Number 2',
        'roll_number_3': 'Roll Number 3', 'Roll Number 4': 'Roll Number 4',
        'roll_number_5': 'Roll Number 5', 'roll_number_6': 'Roll Number 6',
        'roll_number_7': 'Roll Number 7', 'Roll Number 8': 'Roll Number 8',
        'roll_number_9': 'Roll Number 9', 'Roll Number 10': 'Roll Number 10',
        'mode': 'Mode', 'type': 'Type',
        'seat_number_1': 'Seat Number 1', 'seat_number_2': 'Seat Number 2',
        'seat_number_3': 'Seat Number 3', 'seat_number_4': 'Seat Number 4',
        'seat_number_5': 'Seat Number 5', 'seat_number_6': 'Seat Number 6',
        'seat_number_7': 'Seat Number 7', 'seat_number_8': 'Seat Number 8',
        'seat_number_9': 'Seat Number 9', 'seat_number_10': 'Seat Number 10',
        'enrollment_number': 'Enrollment Number', 'session': 'Session',
        'regular_backlog': 'Regular/Backlog', 'father_name': 'Father\'s Name',
        'mother_name': 'Mother\'s Name', 'gender': 'Gender',
        'exam_name': 'Exam Name', 'exam_centre': 'Exam Centre',
        'college_name': 'College Name', 'address': 'Address',
        'paper_1': 'Paper 1', 'paper_2': 'Paper 2', 'paper_3': 'Paper 3',
        'paper_4': 'Paper 4', 'paper_5': 'Paper 5', 'paper_6': 'Paper 6',
        'paper_7': 'Paper 7', 'paper_8': 'Paper 8', 'paper_9': 'Paper 9', 'paper_10': 'Paper 10',
        'report_key': 'report_key', 'room_num': 'room_num',
        'absent_roll_numbers': 'absent_roll_numbers', 'ufm_roll_numbers': 'ufm_roll_numbers',
        'invigilators': 'invigilators',
        'senior_center_superintendent': 'senior_center_superintendent',
        'center_superintendent': 'center_superintendent',
        'assistant_center_superintendent': 'assistant_center_superintendent',
        'permanent_invigilator': 'permanent_invigilator',
        'assistant_permanent_invigilator': 'assistant_permanent_invigilator',
        'class_3_worker': 'class_3_worker', 'class_4_worker': 'class_4_worker',
        # NEW MAPPINGS
        'prep_days': 'prep_days', 'closing_days': 'closing_days',
        'selected_classes': 'selected_classes', 'setting_key': 'setting_key',
        'setting_value': 'setting_value'
    }
    actual_reverse_column_mappings = {k: v for k, v in reverse_column_mappings.items() if k in df.columns}
    df.rename(columns=actual_reverse_column_mappings, inplace=True)
    
    if 'date' in df.columns:
        def format_date_for_csv(d_str):
            if pd.isna(d_str) or not isinstance(d_str, str) or d_str.strip() == '':
                return ''
            try:
                dt_obj = datetime.datetime.strptime(d_str, '%Y-%m-%d')
                return dt_obj.strftime('%d-%m-%Y')
            except ValueError:
                return d_str
        df['date'] = df['date'].apply(format_date_for_csv)
    
    # JSON fields handling
    json_fields_to_str = [
        'absent_roll_numbers', 'ufm_roll_numbers', 'invigilators',
        'senior_center_superintendent', 'center_superintendent',
        'assistant_center_superintendent', 'permanent_invigilator', 
        'assistant_permanent_invigilator', 'class_3_worker', 'class_4_worker',
        'prep_days', 'closing_days', 'selected_classes', 'setting_value'
    ]
    
    for field in json_fields_to_str:
        if field in df.columns:
            df[field] = df[field].apply(lambda x: str(x) if x is not None and x != [] else '')
    
    df = df.fillna('')
    df.to_csv(filename, index=False)
    return True, f"✅ Downloaded {len(df)} rows from `{table_name}` to `{filename}`."


# --- NEW FUNCTION: Download attestation_data_combined to parent folder ---
def download_attestation_data_to_parent_folder():
    current_script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.abspath(os.path.join(current_script_dir, os.pardir))
    
    # Construct the full path for the CSV file in the parent directory
    output_filename = os.path.join(parent_dir, ATTESTATION_DATA_FILE)
    
    st.info(f"Attempting to download '{ATTESTATION_DATA_FILE}' from Supabase to: {output_filename}")
    success, message = download_supabase_to_csv("attestation_data_combined", output_filename)
    
    if success:
        st.success(f"Successfully downloaded '{ATTESTATION_DATA_FILE}' to the parent folder.")
        # Reload the app to ensure the newly downloaded file is recognized
        st.rerun() 
    else:
        st.error(f"Failed to download '{ATTESTATION_DATA_FILE}': {message}")
    
    return success, message

# --- Helper Functions (Place these BEFORE load_data) ---

def _format_roll_number(roll):
    """
    Converts any roll number input to a clean, stripped string,
    removing any trailing '.0' from Excel/float conversions.
    """
    if pd.isna(roll):
        return ""  # Return empty string for NaNs
    
    # Convert to string, strip whitespace, and remove trailing '.0'
    roll_str = str(roll).strip()
    if roll_str.endswith('.0'):
        roll_str = roll_str[:-2]
        
    return roll_str

def _format_paper_code(code):
    """
    Converts any paper code input to a clean, stripped string,
    removing any trailing '.0' from Excel/float conversions.
    """
    if pd.isna(code):
        return ""  # Return empty string for NaNs
    
    # Convert to string, strip whitespace, and remove trailing '.0'
    code_str = str(code).strip()
    if code_str.endswith('.0'):
        code_str = code_str[:-2]
        
    return code_str


# --- Your UPDATED load_data Function ---

def load_data():
    """
    Loads all required CSV data from local files, downloading from Supabase if missing.
    UPDATED: Iterates through ALL system tables to ensure local CSVs are always in sync with Supabase.
    """
    sitting_plan_df = pd.DataFrame()
    timetable_df = pd.DataFrame()
    assigned_seats_df = pd.DataFrame(columns=["Roll Number", "Paper Code", "Paper Name", "Room Number", "Seat Number", "date", "shift"])
    attestation_df = pd.DataFrame()

    # --- 1. Sync ALL Tables from Supabase ---
    tables_to_sync = {
        "timetable": TIMETABLE_FILE,
        "sitting_plan": SITTING_PLAN_FILE,
        "assigned_seats": ASSIGNED_SEATS_FILE,
        "prep_closing_assignments": PREP_CLOSING_ASSIGNMENTS_FILE,
        "global_settings": GLOBAL_SETTINGS_FILE,
        "shift_assignments": SHIFT_ASSIGNMENTS_FILE,
        "room_invigilator_assignments": ROOM_INVIGILATORS_FILE,
        "exam_team_members": EXAM_TEAM_MEMBERS_FILE,
        "cs_reports": CS_REPORTS_FILE,
        "attestation_data_combined": ATTESTATION_DATA_FILE
    }

    for table_name, file_path in tables_to_sync.items():
        if not os.path.exists(file_path) or os.stat(file_path).st_size == 0:
            try:
                # Silently try to download everything on startup
                download_supabase_to_csv(table_name, file_path)
            except Exception:
                pass # Ignore errors during silent sync

    # --- 2. Load DataFrames for the App Session ---
    
    # Load Sitting Plan
    if os.path.exists(SITTING_PLAN_FILE) and os.stat(SITTING_PLAN_FILE).st_size > 0:
        try:
            roll_cols = {f"Roll Number {i}": str for i in range(1, 11)}
            sitting_plan_df = pd.read_csv(SITTING_PLAN_FILE, dtype=roll_cols)
            sitting_plan_df.columns = sitting_plan_df.columns.str.strip().str.replace('\ufeff', '').str.replace('\xa0', ' ')
            
            # Use helper functions
            if 'Paper Code' in sitting_plan_df.columns:
                sitting_plan_df['Paper Code'] = sitting_plan_df['Paper Code'].apply(_format_paper_code)
            
            for i in range(1, 11):
                col_name = f'Roll Number {i}'
                if col_name in sitting_plan_df.columns:
                    sitting_plan_df[col_name] = sitting_plan_df[col_name].apply(_format_roll_number)

        except Exception as e:
            st.error(f"Error loading {SITTING_PLAN_FILE}: {e}")
            sitting_plan_df = pd.DataFrame()

    # Load Timetable
    if os.path.exists(TIMETABLE_FILE) and os.stat(TIMETABLE_FILE).st_size > 0:
        try:
            timetable_df = pd.read_csv(TIMETABLE_FILE, dtype=str)
            timetable_df.columns = timetable_df.columns.str.strip().str.replace('\ufeff', '').str.replace('\xa0', ' ')
            
            if 'Paper Code' in timetable_df.columns:
                timetable_df['Paper Code'] = timetable_df['Paper Code'].apply(_format_paper_code)
            if 'date' in timetable_df.columns:
                timetable_df['date'] = timetable_df['date'].str.strip()
            if 'shift' in timetable_df.columns:
                timetable_df['shift'] = timetable_df['shift'].str.strip()
                
        except Exception as e:
            st.error(f"Error loading {TIMETABLE_FILE}: {e}")
            timetable_df = pd.DataFrame()
    
    # Load Assigned Seats
    if os.path.exists(ASSIGNED_SEATS_FILE) and os.stat(ASSIGNED_SEATS_FILE).st_size > 0:
        try:
            temp_assigned_df = pd.read_csv(ASSIGNED_SEATS_FILE, dtype=str)
            temp_assigned_df.columns = temp_assigned_df.columns.str.strip().str.replace('\ufeff', '').str.replace('\xa0', ' ')

            rename_map = {}
            if 'Roll Numb' in temp_assigned_df.columns: rename_map['Roll Numb'] = 'Roll Number'
            if 'Paper Cod' in temp_assigned_df.columns: rename_map['Paper Cod'] = 'Paper Code'
            if 'Paper Nan' in temp_assigned_df.columns: rename_map['Paper Nan'] = 'Paper Name'
            if 'Room Nur' in temp_assigned_df.columns: rename_map['Room Nur'] = 'Room Number'
            if 'Seat Numi' in temp_assigned_df.columns: rename_map['Seat Numi'] = 'Seat Number'
            if rename_map:
                temp_assigned_df.rename(columns=rename_map, inplace=True)
            
            required_assigned_cols = ["Roll Number", "Paper Code", "Paper Name", "Room Number", "Seat Number", "date", "shift"]
            missing_cols = [col for col in required_assigned_cols if col not in temp_assigned_df.columns]

            if missing_cols:
                st.error(f"Critical Error: Missing essential columns in {ASSIGNED_SEATS_FILE}: {missing_cols}.")
                assigned_seats_df = pd.DataFrame(columns=required_assigned_cols)
            else:
                assigned_seats_df = temp_assigned_df[required_assigned_cols].copy()
                assigned_seats_df['Paper Code'] = assigned_seats_df['Paper Code'].apply(_format_paper_code)
                assigned_seats_df['Roll Number'] = assigned_seats_df['Roll Number'].apply(_format_roll_number)
                assigned_seats_df['date'] = assigned_seats_df['date'].astype(str).str.strip()
                assigned_seats_df['shift'] = assigned_seats_df['shift'].astype(str).str.strip()
                assigned_seats_df['Room Number'] = assigned_seats_df['Room Number'].astype(str).str.strip()
                assigned_seats_df['Seat Number'] = assigned_seats_df['Seat Number'].astype(str).str.strip()

        except Exception as e:
            st.error(f"Error loading {ASSIGNED_SEATS_FILE}: {e}.")
            assigned_seats_df = pd.DataFrame(columns=required_assigned_cols)
    
    # Load Attestation Data
    current_script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.abspath(os.path.join(current_script_dir, os.pardir))
    attestation_file_in_parent = os.path.join(parent_dir, ATTESTATION_DATA_FILE)
    local_attestation = ATTESTATION_DATA_FILE
    path_to_load = attestation_file_in_parent if os.path.exists(attestation_file_in_parent) else local_attestation

    if os.path.exists(path_to_load) and os.stat(path_to_load).st_size > 0:
        try:
            attestation_df = pd.read_csv(path_to_load, dtype=str)
            attestation_df.columns = attestation_df.columns.str.strip().str.replace('\ufeff', '').str.replace('\xa0', ' ')
            if 'Roll Number' in attestation_df.columns:
                attestation_df['Roll Number'] = attestation_df['Roll Number'].apply(_format_roll_number)
            for i in range(1, 11):
                col_name = f'Paper {i}'
                if col_name in attestation_df.columns:
                    attestation_df[col_name] = attestation_df[col_name].fillna('').astype(str)
        except Exception as e:
            pass
  
    st.session_state['sitting_plan'] = sitting_plan_df
    st.session_state['timetable'] = timetable_df
    st.session_state['assigned_seats_df'] = assigned_seats_df
    st.session_state['attestation_df'] = attestation_df

    return sitting_plan_df, timetable_df, assigned_seats_df, attestation_df


def _format_paper_code(code):
    """
    Converts any paper code input to a clean, stripped string,
    removing any trailing '.0' from Excel/float conversions.
    """
    if pd.isna(code):
        return ""  # Return empty string for NaNs
    
    # Convert to string, strip whitespace, and remove trailing '.0'
    code_str = str(code).strip()
    if code_str.endswith('.0'):
        code_str = code_str[:-2]
        
    return code_str

def load_shift_assignments():
    if os.path.exists(SHIFT_ASSIGNMENTS_FILE):
        try:
            # Use a robust engine to handle inconsistent data
            df = pd.read_csv(SHIFT_ASSIGNMENTS_FILE, engine='python')
            
            def safe_literal_eval(val):
                if isinstance(val, str) and val.strip():
                    # Strip any surrounding quotes and whitespace
                    clean_val = val.strip().strip('"')
                    try:
                        # Safely convert the cleaned string to a list
                        return ast.literal_eval(clean_val)
                    except (ValueError, SyntaxError):
                        # If conversion fails, return an empty list
                        return []
                return []

            # Apply the safe parser to all relevant columns
            for role in ["senior_center_superintendent", "center_superintendent", "assistant_center_superintendent", 
                         "permanent_invigilator", "assistant_permanent_invigilator", 
                         "class_3_worker", "class_4_worker"]:
                if role in df.columns:
                    df[role] = df[role].apply(safe_literal_eval)

            return df

        except Exception as e:
            # st.error(f"Error loading shift assignments: {e}. Reinitializing shift assignments file.")
            return pd.DataFrame(columns=['date', 'shift', 'senior_center_superintendent', 'center_superintendent', 
                                         "assistant_center_superintendent", "permanent_invigilator", 
                                         "assistant_permanent_invigilator", "class_3_worker", "class_4_worker"])

    # If the file doesn't exist, create an empty DataFrame
    return pd.DataFrame(columns=['date', 'shift', 'senior_center_superintendent', 'center_superintendent', 
                                 "assistant_center_superintendent", "permanent_invigilator", 
                                 "assistant_permanent_invigilator", "class_3_worker", "class_4_worker"])
def save_shift_assignment(date, shift, assignments):
    assignments_df = load_shift_assignments()
    
    # Create a unique key for the assignment
    assignment_key = f"{date}_{shift}"

    # Prepare data for DataFrame
    data_for_df = {
        'date': date,
        'shift': shift,
        'senior_center_superintendent': str(assignments.get('senior_center_superintendent', [])),
        'center_superintendent': str(assignments.get('center_superintendent', [])), 
        'assistant_center_superintendent': str(assignments.get('assistant_center_superintendent', [])),
        'permanent_invigilator': str(assignments.get('permanent_invigilator', [])),
        'assistant_permanent_invigilator': str(assignments.get('assistant_permanent_invigilator', [])),
        'class_3_worker': str(assignments.get('class_3_worker', [])),
        'class_4_worker': str(assignments.get('class_4_worker', []))
    }
    new_row_df = pd.DataFrame([data_for_df])

    # Check if assignment_key already exists
    if assignment_key in (assignments_df['date'] + '_' + assignments_df['shift']).values:
        idx_to_update = assignments_df[(assignments_df['date'] == date) & (assignments_df['shift'] == shift)].index[0]
        for col, val in data_for_df.items():
            assignments_df.loc[idx_to_update, col] = val
    else:
        assignments_df = pd.concat([assignments_df, new_row_df], ignore_index=True)
    
    try:
        # 1. Save to local CSV
        assignments_df.to_csv(SHIFT_ASSIGNMENTS_FILE, index=False)

        # 2. Sync to Supabase
        if supabase:
            try:
                supabase.table("shift_assignments").delete().neq("id", 0).execute()
                upload_csv_to_supabase("shift_assignments", SHIFT_ASSIGNMENTS_FILE)
            except Exception as db_e:
                 return True, f"Saved locally, but Supabase sync failed: {db_e}"

        return True, "Shift assignments saved and synced to Supabase!"
    except Exception as e:
        return False, f"Error saving shift assignments: {e}"


# Helper function to ensure consistent string formatting for paper codes (remove .0 if numeric)
def _format_paper_code(code_str):
    if pd.isna(code_str) or not code_str:
        return ""
    s = str(code_str).strip()
    # If it looks like a float (e.g., "12345.0"), convert to int string
    if s.endswith('.0') and s[:-2].isdigit():
        return s[:-2]
    return s



# Save uploaded files (for admin panel)
def save_uploaded_file(uploaded_file_content, filename):
    try:
        if isinstance(uploaded_file_content, pd.DataFrame):
            # If it's a DataFrame, convert to CSV bytes
            csv_bytes = uploaded_file_content.to_csv(index=False).encode('utf-8')
        else:
            # Assume it's bytes from st.file_uploader
            # Ensure uploaded_file_content is a BytesIO object or similar with .getbuffer()
            if hasattr(uploaded_file_content, 'getbuffer'):
                csv_bytes = uploaded_file_content.getbuffer()
            else:
                # Fallback for other file-like objects, or if it's already bytes
                csv_bytes = uploaded_file_content.read()


        with open(filename, "wb") as f:
            f.write(csv_bytes)
        return True, f"File {filename} saved successfully!" # Modified: Return a tuple here
    except Exception as e:
        return False, f"Error saving file {filename}: {e}"


# Admin login (simple hardcoded credentials)
def admin_login():
    user = st.text_input("Username", type="default")
    pwd = st.text_input("Password", type="password")
    return user == "admin" and pwd == "admin123"

# Centre Superintendent login (simple hardcoded credentials)
def cs_login():
    user = st.text_input("CS Username", type="default")
    pwd = st.text_input("CS Password", type="password")
    return user == "cs_admin" and pwd == "cs_pass123"

# --- CSV Helper Functions for CS Reports ---
def load_cs_reports_csv():
    if os.path.exists(CS_REPORTS_FILE):
        try:
            df = pd.read_csv(CS_REPORTS_FILE)
            
            # Standardize column names to lowercase and replace spaces with underscores
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

            # Ensure 'class' column exists, add if missing with empty string as default
            if 'class' not in df.columns:
                df['class'] = ""
            
            # Convert string representations of lists back to actual lists
            for col in ['absent_roll_numbers', 'ufm_roll_numbers']:
                if col in df.columns:
                    # Convert to string, then handle 'nan' and empty strings before literal_eval
                    df[col] = df[col].astype(str).apply(
                        lambda x: ast.literal_eval(x) if x.strip() and x.strip().lower() != 'nan' else []
                    )
            return df
        except Exception as e:
            st.error(f"Error loading CS reports from CSV: {e}")
            return pd.DataFrame(columns=['report_key', 'date', 'shift', 'room_num', 'paper_code', 'paper_name', 'class', 'absent_roll_numbers', 'ufm_roll_numbers'])
    else:
        return pd.DataFrame(columns=['report_key', 'date', 'shift', 'room_num', 'paper_code',])

def save_cs_report_csv(report_key, data):
    reports_df = load_cs_reports_csv()
    
    # Convert lists to string representation for CSV storage
    data_for_df = data.copy()
    data_for_df['absent_roll_numbers'] = str(data_for_df.get('absent_roll_numbers', []))
    data_for_df['ufm_roll_numbers'] = str(data_for_df.get('ufm_roll_numbers', []))

    new_row_df = pd.DataFrame([data_for_df])

    if report_key in reports_df['report_key'].values:
        idx_to_update = reports_df[reports_df['report_key'] == report_key].index[0]
        for col, val in data_for_df.items():
            reports_df.loc[idx_to_update, col] = val
    else:
        reports_df = pd.concat([reports_df, new_row_df], ignore_index=True)

    try:
        # 1. Save to local CSV
        reports_df.to_csv(CS_REPORTS_FILE, index=False)

        # 2. Sync to Supabase
        if supabase:
            try:
                supabase.table("cs_reports").delete().neq("id", 0).execute()
                # We must define the json fields for correct parsing in upload_csv_to_supabase
                # Fortunately upload_csv_to_supabase already handles this internally
                upload_csv_to_supabase("cs_reports", CS_REPORTS_FILE)
            except Exception as db_e:
                return True, f"Saved locally, but Supabase sync failed: {db_e}"

        return True, "Report saved and synced to Supabase successfully!"
    except Exception as e:
        return False, f"Error saving report to CSV: {e}"

def load_single_cs_report_csv(report_key):
    reports_df = load_cs_reports_csv()
    filtered_df = reports_df[reports_df['report_key'] == report_key]
    if not filtered_df.empty:
        return True, filtered_df.iloc[0].to_dict()
    else:
        return False, {}

# --- Exam Team Members Functions ---
def load_exam_team_members():
    if os.path.exists(EXAM_TEAM_MEMBERS_FILE):
        try:
            df = pd.read_csv(EXAM_TEAM_MEMBERS_FILE)
            return df['Name'].tolist()
        except Exception as e:
            st.error(f"Error loading exam team members: {e}")
            return []
    return []

def save_exam_team_members(members):
    df = pd.DataFrame({'Name': sorted(list(set(members)))}) # Remove duplicates and sort
    try:
        # 1. Save to local CSV
        df.to_csv(EXAM_TEAM_MEMBERS_FILE, index=False)
        
        # 2. Sync to Supabase (Delete old data, Upload new data)
        if supabase:
            try:
                # Delete all existing rows to prevent duplicates (assumes table has an 'id' column)
                supabase.table("exam_team_members").delete().neq("id", 0).execute()
                # Upload the updated CSV
                upload_csv_to_supabase("exam_team_members", EXAM_TEAM_MEMBERS_FILE)
            except Exception as db_e:
                return True, f"Saved locally, but Supabase sync failed: {db_e}"

        return True, "Exam team members saved and synced to Supabase successfully!"
    except Exception as e:
        return False, f"Error saving exam team members: {e}"

# Refactored helper function to get raw student data for a session
def _get_session_students_raw_data(date_str, shift, assigned_seats_df, timetable_df):
    """
    Collects raw student data for a given date and shift from assigned_seats_df
    and merges with timetable info.
    Returns a list of dictionaries, each representing an assigned student.
    """
    all_students_data = []

    # Filter timetable for the given date and shift
    current_day_exams_tt = timetable_df[
        (timetable_df["date"].astype(str).str.strip() == date_str) &
        (timetable_df["shift"].astype(str).str.strip().str.lower() == shift.lower())
    ].copy()

    if current_day_exams_tt.empty:
        return all_students_data # Return empty list if no exams found

    # Iterate through each exam scheduled for the date/shift in the timetable
    for _, tt_row in current_day_exams_tt.iterrows():
        tt_class = str(tt_row["Class"]).strip()
        tt_paper_code = str(tt_row["Paper Code"]).strip()
        tt_paper_name = str(tt_row["Paper Name"]).strip()

        # Filter assigned_seats_df for students assigned to this specific exam session
        current_exam_assigned_students = assigned_seats_df[
            (assigned_seats_df["date"].astype(str).str.strip() == date_str) &
            (assigned_seats_df["shift"].astype(str).str.strip().str.lower() == shift.lower()) &
            (assigned_seats_df["Paper Code"].astype(str).str.strip() == tt_paper_code) & # Use formatted paper code
            (assigned_seats_df["Paper Name"].astype(str).str.strip() == tt_paper_name)
        ]

        for _, assigned_row in current_exam_assigned_students.iterrows():
            roll_num = str(assigned_row["Roll Number"]).strip()
            room_num = str(assigned_row["Room Number"]).strip()
            seat_num_raw = str(assigned_row["Seat Number"]).strip()

            seat_num_display = ""
            seat_num_sort_key = None
            try:
                # Handle alphanumeric seats for sorting (e.g., 1A, 2A, 1B, 2B)
                if re.match(r'^\d+[A-Z]$', seat_num_raw):
                    num_part = int(re.match(r'^(\d+)', seat_num_raw).group(1))
                    char_part = re.search(r'([A-Z])$', seat_num_raw).group(1)
                    # Assign a tuple for sorting: (char_order, number)
                    seat_num_sort_key = (ord(char_part), num_part)
                    seat_num_display = seat_num_raw
                elif seat_num_raw.isdigit():
                    seat_num_sort_key = (float('inf'), int(seat_num_raw)) # Numeric seats after alphanumeric
                    seat_num_display = str(int(float(seat_num_raw))) # Display as integer string
                else:
                    seat_num_sort_key = (float('inf'), float('inf')) # Fallback for other formats
                    seat_num_display = seat_num_raw if seat_num_raw else "N/A"
            except ValueError:
                seat_num_sort_key = (float('inf'), float('inf')) # Fallback for other formats
                seat_num_display = seat_num_raw if seat_num_raw else "N/A"

            all_students_data.append({
                "roll_num": roll_num,
                "room_num": room_num,
                "seat_num_display": seat_num_display, # This is what will be displayed/exported
                "seat_num_sort_key": seat_num_sort_key, # This is for sorting
                "paper_name": tt_paper_name,
                "paper_code": tt_paper_code,
                "class_name": tt_class,
                "date": date_str,
                "shift": shift
            })
    return all_students_data

def get_all_students_for_date_shift_formatted(date_str, shift, assigned_seats_df, timetable):
    all_students_data = _get_session_students_raw_data(date_str, shift, assigned_seats_df, timetable)

    if not all_students_data:
        return None, "No students found for the selected date and shift.", None

    # Sort the collected data by Room Number, then Seat Number
    all_students_data.sort(key=lambda x: (x['room_num'], x['seat_num_sort_key']))

    # Extract exam_time and class_summary_header from timetable (similar to original logic)
    current_day_exams_tt = timetable[
        (timetable["date"].astype(str).str.strip() == date_str) &
        (timetable["shift"].astype(str).str.strip().str.lower() == shift.lower())
    ]
    exam_time = current_day_exams_tt.iloc[0]["Time"].strip() if "Time" in current_day_exams_tt.columns else "TBD"
    unique_classes = current_day_exams_tt['Class'].dropna().astype(str).str.strip().unique()
    class_summary_header = ""
    if len(unique_classes) == 1:
        class_summary_header = f"{unique_classes[0]} Examination {datetime.datetime.now().year}"
    elif len(unique_classes) > 1:
        class_summary_header = f"Various Classes Examination {datetime.datetime.now().year}"
    else:
        class_summary_header = f"Examination {datetime.datetime.now().year}"

    # --- Prepare text output ---
    output_string_parts = []
    output_string_parts.append("जीवाजी विश्वविद्यालय ग्वालियर")
    output_string_parts.append("परीक्षा केंद्र :- शासकीय विधि महाविद्यालय, मुरेना (म. प्र.) कोड :- G107")
    output_string_parts.append(class_summary_header)
    output_string_parts.append(f"दिनांक :-{date_str}")
    output_string_parts.append(f"पाली :-{shift}")
    output_string_parts.append(f"समय :-{exam_time}")

    students_by_room = {}
    for student in all_students_data:
        room = student['room_num']
        if room not in students_by_room:
            students_by_room[room] = []
        students_by_room[room].append(student)

    for room_num in sorted(students_by_room.keys()):
        output_string_parts.append(f" कक्ष :-{room_num}") # Added space for consistency
        current_room_students = students_by_room[room_num]

        num_cols = 10

        for i in range(0, len(current_room_students), num_cols):
            block_students = current_room_students[i : i + num_cols]

            # Create a single line for 10 students
            single_line_students = []
            for student in block_students:
                # Modified formatting here: removed space after '(' and added '-' before paper_name
                single_line_students.append(
                    f"{student['roll_num']}( कक्ष-{student['room_num']}-सीट-{student['seat_num_display']})-{student['paper_name']}"
                )

            output_string_parts.append("".join(single_line_students)) # Join directly without spaces

    final_text_output = "\n".join(output_string_parts)

    # --- Prepare Excel output data ---
    excel_output_data = []

    # Excel Header
    excel_output_data.append(["जीवाजी विश्वविद्यालय ग्वालियर"])
    excel_output_data.append(["परीक्षा केंद्र :- शासकीय विधि महाविद्यालय, मुरेना (म. प्र.) कोड :- G107"])
    excel_output_data.append([class_summary_header])
    excel_output_data.append([]) # Blank line
    excel_output_data.append(["दिनांक :-", date_str])
    excel_output_data.append(["पाली :-", shift])
    excel_output_data.append(["समय :-", exam_time])
    excel_output_data.append([]) # Blank line

    # Excel Student Data Section (now each block of 10 students is one row, each student is one cell)
    for room_num in sorted(students_by_room.keys()):
        excel_output_data.append([f" कक्ष :-{room_num}"]) # Added space for consistency
        current_room_students = students_by_room[room_num]

        num_cols = 10

        for i in range(0, len(current_room_students), num_cols):
            block_students = current_room_students[i : i + num_cols]

            excel_row_for_students = [""] * num_cols # Prepare 10 cells for this row

            for k, student in enumerate(block_students):
                # Each cell contains the full student string, modified formatting
                excel_row_for_students[k] = (
                    f"{student['roll_num']}( कक्ष-{student['room_num']}-सीट-{student['seat_num_display']})-{student['paper_name']}"
                )

            excel_output_data.append(excel_row_for_students)
            excel_output_data.append([""] * num_cols) # Blank row for spacing

    return final_text_output, None, excel_output_data	

# --- Room Invigilator Assignment Functions (NEW) ---
def load_room_invigilator_assignments():
    if os.path.exists(ROOM_INVIGILATORS_FILE):
        try:
            df = pd.read_csv(ROOM_INVIGILATORS_FILE)
            if 'invigilators' in df.columns:
                df['invigilators'] = df['invigilators'].astype(str).apply(
                    lambda x: ast.literal_eval(x) if x.strip() and x.strip().lower() != 'nan' else []
                )
            return df
        except Exception as e:
            st.error(f"Error loading room invigilator assignments: {e}")
            return pd.DataFrame(columns=['date', 'shift', 'room_num', 'invigilators'])
    return pd.DataFrame(columns=['date', 'shift', 'room_num', 'invigilators'])

def save_room_invigilator_assignment(date, shift, room_num, invigilators):
    inv_df = load_room_invigilator_assignments()
    
    # Create a unique key for the assignment
    assignment_key = f"{date}_{shift}_{room_num}"

    data_for_df = {
        'date': date,
        'shift': shift,
        'room_num': room_num,
        'invigilators': str(invigilators)
    }
    new_row_df = pd.DataFrame([data_for_df])

    # Check if assignment_key already exists
    if assignment_key in (inv_df['date'] + '_' + inv_df['shift'] + '_' + inv_df['room_num'].astype(str)).values:
        idx_to_update = inv_df[
            (inv_df['date'] == date) & 
            (inv_df['shift'] == shift) & 
            (inv_df['room_num'].astype(str) == str(room_num))
        ].index[0]
        for col, val in data_for_df.items():
            inv_df.loc[idx_to_update, col] = val
    else:
        inv_df = pd.concat([inv_df, new_row_df], ignore_index=True)
    
    try:
        # 1. Save to local CSV
        inv_df.to_csv(ROOM_INVIGILATORS_FILE, index=False)

        # 2. Sync to Supabase
        if supabase:
            try:
                supabase.table("room_invigilator_assignments").delete().neq("id", 0).execute()
                upload_csv_to_supabase("room_invigilator_assignments", ROOM_INVIGILATORS_FILE)
            except Exception as db_e:
                return True, f"Saved locally, but Supabase sync failed: {db_e}"

        return True, "Room invigilator assignments saved and synced to Supabase!"
    except Exception as e:
        return False, f"Error saving room invigilator assignments: {e}"


# Get all exams for a roll number (Student View)
def get_all_exams(roll_number, sitting_plan, timetable):
    student_exams = []
    roll_number_str = str(roll_number).strip() # Ensure consistent string comparison

    # Iterate through each row of the sitting plan
    for _, sp_row in sitting_plan.iterrows():
        # Check all possible roll number columns in the current sitting plan row
        for i in range(1, 11):
            r_col = f"Roll Number {i}"
            if r_col in sp_row and str(sp_row[r_col]).strip() == roll_number_str:
                # If roll number matches, extract paper and class details from this sitting plan row
                paper = str(sp_row["Paper"]).strip()
                paper_code = str(sp_row["Paper Code"]).strip()
                paper_name = str(sp_row["Paper Name"]).strip()
                _class = str(sp_row["Class"]).strip()

                # Find all matching entries in the timetable for this paper and class
                matches_in_timetable = timetable[
                    (timetable["Paper"].astype(str).str.strip() == paper) &
                    (timetable["Paper Code"].astype(str).str.strip() == paper_code) &
                    (timetable["Paper Name"].astype(str).str.strip() == paper_name) &
                    (timetable["Class"].astype(str).str.strip().str.lower() == _class.lower())
                ]

                # Add all found timetable matches for this student's paper to the list
                for _, tt_row in matches_in_timetable.iterrows():
                    student_exams.append({
                        "date": tt_row["date"],
                        "shift": tt_row["shift"],
                        "Class": _class,
                        "Paper": paper,
                        "Paper Code": paper_code,
                        "Paper Name": paper_name
                    })
                # Break from inner loop once the roll number is found in a row to avoid duplicate processing for the same row
                # if the roll number appears in multiple 'Roll Number X' columns within the *same* row (unlikely but safe)
                break
    return student_exams

# Get sitting details for a specific roll number and date (Student View)
#
# DELETE your old 'get_sitting_details' function and REPLACE it with this one
#
# This new function uses the CORRECT data sources.
#
def get_student_exam_details(roll_number, date, assigned_seats_df, timetable_df):
    """
    Finds exam details for a student by correctly searching the 
    assigned_seats_df and cross-referencing timetable_df.
    """
    
    # Use the cleaning functions from load_data.
    # Make sure _format_roll_number(roll) exists from my previous answer!
    # If not, use this:
    def _format_roll_number(roll):
        roll_str = str(roll).strip()
        return roll_str[:-2] if roll_str.endswith('.0') else roll_str

    roll_number_str = _format_roll_number(roll_number)
    date_str = str(date).strip() # Should be 'DD-MM-YYYY'

    found_sittings = []
    
    # 1. Find the student's assigned seats on that day
    # (Assumes load_data has cleaned these columns!)
    student_exams_on_date = assigned_seats_df[
        (assigned_seats_df['Roll Number'] == roll_number_str) &
        (assigned_seats_df['date'] == date_str)
    ]

    if student_exams_on_date.empty:
        return [] # No exams found for this roll number and date

    # 2. Merge with timetable to get 'Class' (and 'Mode'/'Type' if they existed)
    # We must merge on all keys to get the *exact* class
    # Note: This assumes timetable_df is also cleaned in load_data
    
    # Ensure merge keys are correct types (paranoia is good here)
    timetable_to_merge = timetable_df[['Paper Code', 'date', 'shift', 'Class']].copy()
    
    # We may not need all keys if Paper Code + date + shift is unique
    final_details = pd.merge(
        student_exams_on_date,
        timetable_to_merge,
        on=['Paper Code', 'date', 'shift'],
        how='left'
    )
    
    # Fill in any missing data from the merge
    final_details['Class'] = final_details['Class'].fillna('N/A')

    # 3. Format the results
    for _, row in final_details.iterrows():
        found_sittings.append({
            "Room Number": row["Room Number"],
            "Seat Number": row["Seat Number"],
            "Class": row["Class"],
            "Paper": row.get("Paper", ""), # 'Paper' (short name) is in timetable, not assigned_seats
            "Paper Code": row["Paper Code"],
            "Paper Name": row["Paper Name"],
            "date": row["date"],
            "shift": row["shift"],
            "Mode": row.get("Mode", "REGULAR"), # Default if not found
            "Type": row.get("Type", "REGULAR")  # Default if not found
        })
        
    return found_sittings


# New function to get all students for a given date and shift, sorted by roll number (Admin Panel)
def get_all_students_roll_number_wise_formatted(date_str, shift, assigned_seats_df, timetable):
    all_students_data = _get_session_students_raw_data(date_str, shift, assigned_seats_df, timetable)
    
    if not all_students_data:
        return None, "No students found for the selected date and shift.", None

    # Sort the collected data by Roll Number (lexicographically as strings)
    all_students_data.sort(key=lambda x: x['roll_num'])

    # Extract exam_time and class_summary_header from timetable (similar to original logic)
    current_day_exams_tt = timetable[
        (timetable["date"].astype(str).str.strip() == date_str) &
        (timetable["shift"].astype(str).str.strip().str.lower() == shift.lower())
    ]
    exam_time = current_day_exams_tt.iloc[0]["Time"].strip() if "Time" in current_day_exams_tt.columns else "TBD"
    unique_classes = current_day_exams_tt['Class'].dropna().astype(str).str.strip().unique()
    class_summary_header = ""
    if len(unique_classes) == 1:
        class_summary_header = f"{unique_classes[0]} Examination {datetime.datetime.now().year}"
    elif len(unique_classes) > 1:
        class_summary_header = f"Various Classes Examination {datetime.datetime.now().year}"
    else:
        class_summary_header = f"Examination {datetime.datetime.now().year}"

    # --- Prepare text output ---
    output_string_parts = []
    output_string_parts.append("जीवाजी विश्वविद्यालय ग्वालियर")
    output_string_parts.append("परीक्षा केंद्र :- शासकीय विधि महाविद्यालय, मुरेना (म. प्र.) कोड :- G107")
    output_string_parts.append(class_summary_header)
    output_string_parts.append(f"दिनांक :-{date_str}")
    output_string_parts.append(f"पाली :-{shift}")
    output_string_parts.append(f"समय :-{exam_time}")
    output_string_parts.append("") # Blank line for separation

    num_cols = 10 
    for i in range(0, len(all_students_data), num_cols):
        block_students = all_students_data[i : i + num_cols]
        
        single_line_students = []
        for student in block_students:
            single_line_students.append(
                f"{student['roll_num']}( कक्ष-{student['room_num']}-सीट-{student['seat_num_display']}){student['paper_name']}"
            )
        output_string_parts.append("".join(single_line_students))

    final_text_output = "\n".join(output_string_parts)

    # --- Prepare Excel output data ---
    excel_output_data = []

    # Excel Header
    excel_output_data.append(["जीवाजी विश्वविद्यालय ग्वालियर"])
    excel_output_data.append(["परीक्षा केंद्र :- शासकीय विधि महाविद्यालय, मुरेना (म. प्र.) कोड :- G107"])
    excel_output_data.append([class_summary_header])
    excel_output_data.append([]) # Blank line
    excel_output_data.append(["दिनांक :-", date_str])
    excel_output_data.append(["पाली :-", shift])
    excel_output_data.append(["समय :-", exam_time])
    excel_output_data.append([]) # Blank line

    # Excel Student Data Section
    for i in range(0, len(all_students_data), num_cols):
        block_students = all_students_data[i : i + num_cols]
        
        excel_row_for_students = [""] * num_cols

        for k, student in enumerate(block_students):
            excel_row_for_students[k] = (
                f"{student['roll_num']}( कक्ष-{student['room_num']}-सीट-{student['seat_num_display']}){student['paper_name']}"
            )
        
        excel_output_data.append(excel_row_for_students)
        excel_output_data.append([""] * num_cols) # Blank row for spacing

    return final_text_output, None, excel_output_data

# New helper function based on pdftocsv.py's extract_metadata, but using "UNSPECIFIED" defaults
def extract_metadata_from_pdf_text(text):
    # Extract Class Group, Year/Semester, and Session like "BSC", "1YEAR", "MAR-2025"
    # Looking for pattern like "BSC / 1YEAR / REGULAR / EXR / MAR-2025" or "LLB / 6SEM / REGULAR / EXR / JUN-2025"
    pattern_match = re.search(r'([A-Z]+)\s*/\s*(\d+(?:SEM|YEAR))\s*/\s*([A-Z]+)\s*/\s*([A-Z]+)\s*/\s*([A-Z]{3}-20\d{2})', text)
   
    if pattern_match:
        class_part = pattern_match.group(1)  # BSC/LLB
        year_part = pattern_match.group(2)   # 1YEAR/6SEM
        session_part = pattern_match.group(5)  # MAR-2025/JUN-2025/DEC-2025
        
        # Format session from "MAR-2025" to "MAR 25", "JUN-2025" to "JUN 25", etc.
        session_formatted = session_part.replace("-20", " ")  # "MAR-2025" -> "MAR 25"
        
        class_val = f"{class_part} {year_part} - {session_formatted}"
        mode_type = pattern_match.group(3)  # Third element (PRIVATE/REGULAR)
        type_type = pattern_match.group(4)  # Fourth element (SUPP/EXR/REGULAR/etc)
    else:
        # Fallback: try to extract class and year separately
        class_match = re.search(r'([A-Z]+)\s*/?\s*(\d+(?:SEM|YEAR))', text)
        session_match = re.search(r'([A-Z]{3}-20\d{2})', text)  # Any 3-letter month + year
        
        if class_match and session_match:
            class_part = class_match.group(1)
            year_part = class_match.group(2)
            session_formatted = session_match.group(1).replace("-20", " ")  # Format any session
            class_val = f"{class_part} {year_part} - {session_formatted}"
        elif class_match:
            class_val = f"{class_match.group(1)} {class_match.group(2)}"
        else:
            class_val = "UNSPECIFIED_CLASS"
        
        # Fallback to original logic for mode and type
        mode_type = "UNSPECIFIED_MODE"
        # Check for PRIVATE first since it's more specific
        for keyword_mode in ["PRIVATE", "REGULAR"]:
            if keyword_mode in text.upper():
                mode_type = keyword_mode
                break
               
        type_type = "UNSPECIFIED_TYPE"
        # Check for more specific types first
        for keyword_type in ["ATKT", "SUPP", "EXR", "REGULAR", "PRIVATE"]:
            if keyword_type in text.upper():
                type_type = keyword_type
                break

    paper_code = re.search(r'Paper Code[:\s]*([A-Z0-9]+)', text, re.IGNORECASE)
    paper_code_val = _format_paper_code(paper_code.group(1)) if paper_code else "UNSPECIFIED_PAPER_CODE" # Use formatter
   
    paper_name = re.search(r'Paper Name[:\s]*(.+?)(?:\n|$)', text)
    paper_name_val = paper_name.group(1).strip() if paper_name else "UNSPECIFIED_PAPER_NAME"
   
    return {
        "class": class_val,
        "mode": mode_type,
        "type": type_type,  
        "room_number": "",
        "seat_numbers": [""] * 10,
        "paper_code": paper_code_val,
        "paper_name": paper_name_val
    }

# --- Integration of pdftocsv.py logic ---
def process_sitting_plan_pdfs(zip_file_buffer, output_sitting_plan_path, output_timetable_path):
    all_rows = []
    sitting_plan_columns = [f"Roll Number {i+1}" for i in range(10)]
    sitting_plan_columns += ["Class", "Mode", "Type", "Room Number"]
    sitting_plan_columns += [f"Seat Number {i+1}" for i in range(10)]
    sitting_plan_columns += ["Paper", "Paper Code", "Paper Name"]

    def extract_roll_numbers(text):
        # Use a set to automatically handle duplicates during extraction
        return sorted(list(set(re.findall(r'\b\d{9}\b', text)))) # De-duplicate and sort

    def format_sitting_plan_rows(rolls, paper_folder_name, meta):
        rows = []
        for i in range(0, len(rolls), 10):
            row = rolls[i:i+10]
            while len(row) < 10:
                row.append("")  # pad to ensure 10 roll number columns
            row.extend([
                meta["class"],
                meta["mode"],
                meta["type"],
                meta["room_number"]
            ])
            row.extend(meta["seat_numbers"]) # These are initially blank, filled later by assignment
            row.append(paper_folder_name)  # Use folder name as Paper
            row.append(meta["paper_code"])
            row.append(meta["paper_name"])
            rows.append(row)
        return rows

    unique_exams_for_timetable = [] # To collect data for incomplete timetable

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(zip_file_buffer, 'r') as zip_ref:
            zip_ref.extractall(tmpdir)
        
        base_dir = tmpdir
        # Check if there's a 'pdf_folder' sub-directory inside the extracted content
        # This handles cases where the zip contains 'pdf_folder' directly or files/folders at root
        extracted_contents = os.listdir(tmpdir)
        if 'pdf_folder' in extracted_contents and os.path.isdir(os.path.join(tmpdir, 'pdf_folder')):
            base_dir = os.path.join(tmpdir, 'pdf_folder')
        elif len(extracted_contents) == 1 and os.path.isdir(os.path.join(tmpdir, extracted_contents[0])):
            # If there's only one folder at the root, assume it's the base_dir
            base_dir = os.path.join(tmpdir, extracted_contents[0])


        processed_files_count = 0
        for folder_name in os.listdir(base_dir):
            folder_path = os.path.join(base_dir, folder_name)
            if os.path.isdir(folder_path):
                for file in os.listdir(folder_path):
                    if file.lower().endswith(".pdf"):
                        pdf_path = os.path.join(folder_path, file)
                        try:
                            doc = fitz.open(pdf_path)
                            full_text = "\n".join(page.get_text() for page in doc)
                            doc.close()
                            
                            # Use the new extract_metadata_from_pdf_text function
                            current_meta = extract_metadata_from_pdf_text(full_text)
                            
                            # Ensure paper_code and paper_name fallback to folder_name if still unspecified
                            if current_meta['paper_code'] == "UNSPECIFIED_PAPER_CODE":
                                current_meta['paper_code'] = folder_name
                            if current_meta['paper_name'] == "UNSPECIFIED_PAPER_NAME":
                                current_meta['paper_name'] = folder_name

                            rolls = extract_roll_numbers(full_text) # This now de-duplicates and sorts
                            rows = format_sitting_plan_rows(rolls, paper_folder_name=folder_name, meta=current_meta)
                            all_rows.extend(rows)
                            processed_files_count += 1
                            st.info(f"✔ Processed: {file} ({len(rolls)} unique roll numbers)")

                            # Collect unique exam details for timetable generation
                            unique_exams_for_timetable.append({
                                'Class': current_meta['class'],
                                'Paper': folder_name, # Use folder name as Paper
                                'Paper Code': current_meta['paper_code'],
                                'Paper Name': current_meta['paper_name']
                            })

                        except Exception as e:
                            st.error(f"❌ Failed to process {file}: {e}")
    
    # --- Sitting Plan Update Logic ---
    if all_rows:
        df_new_sitting_plan = pd.DataFrame(all_rows, columns=sitting_plan_columns)

        # Load existing sitting plan data
        existing_sitting_plan_df = pd.DataFrame()
        if os.path.exists(output_sitting_plan_path):
            try:
                existing_sitting_plan_df = pd.read_csv(output_sitting_plan_path, dtype={
                    f"Roll Number {i}": str for i in range(1, 11)
                })
                existing_sitting_plan_df.columns = existing_sitting_plan_df.columns.str.strip()
                if 'Paper Code' in existing_sitting_plan_df.columns:
                    existing_sitting_plan_df['Paper Code'] = existing_sitting_plan_df['Paper Code'].apply(_format_paper_code)
            except Exception as e:
                st.warning(f"Could not load existing sitting plan data for update: {e}. Starting fresh for sitting plan.")
                existing_sitting_plan_df = pd.DataFrame(columns=sitting_plan_columns)

        # Ensure all columns are present in both DataFrames before concatenation
        # Add missing columns to df_new_sitting_plan from existing_sitting_plan_df
        for col in existing_sitting_plan_df.columns:
            if col not in df_new_sitting_plan.columns:
                df_new_sitting_plan[col] = pd.NA
        # Add missing columns to existing_sitting_plan_df from df_new_sitting_plan
        for col in df_new_sitting_plan.columns:
            if col not in existing_sitting_plan_df.columns:
                existing_sitting_plan_df[col] = pd.NA

        # Reorder columns to match existing_sitting_plan_df before concatenation
        df_new_sitting_plan = df_new_sitting_plan[existing_sitting_plan_df.columns]

        # Concatenate and remove duplicates
        combined_sitting_plan_df = pd.concat([existing_sitting_plan_df, df_new_sitting_plan], ignore_index=True)

        # Define columns for identifying unique sitting plan entries.
        roll_num_cols = [f"Roll Number {i+1}" for i in range(10)]
        
        # Using all relevant columns to define uniqueness for sitting plan entries
        subset_cols_sitting_plan = roll_num_cols + ["Class", "Mode", "Type", "Room Number", "Paper", "Paper Code", "Paper Name"]
        
        # Filter subset_cols_sitting_plan to only include columns actually present in the DataFrame
        existing_subset_cols_sitting_plan = [col for col in subset_cols_sitting_plan if col in combined_sitting_plan_df.columns]

        # Fill NaN values with empty strings before dropping duplicates for consistent hashing
        combined_sitting_plan_df_filled = combined_sitting_plan_df.fillna('')
        df_sitting_plan_final = combined_sitting_plan_df_filled.drop_duplicates(subset=existing_subset_cols_sitting_plan, keep='first')

        df_sitting_plan_final.to_csv(output_sitting_plan_path, index=False)
        st.success(f"Successfully processed {processed_files_count} PDFs and updated sitting plan to {output_sitting_plan_path}")
    else:
        st.warning("No roll numbers extracted from PDFs to update sitting plan.")

    # --- Timetable Update Logic ---
    if unique_exams_for_timetable:
        df_new_timetable_entries = pd.DataFrame(unique_exams_for_timetable).drop_duplicates().reset_index(drop=True)

        # Define expected structure
        expected_columns = ["SN", "date", "shift", "Time", "Class", "Paper", "Paper Code", "Paper Name"]

        # Load existing timetable if exists
        if os.path.exists(output_timetable_path):
            try:
                existing_timetable_df = pd.read_csv(output_timetable_path)
                existing_timetable_df.columns = existing_timetable_df.columns.str.strip()
                if 'Paper Code' in existing_timetable_df.columns:
                    existing_timetable_df['Paper Code'] = existing_timetable_df['Paper Code'].astype(str).str.strip()
            except Exception as e:
                st.warning(f"Could not load existing timetable: {e}. Starting fresh.")
                existing_timetable_df = pd.DataFrame(columns=expected_columns)
        else:
            existing_timetable_df = pd.DataFrame(columns=expected_columns)

        # Add missing columns to both DataFrames
        for col in expected_columns:
            if col not in df_new_timetable_entries.columns:
                df_new_timetable_entries[col] = pd.NA
            if col not in existing_timetable_df.columns:
                existing_timetable_df[col] = pd.NA

        # Reorder columns
        df_new_timetable_entries = df_new_timetable_entries[expected_columns]
        existing_timetable_df = existing_timetable_df[expected_columns]

        # Concatenate and deduplicate using relevant fields
        combined_df = pd.concat([existing_timetable_df, df_new_timetable_entries], ignore_index=True)

        # Fields that define uniqueness of a timetable entry (excluding SN)
        unique_fields = ["date", "shift", "Time", "Class", "Paper", "Paper Code", "Paper Name"]

        # Remove duplicates based on content
        df_timetable_final = combined_df.drop_duplicates(subset=unique_fields, keep='first').reset_index(drop=True)

        # Reassign serial numbers
        df_timetable_final["SN"] = range(1, len(df_timetable_final) + 1)

        # Save final CSV
        df_timetable_final.to_csv(output_timetable_path, index=False)
        st.success(f"Timetable updated at {output_timetable_path}.")
        return True, "Timetable deduplicated and saved successfully."
    
    else:
        st.warning("No unique exam details found to generate timetable.")
        return False, "No data to process."   
    return True, "PDF processing complete."   


# --- Integration of rasa_pdf.py logic ---
def process_attestation_pdfs(zip_file_buffer, output_csv_path):
    all_data = []

    def parse_pdf_content(text):
        students = re.split(r"\n?RollNo\.\:\s*", text)
        students = [s.strip() for s in students if s.strip()]

        student_records = []

        for s in students:
            lines = s.splitlines()
            lines = [line.strip() for line in lines if line.strip()]

            def extract_after(label):
                for i, line in enumerate(lines):
                    if line.startswith(label):
                        value = line.replace(label, "", 1).strip() # Use count=1 for replace
                        if value:
                            return value
                        elif i+1 < len(lines):
                            return lines[i+1].strip()
                    # Special handling for "Regular/Backlog" as it might be on the next line
                    if label == "Regular/ Backlog:" and line.startswith("Regular/Backlog"):
                        value = line.replace("Regular/Backlog", "", 1).strip() # Use count=1 for replace
                        if value:
                            return value
                        elif i+1 < len(lines):
                            return lines[i+1].strip()
                return "" # Return empty string if label not found or value is empty

            roll_no = re.match(r"(\d{9})", lines[0]).group(1) if lines and re.match(r"(\d{9})", lines[0]) else ""
            enrollment = extract_after("Enrollment No.:")
            session = extract_after("Session:")
            regular = extract_after("Regular/ Backlog:")
            student_name = extract_after("Name:")
            father = extract_after("Father's Name:")
            mother = extract_after("Mother's Name:")
            gender = extract_after("Gender:")
            exam_name = extract_after("Exam Name:")
            centre = extract_after("Exam Centre:")
            college = extract_after("College Nmae:") # Note: Original script had 'Nmae'
            address = extract_after("Address:")

            papers = re.findall(r"([^\n]+?\[\d{5}\][^\n]*)", s) # Corrected regex for paper code

            student_data = {
                "Roll Number": roll_no,
                "Enrollment Number": enrollment,
                "Session": session,
                "Regular/Backlog": regular,
                "Name": student_name,
                "Father's Name": father,
                "Mother's Name": mother,
                "Gender": gender,
                "Exam Name": exam_name,
                "Exam Centre": centre,
                "College Name": college,
                "Address": address
            }

            for i, paper in enumerate(papers[:10]):
                student_data[f"Paper {i+1}"] = paper.strip()

            student_records.append(student_data)
        return student_records

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(zip_file_buffer, 'r') as zip_ref:
            zip_ref.extractall(tmpdir)
        
        # Assuming PDFs are directly in the extracted folder or a subfolder named 'rasa_pdf'
        pdf_base_dir = tmpdir
        if 'rasa_pdf' in os.listdir(tmpdir) and os.path.isdir(os.path.join(tmpdir, 'rasa_pdf')):
            pdf_base_dir = os.path.join(tmpdir, 'rasa_pdf')

        processed_files_count = 0
        for filename in os.listdir(pdf_base_dir):
            if filename.lower().endswith(".pdf"):
                pdf_path = os.path.join(pdf_base_dir, filename)
                try:
                    doc = fitz.open(pdf_path)
                    text = "\n".join([page.get_text() for page in doc])
                    doc.close()
                    st.info(f"📄 Extracting: {filename}")
                    all_data.extend(parse_pdf_content(text))
                    processed_files_count += 1
                except Exception as e:
                    st.error(f"❌ Failed to process {filename}: {e}")
    
    if all_data:
        df = pd.DataFrame(all_data)
        df.to_csv(output_csv_path, index=False)
        return True, f"Successfully processed {processed_files_count} attestation PDFs and saved to {output_csv_path}"
    else:
        return False, "No data extracted from attestation PDFs."

# --- Integration of college_statistic.py logic ---
def generate_college_statistics(input_csv_path, output_csv_path):
    if not os.path.exists(input_csv_path):
        return False, f"Input file not found: {input_csv_path}. Please process attestation PDFs first."

    try:
        # Load data
        df = pd.read_csv(input_csv_path, dtype={"Roll Number": str, "Enrollment Number": str})

        # Basic cleaning
        df['College Name'] = df['College Name'].fillna('UNKNOWN').astype(str).str.strip().str.upper()
        df['Exam Name'] = df['Exam Name'].fillna('UNKNOWN').astype(str).str.strip().str.upper()
        df['Regular/Backlog'] = df['Regular/Backlog'].astype(str).str.strip().str.upper()

        # Extract class group and year
        def extract_class_group_and_year(exam_name):
            if pd.isna(exam_name):
                return "UNKNOWN", "UNKNOWN"

            exam_name = str(exam_name).upper().strip()

            # Match pattern like BCOM - Commerce [C032] - 1YEAR or BED - PLAIN[PLAIN] - 2SEM
            match = re.match(r'^([A-Z]+)\s*-\s*.+\[\w+\]\s*-\s*(\d+(ST|ND|RD|TH)?(YEAR|SEM))$', exam_name)
            if match:
                class_group = match.group(1).strip()
                year_or_sem = match.group(2).strip()
                return class_group, year_or_sem

            # Fallback: try to extract roman numeral patterns like II YEAR
            roman = re.search(r'\b([IVXLCDM]+)\s*(YEAR|SEM)\b', exam_name)
            if roman:
                return "UNKNOWN", roman.group(0).strip()

            return "UNKNOWN", "UNKNOWN"



        df[["Class Group", "Year"]] = df["Exam Name"].apply(lambda x: pd.Series(extract_class_group_and_year(x)))
        

        # Group definitions
        class_groups = sorted(df["Class Group"].dropna().unique())
        college_list = sorted(df["College Name"].dropna().unique())

        # Count function
        def get_counts(df, college, group, year):
            subset = df[(df["College Name"] == college) & (df["Class Group"] == group) & (df["Year"] == year)]
            total = len(subset)
            regular = len(subset[subset["Regular/Backlog"] == "REGULAR"])
            private = len(subset[subset["Regular/Backlog"] == "PRIVATE"])
            exr = len(subset[subset["Regular/Backlog"] == "EXR"])
            supp = len(subset[subset["Regular/Backlog"] == "SUPP"])
            atkt = len(subset[subset["Regular/Backlog"] == "ATKT"])
            return [total, regular, private, exr, atkt, supp]

        # Prepare output structure
        output_rows = []

        for group in class_groups:
            years = sorted(df[df["Class Group"] == group]["Year"].dropna().unique())

            # Header rows
            header_row1 = ["Class"] + [f"{group} - {year}" for year in years for _ in range(5)]
            header_row2 = ["College", "Grand Total"] + ["Total", "Regular", "Private", "EXR", "ATKT", "SUPP"] * len(years)

            block_data = []
            for college in college_list:
                row = [college]
                grand_total = 0
                for year in years:
                    t, r, p, x, a, s = get_counts(df, college, group, year)
                    row += [t, r, p, x, a, s]
                    grand_total += t
                row.insert(1, grand_total)
                block_data.append(row)

            output_rows.append(header_row1)
            output_rows.append(header_row2)
            output_rows += block_data
            output_rows.append([])

        # Final Summary Block
        output_rows.append(["College", "Total of all"])
        for college in college_list:
            total = len(df[df["College Name"] == college])
            output_rows.append([college, total])

        # Save final output
        pd.DataFrame(output_rows).to_csv(output_csv_path, index=False, header=False)
        return True, f"✅ College statistics saved to {output_csv_path}"

    except Exception as e:
        return False, f"❌ Error generating college statistics: {e}"


# New helper function to generate sequential seat numbers based on a range string (from assign_seats_app.py)
def generate_sequential_seats(seat_range_str, num_students):
    generated_seats = []
    seat_range_str = seat_range_str.strip().upper() # Normalize input

    if '-' in seat_range_str:
        start_seat_str, end_seat_str = seat_range_str.split('-')
        
        # Handle alphanumeric like "1A-60A"
        if re.match(r'^\d+[A-Z]$', start_seat_str) and re.match(r'^\d+[A-Z]$', end_seat_str):
            start_num = int(re.match(r'^(\d+)', start_seat_str).group(1))
            start_char = re.search(r'([A-Z])$', start_seat_str).group(1)
            end_num = int(re.match(r'^(\d+)', end_seat_str).group(1))
            end_char = re.search(r'([A-Z])$', end_seat_str).group(1)

            if start_char != end_char:
                raise ValueError("For alphanumeric seat ranges (e.g., 1A-60A), the alphabet part must be the same.")
            
            for i in range(start_num, end_num + 1):
                generated_seats.append(f"{i}{start_char}")
        # Handle numeric like "1-60"
        elif start_seat_str.isdigit() and end_seat_str.isdigit():
            start_num = int(start_seat_str)
            end_num = int(end_seat_str)
            for i in range(start_num, end_num + 1):
                generated_seats.append(str(i))
        else:
            raise ValueError("Invalid seat number range format. Use '1-60' or '1A-60A'.")
    elif seat_range_str.isdigit() or re.match(r'^\d+[A-Z]$', seat_range_str):
        generated_seats.append(seat_range_str)
    else:
        raise ValueError("Invalid seat number format. Use a single number, '1-60', or '1A-60A'.")

    # Return only as many seats as there are students, or all generated seats if fewer students
    return generated_seats[:num_students]


# NEW FUNCTION: Get unassigned students for a given date and shift
def get_unassigned_students_for_session(date_str, shift, sitting_plan_df, timetable_df):
    unassigned_roll_numbers_details = {} # {roll_num: {class, paper, paper_code, paper_name}}

    # 1. Filter timetable for the given date and shift
    relevant_tt_exams = timetable_df[
        (timetable_df["date"].astype(str).str.strip() == date_str) &
        (timetable_df["shift"].astype(str).str.strip().str.lower() == shift.lower())
    ].copy()

    if relevant_tt_exams.empty:
        return []

    # Create a unique identifier for exams in timetable for easier matching
    relevant_tt_exams['exam_key'] = relevant_tt_exams['Class'].astype(str).str.strip().str.lower() + "_" + \
                                     relevant_tt_exams['Paper'].astype(str).str.strip() + "_" + \
                                     relevant_tt_exams['Paper Code'].astype(str).str.strip() + "_" + \
                                     relevant_tt_exams['Paper Name'].astype(str).str.strip()

    # Iterate through sitting plan to find students for these exams
    for _, sp_row in sitting_plan_df.iterrows():
        # Create exam_key for this sitting plan row
        sp_exam_key = str(sp_row['Class']).strip().lower() + "_" + \
                      str(sp_row['Paper']).strip() + "_" + \
                      str(sp_row['Paper Code']).strip() + "_" + \
                      str(sp_row['Paper Name']).strip()

        # Check if this sitting plan entry corresponds to a relevant exam session
        if sp_exam_key in relevant_tt_exams['exam_key'].values:
            room_assigned = str(sp_row['Room Number']).strip()
            
            # Check all roll numbers in this sitting plan row
            for i in range(1, 11):
                roll_col = f"Roll Number {i}"
                if roll_col in sp_row and pd.notna(sp_row[roll_col]) and str(sp_row[roll_col]).strip() != '':
                    roll_num = str(sp_row[roll_col]).strip()
                    # If room is blank, this student is unassigned for this paper
                    if not room_assigned: # If room_assigned is an empty string
                        # Store details for display
                        unassigned_roll_numbers_details[roll_num] = {
                            'Class': str(sp_row['Class']).strip(),
                            'Paper': str(sp_row['Paper']).strip(),
                            'Paper Code': str(sp_row['Paper Code']).strip(),
                            'Paper Name': str(sp_row['Paper Name']).strip()
                        }
    
    # Convert to a list of dictionaries for display, sorted by roll number
    sorted_unassigned_list = []
    for roll, details in sorted(unassigned_roll_numbers_details.items()):
        sorted_unassigned_list.append({
            "Roll Number": roll,
            "Class": details['Class'],
            "Paper": details['Paper'],
            "Paper Code": details['Paper Code'],
            "Paper Name": details['Paper Name']
        })
    
    return sorted_unassigned_list

# NEW FUNCTION: Get summary of students by paper for a given session (assigned + unassigned)
def get_session_paper_summary(date_str, shift, sitting_plan_df, assigned_seats_df, timetable_df):
    summary_data = []

    # Filter timetable for the given date and shift
    relevant_tt_exams = timetable_df[
        (timetable_df["date"].astype(str).str.strip() == date_str) &
        (timetable_df["shift"].astype(str).str.strip().str.lower() == shift.lower())
    ].copy()

    if relevant_tt_exams.empty:
        return pd.DataFrame(columns=['Paper Name', 'Paper Code', 'Total Expected', 'Assigned', 'Unassigned'])

    # Iterate through each unique paper in the relevant timetable exams
    for _, tt_row in relevant_tt_exams.drop_duplicates(subset=['Paper Code', 'Paper Name']).iterrows():
        paper_code = str(tt_row['Paper Code']).strip()
        paper_name = str(tt_row['Paper Name']).strip()
        
        # Get all expected roll numbers for this specific paper (from sitting plan)
        expected_rolls = set()
        paper_sitting_rows = sitting_plan_df[sitting_plan_df['Paper Code'].astype(str).str.strip() == paper_code]
        for _, sp_row in paper_sitting_rows.iterrows():
            for i in range(1, 11):
                roll_col = f"Roll Number {i}"
                if roll_col in sp_row and pd.notna(sp_row[roll_col]) and str(sp_row[roll_col]).strip() != '':
                    expected_rolls.add(str(sp_row[roll_col]).strip())
        
        total_expected_students = len(expected_rolls)

        # Get assigned roll numbers for this specific paper, date, and shift
        assigned_rolls_for_paper = set(
            assigned_seats_df[
                (assigned_seats_df["Paper Code"].astype(str).str.strip() == paper_code) & # Use formatted paper code
                (assigned_seats_df["date"] == date_str) &
                (assigned_seats_df["shift"] == shift)
            ]["Roll Number"].astype(str).tolist()
        )
        num_assigned_students = len(assigned_rolls_for_paper)

        # Calculate unassigned students
        num_unassigned_students = total_expected_students - num_assigned_students

        summary_data.append({
            'Paper Name': paper_name,
            'Paper Code': paper_code,
            'Total Expected': total_expected_students,
            'Assigned': num_assigned_students,
            'Unassigned': num_unassigned_students
        })
    
    return pd.DataFrame(summary_data)

# NEW FUNCTION: Display Room Occupancy Report
def display_room_occupancy_report(sitting_plan_df, assigned_seats_df, timetable_df):
    st.subheader("📊 Room Occupancy Report")
    st.info("View detailed occupancy for each room based on Assigned Seats.")

    # We strictly need assigned_seats_df and timetable_df (for dates)
    if assigned_seats_df.empty or timetable_df.empty:
        st.warning("Please upload 'assigned_seats.csv' and 'timetable.csv' (and ensure seats are assigned) to generate this report.")
        return

    # --- 1. Standardize Columns ---
    assigned_seats_df['date'] = assigned_seats_df['date'].astype(str).str.strip()
    assigned_seats_df['shift'] = assigned_seats_df['shift'].astype(str).str.strip().str.lower()
    assigned_seats_df['Room Number'] = assigned_seats_df['Room Number'].astype(str).str.strip()
    assigned_seats_df['Seat Number'] = assigned_seats_df['Seat Number'].astype(str).str.strip()
    
    timetable_df['date'] = timetable_df['date'].astype(str).str.strip()
    timetable_df['shift'] = timetable_df['shift'].astype(str).str.strip().str.lower()

    # --- 2. Date and Shift Selection ---
    # Get options from Timetable (or Assigned Seats if Timetable is partial, but Timetable is safer for full list)
    report_date_options = sorted(timetable_df["date"].unique())
    report_shift_options = sorted(timetable_df["shift"].str.title().unique())

    col1, col2 = st.columns(2)
    with col1:
        selected_report_date = st.selectbox("Select Date", report_date_options, key="room_report_date")
    with col2:
        selected_report_shift = st.selectbox("Select Shift", report_shift_options, key="room_report_shift")

    if st.button("Generate Room Occupancy Report"):
        # --- 3. Filter Assigned Seats (The Source of Truth) ---
        relevant_assignments = assigned_seats_df[
            (assigned_seats_df['date'] == selected_report_date) &
            (assigned_seats_df['shift'] == selected_report_shift.lower())
        ].copy()

        if relevant_assignments.empty:
            st.warning(f"No students have been assigned seats for {selected_report_date} ({selected_report_shift}).")
            return

        # --- 4. Aggregate Data by Room ---
        room_stats = {} 

        for _, row in relevant_assignments.iterrows():
            room_num = row['Room Number']
            if room_num not in room_stats:
                room_stats[room_num] = {
                    'count': 0, 
                    'details': []
                }
            
            room_stats[room_num]['count'] += 1
            
            # Create detail string: "RollNumber (Seat)"
            # You can add Paper Code if needed: f"{row['Roll Number']} ({row['Seat Number']}) [{row['Paper Code']}]"
            detail_str = f"{row['Roll Number']} ({row['Seat Number']})"
            room_stats[room_num]['details'].append(detail_str)

        # --- 5. Build Final DataFrame ---
        room_occupancy_data = []
        for room_num, stats in room_stats.items():
            # Sort student details by seat number for readability
            # Logic attempts to sort numerically if seat is "1", "2" or "1A", "2A"
            try:
                stats['details'].sort(key=lambda x: int(''.join(filter(str.isdigit, x.split('(')[1]))))
            except:
                stats['details'].sort() # Fallback to string sort

            details_display = ", ".join(stats['details'])

            # Since we are using Assigned Seats as the source, "Total Expected" is equal to "Total Assigned"
            # We can rename the column to "Student Count" to be more accurate.
            room_occupancy_data.append({
                'Room Number': room_num,
                'Student Count': stats['count'],
                'Assigned Student Details': details_display
            })

        if room_occupancy_data:
            df_occupancy = pd.DataFrame(room_occupancy_data)
            
            # Sort by Room Number (Numeric if possible)
            try:
                df_occupancy['room_sort'] = pd.to_numeric(df_occupancy['Room Number'], errors='coerce')
                df_occupancy = df_occupancy.sort_values('room_sort').drop(columns=['room_sort'])
            except:
                df_occupancy = df_occupancy.sort_values('Room Number')

            # Reorder columns
            st.dataframe(df_occupancy[['Room Number', 'Student Count', 'Assigned Student Details']], use_container_width=True)
            
            # Download Button
            csv_occupancy = df_occupancy.to_csv(index=False).encode('utf-8')
            file_name = f"room_occupancy_{selected_report_date}_{selected_report_shift}.csv"
            st.download_button(
                label="Download Room Occupancy Report as CSV",
                data=csv_occupancy,
                file_name=file_name,
                mime="text/csv",
            )
        else:
            st.info("No occupancy data generated.")
            
# NEW FUNCTION: Generate Room Chart in specified format
def generate_room_chart_report(date_str, shift, sitting_plan_df, assigned_seats_df, timetable_df):
    output_string_parts = []

    # --- Robust Checks for essential columns ---
    required_timetable_cols = ["date", "shift", "Time", "Class", "Paper Code", "Paper Name"]
    for col in required_timetable_cols:
        if col not in timetable_df.columns:
            return f"Error: Missing essential column '{col}' in timetable.csv. Please ensure the file is correctly formatted."

    required_assigned_seats_cols = ["Roll Number", "Paper Code", "Paper Name", "Room Number", "Seat Number", "date", "shift"]
    for col in required_assigned_seats_cols:
        if col not in assigned_seats_df.columns:
            return f"Error: Missing essential column '{col}' in assigned_seats.csv. Please ensure seats are assigned and the file is correctly formatted."

    # 1. Get header information from timetable
    relevant_tt_exams = timetable_df[
        (timetable_df["date"].astype(str).str.strip() == date_str) &
        (timetable_df["shift"].astype(str).str.strip().str.lower() == shift.lower())
    ]

    if relevant_tt_exams.empty:
        return "No exams found for the selected date and shift to generate room chart."

    # Extract common info for header (assuming they are consistent for a given date/shift)
    exam_time = relevant_tt_exams.iloc[0]["Time"].strip() if "Time" in relevant_tt_exams.columns else ""
    
    # Determine the class summary for the header
    unique_classes = relevant_tt_exams['Class'].dropna().astype(str).str.strip().unique()
    class_summary_header = ""
    if len(unique_classes) == 1:
        class_summary_header = f"{unique_classes[0]} Examination {datetime.datetime.now().year}"
    elif len(unique_classes) > 1:
        class_summary_header = f"Various Classes Examination {datetime.datetime.now().year}"
    else:
        class_summary_header = f"Examination {datetime.datetime.now().year}"

    # Static header lines
    output_string_parts.append(",,,,,,,,,\nजीवाजी विश्वविद्यालय ग्वालियर ,,,,,,,,,\n\"परीक्षा केंद्र :- शासकीय विधि महाविद्यालय, मुरेना (म. प्र.) कोड :- G107 \",,,,,,,,,\n")
    output_string_parts.append(f"{class_summary_header},,,,,,,,,\n")
    output_string_parts.append(f"date :- ,,{date_str},,shift :-,{shift},,Time :- ,,\n")

    # 2. Get all assigned students for the given date and shift
    assigned_students_for_session = assigned_seats_df[
        (assigned_seats_df["date"] == date_str) &
        (assigned_seats_df["shift"] == shift)
    ].copy()

    if assigned_students_for_session.empty:
        output_string_parts.append("\nNo students assigned seats for this date and shift.")
        return "".join(output_string_parts)

    # Merge with timetable to get full paper names and Class
    # Ensure paper codes are comparable (e.g., int vs str)
    assigned_students_for_session['Paper Code'] = assigned_students_for_session['Paper Code'].astype(str)
    timetable_df['Paper Code'] = timetable_df['Paper Code'].astype(str)

    assigned_students_for_session = pd.merge(
        assigned_students_for_session,
        timetable_df[['Paper Code', 'Paper Name', 'Class']], # Need Class for the summary line
        on='Paper Code',
        how='left',
        suffixes=('', '_tt') # Suffixes are applied only if column names are duplicated in both DFs
    )
    # Use Paper Name from timetable if available, otherwise from assigned_seats_df
    assigned_students_for_session['Paper Name'] = assigned_students_for_session['Paper Name_tt'].fillna(assigned_students_for_session['Paper Name'])
    
    # Corrected line: Access 'Class' directly, as it would not have been suffixed if not present in assigned_seats_df
    # The 'Class' column from timetable_df is merged directly if assigned_seats_df doesn't have one,
    # otherwise it would be 'Class_tt'. We need to check which one exists.
    if 'Class_tt' in assigned_students_for_session.columns:
        assigned_students_for_session['Class'] = assigned_students_for_session['Class_tt'].fillna('')
    elif 'Class' in assigned_students_for_session.columns: # Fallback if 'Class' was already in assigned_seats_df
        assigned_students_for_session['Class'] = assigned_students_for_session['Class'].fillna('')
    else:
        assigned_students_for_session['Class'] = '' # Default if neither exists, though this should be caught by earlier checks

    # Sort by Room Number, then by Seat Number
    def sort_seat_number_key(seat):
        if isinstance(seat, str):
            match_a = re.match(r'(\d+)A', seat)
            match_b = re.match(r'(\d+)B', seat)
            if match_a:
                return (0, int(match_a.group(1))) # A-seats first
            elif match_b:
                return (1, int(match_b.group(1))) # B-seats second
            elif seat.isdigit():
                return (2, int(seat)) # Numeric seats last
        return (3, seat) # Fallback for unexpected formats

    assigned_students_for_session['sort_key'] = assigned_students_for_session['Seat Number'].apply(sort_seat_number_key)
    assigned_students_for_session = assigned_students_for_session.sort_values(by=['Room Number', 'sort_key']).drop(columns=['sort_key'])

    # Group by room for output
    students_by_room = assigned_students_for_session.groupby('Room Number')

    for room_num, room_data in students_by_room:
        output_string_parts.append(f"\n,,,कक्ष  :-,{room_num}  ,,,,\n") # Room header
        
        # Get unique papers for this room and session for the "परीक्षा का नाम" line
        unique_papers_in_room = room_data[['Class', 'Paper Code', 'Paper Name']].drop_duplicates()
        
        for _, paper_row in unique_papers_in_room.iterrows():
            paper_class = str(paper_row['Class']).strip()
            paper_code = str(paper_row['Paper Code']).strip()
            paper_name = str(paper_row['Paper Name']).strip()
            
            # Count students for this specific paper in this room
            students_for_this_paper_in_room = room_data[
                (room_data['Paper Code'].astype(str).str.strip() == paper_code) &
                (room_data['Paper Name'].astype(str).str.strip() == paper_name)
            ]
            num_students_for_paper = len(students_for_this_paper_in_room)

            output_string_parts.append(
                f"परीक्षा का नाम (Class - mode - Type),,,प्रश्न पत्र (paper- paper code - paper name),,,,उत्तर पुस्तिकाएं (number of students),,\n"
                f",,,,,,,प्राप्त ,प्रयुक्त ,शेष \n"
                f"{paper_class} - Regular - Regular,,,{paper_code} - {paper_name}        ,,,,{num_students_for_paper},,\n" # Assuming Regular for now
            )
            output_string_parts.append(",,,,,,,,,\n") # Blank line

        output_string_parts.append(",,,,,,,,,\n") # Blank line
        output_string_parts.append(f",,,Total,,,,{len(room_data)},,\n") # Total for the room
        output_string_parts.append(",,,,,,,,,\n") # Blank line
        output_string_parts.append("roll number - (room number-seat number) - 20 letters of paper name,,,,,,,,,\n")

        # Now add the roll number lines
        current_line_students = []
        for _, student_row in room_data.iterrows():
            roll_num = str(student_row['Roll Number']).strip()
            room_num_display = str(student_row['Room Number']).strip()
            seat_num_display = str(student_row['Seat Number']).strip()
            paper_name_display = str(student_row['Paper Name']).strip()
            
            # Truncate paper name to first 20 characters
            truncated_paper_name = paper_name_display[:20]

            student_entry = f"{roll_num}( कक्ष-{room_num_display}-सीट-{seat_num_display})-{truncated_paper_name}"
            current_line_students.append(student_entry)

            if len(current_line_students) == 10:
                output_string_parts.append(",".join(current_line_students) + "\n")
                current_line_students = []
        
        # Add any remaining students in the last line for the room
        if current_line_students:
            output_string_parts.append(",".join(current_line_students) + "\n")
        
        output_string_parts.append("\n") # Add an extra newline between rooms

    return "".join(output_string_parts)

# Function to generate UFM print form
# Corrected function to generate UFM print form
def generate_ufm_print_form(ufm_roll_number, attestation_df, assigned_seats_df, timetable_df,
                            report_date, report_shift, report_paper_code, report_paper_name):
    """
    Generates a printable UFM form for a given roll number within a specific exam context.

    Args:
        ufm_roll_number (str): The roll number of the student with UFM.
        attestation_df (pd.DataFrame): DataFrame loaded from attestation_data_combined.csv.
        assigned_seats_df (pd.DataFrame): DataFrame containing assigned seats information.
        timetable_df (pd.DataFrame): DataFrame containing the examination timetable.
        report_date (str): The date of the exam session (DD-MM-YYYY).
        report_shift (str): The shift of the exam session (Morning/Evening).
        report_paper_code (str): The paper code of the exam session.
        report_paper_name (str): The paper name of the exam session.

    Returns:
        str: A formatted string containing the UFM print form details, or an error message.
    """
    ufm_roll_number = str(ufm_roll_number).strip()

    # Retrieve student details from attestation_df
    student_details = attestation_df[attestation_df['Roll Number'].astype(str).str.strip() == ufm_roll_number]
    if student_details.empty:
        return f"Error: Student with Roll Number {ufm_roll_number} not found in attestation data."
    student_detail = student_details.iloc[0]

    # Get exam details specific to the UFM incident from assigned_seats and timetable
    relevant_assigned_seat = assigned_seats_df[
        (assigned_seats_df['Roll Number'].astype(str).str.strip() == ufm_roll_number) &
        (assigned_seats_df['date'].astype(str).str.strip() == report_date) &
        (assigned_seats_df['shift'].astype(str).str.strip() == report_shift) &
        (assigned_seats_df['Paper Code'].astype(str).str.strip() == _format_paper_code(report_paper_code)) &
        (assigned_seats_df['Paper Name'].astype(str).str.strip() == report_paper_name)
    ]
    
    exam_room_number = "N/A"
    exam_paper_code = _format_paper_code(report_paper_code)
    exam_paper_name = report_paper_name
    exam_time = "N/A"
    exam_class = "N/A"

    if not relevant_assigned_seat.empty:
        assigned_info = relevant_assigned_seat.iloc[0]
        exam_room_number = str(assigned_info['Room Number']).strip()
        
        matching_timetable_entry = timetable_df[
            (timetable_df['date'].astype(str).str.strip() == report_date) &
            (timetable_df['shift'].astype(str).str.strip() == report_shift) &
            (timetable_df['Paper Code'].astype(str).str.strip() == _format_paper_code(report_paper_code)) &
            (timetable_df['Paper Name'].astype(str).str.strip() == report_paper_name)
        ]
        if not matching_timetable_entry.empty:
            exam_time = str(matching_timetable_entry.iloc[0]['Time']).strip()
            exam_class = str(matching_timetable_entry.iloc[0]['Class']).strip()
    else:
        # Fallback if student is UFM'd but not found in assigned_seats for that specific session.
        matching_timetable_entry = timetable_df[
            (timetable_df['date'].astype(str).str.strip() == report_date) &
            (timetable_df['shift'].astype(str).str.strip() == report_shift) &
            (timetable_df['Paper Code'].astype(str).str.strip() == _format_paper_code(report_paper_code)) &
            (timetable_df['Paper Name'].astype(str).str.strip() == report_paper_name)
        ]
        if not matching_timetable_entry.empty:
            exam_time = str(matching_timetable_entry.iloc[0]['Time']).strip()
            exam_class = str(matching_timetable_entry.iloc[0]['Class']).strip()
        else:
            exam_time = "Not Found in Timetable"
            exam_class = "Not Found in Timetable"


    form_parts = []
    form_parts.append("--- UFM Case Print Form ---")
    form_parts.append("\n**1. Jiwaji University, Gwalior**")
    form_parts.append(f"\n**2. Class:** {exam_class} - {datetime.datetime.now().strftime('%B')}-{datetime.datetime.now().year}-Examination")
    form_parts.append(f"\n**3. Roll Number:** {ufm_roll_number}")
    form_parts.append(f"\n**4. Name of Student:** {student_detail.get('Name', 'N/A')}")
    form_parts.append(f"   **Address:** {student_detail.get('Address', 'N/A')}")
    form_parts.append(f"\n**5. Father's Name:** {student_detail.get('Father\'s Name', 'N/A')}")
    form_parts.append(f"\n**6. College Name:** {student_detail.get('College Name', 'N/A')}")
    form_parts.append(f"\n**7. Exam Center Name:** {student_detail.get('Exam Centre', 'N/A')} Code: G107")
    form_parts.append(f"\n**8. Paper Code & Paper Name:** {exam_paper_code} - {exam_paper_name}")
    form_parts.append(f"\n**9. date:** {report_date}")
    form_parts.append(f"**10. Time:** {report_shift} shift ({exam_time})")
    form_parts.append(f"\n**11. Time of UFM:** _________________________")
    form_parts.append(f"**12. Name of Book/Material:** _________________________")
    form_parts.append(f"**13. Number of pages/details:** _________________________")
    form_parts.append(f"\n**Room Number (where UFM occurred):** {exam_room_number}")
    form_parts.append("\n\n_________________________")
    form_parts.append("Signature of Invigilator(s)")
    form_parts.append("\n\n_________________________")
    form_parts.append("Signature of Centre Superintendent")
    form_parts.append("\n\n--- End of UFM Case Print Form ---")

    return "\n".join(form_parts)


    def _generate_ufm_print_form_with_context(ufm_roll_number, attestation_df, assigned_seats_df, timetable_df,
                                              report_date, report_shift, report_paper_code, report_paper_name):
        """
        Generates a printable UFM form for a given roll number within a specific exam context.
        """
        ufm_roll_number = str(ufm_roll_number).strip()

        # Retrieve student details from attestation_df
        student_details = attestation_df[attestation_df['Roll Number'].astype(str).str.strip() == ufm_roll_number]
        if student_details.empty:
            return f"Error: Student with Roll Number {ufm_roll_number} not found in attestation data."
        student_detail = student_details.iloc[0]

        # Get exam details specific to the UFM incident from assigned_seats and timetable
        # Filter assigned_seats by roll number, date, shift, paper code, paper name
        relevant_assigned_seat = assigned_seats_df[
            (assigned_seats_df['Roll Number'].astype(str).str.strip() == ufm_roll_number) &
            (assigned_seats_df['date'].astype(str).str.strip() == report_date) &
            (assigned_seats_df['shift'].astype(str).str.strip() == report_shift) &
            (assigned_seats_df['Paper Code'].astype(str).str.strip() == _format_paper_code(report_paper_code)) &
            (assigned_seats_df['Paper Name'].astype(str).str.strip() == report_paper_name)
        ]
        
        exam_room_number = "N/A"
        exam_paper_code = _format_paper_code(report_paper_code)
        exam_paper_name = report_paper_name
        exam_time = "N/A"
        exam_class = "N/A" # Will get from timetable

        if not relevant_assigned_seat.empty:
            assigned_info = relevant_assigned_seat.iloc[0]
            exam_room_number = str(assigned_info['Room Number']).strip()
            
            # Get exam time and class from timetable using date, shift, paper code, paper name
            matching_timetable_entry = timetable_df[
                (timetable_df['date'].astype(str).str.strip() == report_date) &
                (timetable_df['shift'].astype(str).str.strip() == report_shift) &
                (timetable_df['Paper Code'].astype(str).str.strip() == _format_paper_code(report_paper_code)) &
                (timetable_df['Paper Name'].astype(str).str.strip() == report_paper_name)
            ]
            if not matching_timetable_entry.empty:
                exam_time = str(matching_timetable_entry.iloc[0]['Time']).strip()
                exam_class = str(matching_timetable_entry.iloc[0]['Class']).strip()
        else:
            # Fallback if student is UFM'd but not found in assigned_seats for that specific session.
            # This might happen if they were unassigned but still appeared for exam.
            # In such cases, we still use the report_paper_code/name but Room will be N/A.
            matching_timetable_entry = timetable_df[
                (timetable_df['date'].astype(str).str.strip() == report_date) &
                (timetable_df['shift'].astype(str).str.strip() == report_shift) &
                (timetable_df['Paper Code'].astype(str).str.strip() == _format_paper_code(report_paper_code)) &
                (timetable_df['Paper Name'].astype(str).str.strip() == report_paper_name)
            ]
            if not matching_timetable_entry.empty:
                exam_time = str(matching_timetable_entry.iloc[0]['Time']).strip()
                exam_class = str(matching_timetable_entry.iloc[0]['Class']).strip()
            else:
                exam_time = "Not Found in Timetable"
                exam_class = "Not Found in Timetable"


        # Prepare the formatted string
        form_parts = []
        form_parts.append("--- UFM Case Print Form ---")
        form_parts.append("\n**1. Jiwaji University, Gwalior**")
        form_parts.append(f"\n**2. Class:** {exam_class} - {datetime.datetime.now().strftime('%B')}-{datetime.datetime.now().year}-Examination")
        form_parts.append(f"\n**3. Roll Number:** {ufm_roll_number}")
        form_parts.append(f"\n**4. Name of Student:** {student_detail.get('Name', 'N/A')}")
        form_parts.append(f"   **Address:** {student_detail.get('Address', 'N/A')}")
        form_parts.append(f"\n**5. Father's Name:** {student_detail.get('Father\'s Name', 'N/A')}")
        form_parts.append(f"\n**6. College Name:** {student_detail.get('College Name', 'N/A')}")
        form_parts.append(f"\n**7. Exam Center Name:** {student_detail.get('Exam Centre', 'N/A')} Code: G107")
        form_parts.append(f"\n**8. Paper Code & Paper Name:** {exam_paper_code} - {exam_paper_name}")
        form_parts.append(f"\n**9. date:** {report_date}")
        form_parts.append(f"**10. Time:** {report_shift} shift ({exam_time})")
        form_parts.append(f"\n**11. Time of UFM:** _________________________")
        form_parts.append(f"**12. Name of Book/Material:** _________________________")
        form_parts.append(f"**13. Number of pages/details:** _________________________")
        form_parts.append(f"\n**Room Number (where UFM occurred):** {exam_room_number}")
        form_parts.append("\n\n_________________________")
        form_parts.append("Signature of Invigilator(s)")
        form_parts.append("\n\n_________________________")
        form_parts.append("Signature of Centre Superintendent")
        form_parts.append("\n\n--- End of UFM Case Print Form ---")

        return "\n".join(form_parts)

    return _generate_ufm_print_form_with_context # Return the inner function so it can be called with context


def display_report_panel():
    st.subheader("📊 Exam Session Reports")

    # 1. Load all necessary data
    sitting_plan, timetable, assigned_seats_df, attestation_df = load_data()
    all_reports_df = load_cs_reports_csv()
    room_invigilators_df = load_room_invigilator_assignments()

    if all_reports_df.empty and room_invigilators_df.empty:
        st.info("No Centre Superintendent reports or invigilator assignments available yet for statistics.")
        return

    if assigned_seats_df.empty:
        st.warning("Assigned seats data is required to calculate expected student counts. Please assign seats first.")
        return

    # 2. Prepare "Expected Students" from Assigned Seats (Source of Truth)
    # Normalize text columns to ensure merges work correctly
    assigned_seats_df['Room Number'] = assigned_seats_df['Room Number'].astype(str).str.strip()
    assigned_seats_df['Paper Code'] = assigned_seats_df['Paper Code'].astype(str).str.strip().str.lower()
    assigned_seats_df['Paper Name'] = assigned_seats_df['Paper Name'].astype(str).str.strip().str.lower()
    assigned_seats_df['date'] = assigned_seats_df['date'].astype(str).str.strip()
    assigned_seats_df['shift'] = assigned_seats_df['shift'].astype(str).str.strip().str.lower()

    # We need 'Class' info which is in Timetable, not usually in Assigned Seats
    timetable['Paper Code'] = timetable['Paper Code'].astype(str).str.strip().str.lower()
    timetable['date'] = timetable['date'].astype(str).str.strip()
    timetable['shift'] = timetable['shift'].astype(str).str.strip().str.lower()
    
    # Create a lookup for Class based on Paper/Date/Shift
    timetable_lookup = timetable[['date', 'shift', 'Paper Code', 'Class']].drop_duplicates()
    timetable_lookup['Class'] = timetable_lookup['Class'].astype(str).str.strip().str.lower()

    # Merge Class info into Assigned Seats
    assigned_seats_with_class = pd.merge(
        assigned_seats_df,
        timetable_lookup,
        on=['date', 'shift', 'Paper Code'],
        how='left'
    )
    assigned_seats_with_class['Class'] = assigned_seats_with_class['Class'].fillna('unknown')

    # 3. Aggregate: Count students per Room/Paper/Session
    # This eliminates any duplicate rows issues
    expected_students_aggregated = assigned_seats_with_class.groupby(
        ['date', 'shift', 'Room Number', 'Paper Code', 'Paper Name', 'Class']
    )['Roll Number'].count().reset_index()
    
    expected_students_aggregated.rename(columns={'Roll Number': 'expected_students_count'}, inplace=True)

    # 4. Prepare Reports Data
    all_reports_df['date'] = all_reports_df['date'].astype(str).str.strip()
    all_reports_df['shift'] = all_reports_df['shift'].astype(str).str.strip().str.lower()
    all_reports_df['room_num'] = all_reports_df['room_num'].astype(str).str.strip()
    all_reports_df['paper_code'] = all_reports_df['paper_code'].astype(str).str.strip().str.lower()
    all_reports_df['paper_name'] = all_reports_df['paper_name'].astype(str).str.strip().str.lower()
    all_reports_df['class'] = all_reports_df['class'].astype(str).str.strip().str.lower()

    # 5. Merge Reports with Expected Counts
    # Key fix: Matching strictly on Date + Shift + Room + Paper
    merged_reports_df = pd.merge(
        all_reports_df,
        expected_students_aggregated,
        left_on=['date', 'shift', 'room_num', 'paper_code', 'paper_name', 'class'],
        right_on=['date', 'shift', 'Room Number', 'Paper Code', 'Paper Name', 'Class'],
        how='left',
        suffixes=('_report', '_assigned')
    )

    # Fill NaN expected_students_count with 0 (if a report exists for a room with 0 assigned seats, theoretically impossible but safe to handle)
    merged_reports_df['expected_students_count'] = merged_reports_df['expected_students_count'].fillna(0).astype(int)

    # 6. Merge Invigilators
    if not room_invigilators_df.empty:
        room_invigilators_df['date'] = room_invigilators_df['date'].astype(str).str.strip()
        room_invigilators_df['shift'] = room_invigilators_df['shift'].astype(str).str.strip().str.lower()
        room_invigilators_df['room_num'] = room_invigilators_df['room_num'].astype(str).str.strip()

        merged_reports_df = pd.merge(
            merged_reports_df,
            room_invigilators_df[['date', 'shift', 'room_num', 'invigilators']],
            on=['date', 'shift', 'room_num'],
            how='left',
            suffixes=('', '_room_inv') 
        )
        merged_reports_df['invigilators'] = merged_reports_df['invigilators'].apply(lambda x: x if isinstance(x, list) else [])
    else:
        merged_reports_df['invigilators'] = [[]] * len(merged_reports_df)

    # 7. Calculate & Display Overall Statistics
    st.markdown("---")
    st.subheader("Overall Statistics")

    total_reports = len(merged_reports_df)
    unique_sessions = merged_reports_df['report_key'].nunique()
    
    # Calculate Expected Students sum
    total_expected_students = merged_reports_df['expected_students_count'].sum()
    
    # Calculate Absent/UFM
    total_absent = merged_reports_df['absent_roll_numbers'].apply(len).sum()
    total_ufm = merged_reports_df['ufm_roll_numbers'].apply(len).sum()
    
    total_present_students = total_expected_students - total_absent
    total_answer_sheets_collected = total_present_students - total_ufm

    overall_attendance_percentage = 0
    if total_expected_students > 0:
        overall_attendance_percentage = (total_present_students / total_expected_students) * 100

    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.metric("Total Reports Submitted", total_reports)
    with col2:
        st.metric("Unique Exam Sessions Reported", unique_sessions)
    with col3:
        st.metric("Total Expected Students", total_expected_students)
    with col4:
        st.metric("Total Absent Students", total_absent)
    with col5:
        st.metric("Overall Attendance (%)", f"{overall_attendance_percentage:.2f}%")
    
    col_metrics_2_1, col_metrics_2_2, col_metrics_2_3 = st.columns(3)
    with col_metrics_2_1:
        st.metric("Total Present Students", total_present_students)
    with col_metrics_2_2:
        st.metric("Total UFM Cases", total_ufm)
    with col_metrics_2_3:
        st.metric("Total Answer Sheets Collected", total_answer_sheets_collected)


    # --- Paper-wise Statistics ---
    st.markdown("---")
    st.subheader("Paper-wise Statistics")

    # Group aggregated expected data by Paper
    expected_by_paper = expected_students_aggregated.groupby(['Paper Name', 'Paper Code']).agg(
        expected_students=('expected_students_count', 'sum')
    ).reset_index()
    
    expected_by_paper.rename(columns={'Paper Name': 'paper_name', 'Paper Code': 'paper_code'}, inplace=True)
    expected_by_paper['paper_name'] = expected_by_paper['paper_name'].astype(str).str.strip().str.lower()
    expected_by_paper['paper_code'] = expected_by_paper['paper_code'].astype(str).str.strip().str.lower()

    # Group reported data by Paper
    reported_by_paper = merged_reports_df.groupby(['paper_name', 'paper_code']).agg(
        total_absent=('absent_roll_numbers', lambda x: x.apply(len).sum()),
        total_ufm=('ufm_roll_numbers', lambda x: x.apply(len).sum())
    ).reset_index()

    # Merge
    paper_stats = pd.merge(
        expected_by_paper,
        reported_by_paper,
        on=['paper_name', 'paper_code'],
        how='left'
    )

    paper_stats['total_absent'] = paper_stats['total_absent'].fillna(0).astype(int)
    paper_stats['total_ufm'] = paper_stats['total_ufm'].fillna(0).astype(int)
    paper_stats['total_present'] = paper_stats['expected_students'] - paper_stats['total_absent']
    paper_stats['total_answer_sheets_collected'] = paper_stats['total_present'] - paper_stats['total_ufm']
    paper_stats['attendance_percentage'] = paper_stats.apply(
        lambda row: (row['total_present'] / row['expected_students'] * 100) if row['expected_students'] > 0 else 0,
        axis=1
    )
    paper_stats['attendance_percentage'] = paper_stats['attendance_percentage'].map('{:.2f}%'.format)

    # Rename for display
    paper_stats.rename(columns={
        'paper_name': 'Paper Name', 'paper_code': 'Paper Code',
        'expected_students': 'Expected Students', 'total_absent': 'Absent Students',
        'total_present': 'Present Students', 'total_ufm': 'UFM Cases',
        'total_answer_sheets_collected': 'Answer Sheets Collected', 'attendance_percentage': 'Attendance (%)'
    }, inplace=True)

    st.dataframe(paper_stats[['Paper Name', 'Paper Code', 'Expected Students', 'Present Students', 'Absent Students', 'UFM Cases', 'Answer Sheets Collected', 'Attendance (%)']])

    # --- Class-wise Statistics ---
    # (Replaced "Type-wise" because Assigned Seats doesn't usually have Type/Mode columns)
    st.markdown("---")
    st.subheader("Class-wise Statistics")

    expected_by_class = expected_students_aggregated.groupby(['Class']).agg(
        expected_students=('expected_students_count', 'sum')
    ).reset_index()
    expected_by_class['Class'] = expected_by_class['Class'].astype(str).str.strip().str.lower()

    reported_by_class = merged_reports_df.groupby(['class']).agg(
        total_absent=('absent_roll_numbers', lambda x: x.apply(len).sum()),
        total_ufm=('ufm_roll_numbers', lambda x: x.apply(len).sum())
    ).reset_index()
    reported_by_class.rename(columns={'class': 'Class'}, inplace=True)

    class_stats = pd.merge(expected_by_class, reported_by_class, on='Class', how='left')

    class_stats['total_absent'] = class_stats['total_absent'].fillna(0).astype(int)
    class_stats['total_ufm'] = class_stats['total_ufm'].fillna(0).astype(int)
    class_stats['total_present'] = class_stats['expected_students'] - class_stats['total_absent']
    class_stats['total_answer_sheets_collected'] = class_stats['total_present'] - class_stats['total_ufm']
    class_stats['attendance_percentage'] = class_stats.apply(
        lambda row: (row['total_present'] / row['expected_students'] * 100) if row['expected_students'] > 0 else 0,
        axis=1
    )
    class_stats['attendance_percentage'] = class_stats['attendance_percentage'].map('{:.2f}%'.format)

    class_stats.rename(columns={
        'expected_students': 'Expected Students', 'total_absent': 'Absent Students',
        'total_present': 'Present Students', 'total_ufm': 'UFM Cases',
        'total_answer_sheets_collected': 'Answer Sheets Collected', 'attendance_percentage': 'Attendance (%)'
    }, inplace=True)

    st.dataframe(class_stats[['Class', 'Expected Students', 'Present Students', 'Absent Students', 'UFM Cases', 'Answer Sheets Collected', 'Attendance (%)']])


    # --- Filtering & Detailed Lists (Unchanged logic) ---
    st.markdown("---")
    st.subheader("Filter and View Reports")

    unique_dates = sorted(merged_reports_df['date'].unique())
    unique_shifts = sorted(merged_reports_df['shift'].unique())
    unique_rooms = sorted(merged_reports_df['room_num'].unique())
    unique_papers = sorted(merged_reports_df['paper_name'].unique())

    filter_date = st.selectbox("Filter by Date", ["All"] + unique_dates, key="report_filter_date")
    filter_shift = st.selectbox("Filter by Shift", ["All"] + unique_shifts, key="report_filter_shift")
    filter_room = st.selectbox("Filter by Room Number", ["All"] + unique_rooms, key="report_filter_room")
    filter_paper = st.selectbox("Filter by Paper Name", ["All"] + unique_papers, key="report_filter_paper")

    filtered_reports_df = merged_reports_df.copy()

    if filter_date != "All": filtered_reports_df = filtered_reports_df[filtered_reports_df['date'] == filter_date]
    if filter_shift != "All": filtered_reports_df = filtered_reports_df[filtered_reports_df['shift'] == filter_shift]
    if filter_room != "All": filtered_reports_df = filtered_reports_df[filtered_reports_df['room_num'] == filter_room]
    if filter_paper != "All": filtered_reports_df = filtered_reports_df[filtered_reports_df['paper_name'] == filter_paper]

    if filtered_reports_df.empty:
        st.info("No reports match the selected filters.")
    else:
        st.dataframe(filtered_reports_df[[
            'date', 'shift', 'room_num', 'paper_code', 'paper_name', 'invigilators', 
            'absent_roll_numbers', 'ufm_roll_numbers'
        ]])
        
        # --- Downloads for Absent/UFM Lists ---
        st.markdown("---")
        st.subheader("Detailed Absentee List (Filtered)")
        absent_list_data = []
        for _, row in filtered_reports_df.iterrows():
            for roll in row['absent_roll_numbers']:
                absent_list_data.append({
                    'date': row['date'], 'shift': row['shift'], 'Room': row['room_num'],
                    'Paper Code': row['paper_code'], 'Paper Name': row['paper_name'], 'Absent Roll Number': roll
                })
        
        if absent_list_data:
            df_absent = pd.DataFrame(absent_list_data)
            st.dataframe(df_absent)
            csv_absent = df_absent.to_csv(index=False).encode('utf-8')
            st.download_button("Download Absentee List as CSV", csv_absent, f"absent_list_{filter_date}_{filter_shift}.csv", "text/csv")
        else:
            st.info("No absent students in the filtered reports.")

        st.markdown("---")
        st.subheader("Detailed UFM List (Filtered)")
        ufm_list_data = []
        for _, row in filtered_reports_df.iterrows():
            for roll in row['ufm_roll_numbers']:
                ufm_list_data.append({
                    'date': row['date'], 'shift': row['shift'], 'Room': row['room_num'],
                    'Paper Code': row['paper_code'], 'Paper Name': row['paper_name'], 'UFM Roll Number': roll
                })
        
        if ufm_list_data:
            df_ufm = pd.DataFrame(ufm_list_data)
            st.dataframe(df_ufm)
            csv_ufm = df_ufm.to_csv(index=False).encode('utf-8')
            st.download_button("Download UFM List as CSV", csv_ufm, f"ufm_list_{filter_date}_{filter_shift}.csv", "text/csv")
        else:
            st.info("No UFM cases in the filtered reports.")
            
# --- Updated Remuneration Calculation Functions (from bill.py) ---

def calculate_remuneration(shift_assignments_df, room_invigilator_assignments_df, timetable_df, assigned_seats_df,
                           manual_rates, prep_closing_assignments, holiday_dates, selected_classes_for_bill):
    """
    Calculates the remuneration for all team members based on assignments and rules,
    including individually selected preparation and closing days and holiday conveyance allowance.

    Updated Rules:
    1. Person gets conveyance only if they worked in both shifts of the same date
       (in selected and/or non-selected class exam).
    2. If eligible, conveyance will be paid in the evening shift of a selected exam only.
    3. Even if eligible, conveyance will not be paid in the morning shift of a selected exam.
    4. Senior CS gets daily remuneration if worked in either shift of selected exam in bill of selected exam.
    5. Senior CS doesn't get daily remuneration if worked in morning shift of selected exam and evening shift of non-selected exam in bill of selected exam.
    6. Senior CS gets daily remuneration if worked in both shifts of selected exam in bill of selected exam.
    7. Senior CS will also get exam day conveyance like others if they worked in both shifts.
    """
    remuneration_rules = {
        'senior_center_superintendent': {'role_display': 'Senior Center Superintendent', 'rate': manual_rates['senior_center_superintendent_rate'], 'unit': 'day', 'eligible_prep_close': True, 'exam_conveyance': True},
        'center_superintendent': {'role_display': 'Center Superintendent', 'rate': manual_rates['center_superintendent_rate'], 'unit': 'shift', 'eligible_prep_close': True, 'exam_conveyance': True},
        'assistant_center_superintendent': {'role_display': 'Assistant Center Superintendent', 'rate': manual_rates['assistant_center_superintendent_rate'], 'unit': 'shift', 'eligible_prep_close': True, 'exam_conveyance': True},
        'permanent_invigilator': {'role_display': 'Permanent Invigilator', 'rate': manual_rates['permanent_invigilator_rate'], 'unit': 'shift', 'eligible_prep_close': True, 'exam_conveyance': True},
        'assistant_permanent_invigilator': {'role_display': 'Assistant Permanent Invigilator', 'rate': manual_rates['assistant_permanent_invigilator_rate'], 'unit': 'shift', 'eligible_prep_close': False, 'exam_conveyance': True},
        'invigilator': {'role_display': 'Invigilator', 'rate': manual_rates['invigilator_rate'], 'unit': 'shift', 'eligible_prep_close': False, 'exam_conveyance': True},
    }

    class_worker_rates = {
        'class_3_worker': {'role_display': 'Class 3 Worker', 'rate_per_student': manual_rates['class_3_worker_rate_per_student']},
        'class_4_worker': {'role_display': 'Class 4 Worker', 'rate_per_student': manual_rates['class_4_worker_rate_per_student']},
    }

    remuneration_data_detailed_raw = []
    
    unique_class_3_workers = set()
    unique_class_4_workers = set()

    unified_assignments = []

    # Get the unique dates and shifts of the selected classes
    if selected_classes_for_bill:
        timetable_df['Class'] = timetable_df['Class'].str.strip()
        filtered_timetable_dates = timetable_df[timetable_df['Class'].isin(selected_classes_for_bill)]
        selected_dates = set(filtered_timetable_dates['date'].unique())
        selected_shifts = set(filtered_timetable_dates['shift'].unique())
    else:
        selected_dates = set(timetable_df['date'].unique())
        selected_shifts = set(timetable_df['shift'].unique())

    # --- UPdateD LOGIC FOR CLASS 3 AND 4 WORKERS ---
    # Populate the worker lists only from assignments on selected exam dates
    filtered_shift_assignments = shift_assignments_df[shift_assignments_df['date'].isin(selected_dates)]
    
    for index, row in filtered_shift_assignments.iterrows():
        if 'class_3_worker' in row and isinstance(row['class_3_worker'], list):
            unique_class_3_workers.update(row['class_3_worker'])
        if 'class_4_worker' in row and isinstance(row['class_4_worker'], list):
            unique_class_4_workers.update(row['class_4_worker'])
    
    # OLD LOGIC for other roles - no changes here
    for index, row in shift_assignments_df.iterrows():
        current_date = row['date']
        current_shift = row['shift']
        
        for role_col in remuneration_rules.keys():
            if role_col in row and isinstance(row[role_col], list):
                for person in row[role_col]:
                    unified_assignments.append({
                        'Name': person,
                        'Role_Key': role_col,
                        'date': current_date,
                        'shift': current_shift,
                        'Source': 'shift_assignments'
                    })
    
    for index, row in room_invigilator_assignments_df.iterrows():
        current_date = row['date']
        current_shift = row['shift']
        invigilators_list = row['invigilators']

        for invigilator in invigilators_list:
            is_assigned_higher_role = False
            for assignment in unified_assignments:
                if (assignment['Name'] == invigilator and
                    assignment['date'] == current_date and
                    assignment['shift'] == current_shift and
                    assignment['Role_Key'] != 'invigilator'):
                    is_assigned_higher_role = True
                    break
            
            if not is_assigned_higher_role:
                unified_assignments.append({
                    'Name': invigilator,
                    'Role_Key': 'invigilator',
                    'date': current_date,
                    'shift': current_shift,
                    'Source': 'room_invigilator_assignments'
                })

    df_assignments = pd.DataFrame(unified_assignments)
    
    session_classes_map = {}
    for _, tt_row in timetable_df.iterrows():
        date_shift_key = (str(tt_row['date']), str(tt_row['shift']))
        if date_shift_key not in session_classes_map:
            session_classes_map[date_shift_key] = set()
        session_classes_map[date_shift_key].add(str(tt_row['Class']).strip())

    workers_with_both_shifts = set()
    if not df_assignments.empty:
        df_assignments['date_dt'] = pd.to_datetime(df_assignments['date'], format='%d-%m-%Y', errors='coerce')
        shift_counts = df_assignments.groupby(['Name', 'date'])['shift'].nunique().reset_index()
        eligible_workers_df = shift_counts[shift_counts['shift'] == 2]
        for _, row in eligible_workers_df.iterrows():
            workers_with_both_shifts.add((row['Name'], row['date']))

    for assignment in unified_assignments:
        name = assignment['Name']
        role_key = assignment['Role_Key']
        date = assignment['date']
        shift = assignment['shift']

        session_classes = list(session_classes_map.get((date, shift), set()))
        is_selected_exam = any(cls in selected_classes_for_bill for cls in [c.strip() for c in session_classes]) if selected_classes_for_bill else True
        
        base_rem_for_shift = remuneration_rules[role_key]['rate']
        
        conveyance = 0
        if remuneration_rules[role_key]['exam_conveyance']:
            if (name, date) in workers_with_both_shifts:
                if shift == 'Evening' and is_selected_exam:
                    conveyance = manual_rates['conveyance_rate']
                elif shift == 'Morning' and is_selected_exam:
                    conveyance = 0
        
        remuneration_data_detailed_raw.append({
            'Name': name,
            'Role_Key': role_key,
            'Role_Display': remuneration_rules[role_key]['role_display'],
            'date': date,
            'shift': shift,
            'Base_Remuneration_Per_shift_Unfiltered': base_rem_for_shift,
            'Conveyance': conveyance,
            'Is_Selected_Exam': is_selected_exam,
            'Classes_in_Session': session_classes,
        })
    
    df_detailed_remuneration = pd.DataFrame(remuneration_data_detailed_raw)

    # --- Generate Individual Bills (Corrected Sorting) ---
    individual_bills = []
    
    role_order_keys = ['senior_center_superintendent', 'center_superintendent', 'assistant_center_superintendent', 'permanent_invigilator', 'assistant_permanent_invigilator', 'invigilator']
    unique_person_roles = df_detailed_remuneration[['Name', 'Role_Display', 'Role_Key']].drop_duplicates()
    
    unique_person_roles_sorted = sorted(
        unique_person_roles.to_dict('records'),
        key=lambda x: role_order_keys.index(x['Role_Key'])
    )

    for i, person in enumerate(unique_person_roles_sorted):
        name = person['Name']
        role_display = person['Role_Display']
        role_key = person['Role_Key']
        
        person_data = df_detailed_remuneration[(df_detailed_remuneration['Name'] == name) & (df_detailed_remuneration['Role_Display'] == role_display)].copy()
        
        if selected_classes_for_bill:
            filtered_person_data = person_data[person_data['Is_Selected_Exam'] == True].copy()
        else:
            filtered_person_data = person_data.copy()

        duty_dates_morning_str = ""
        morning_shifts_df = filtered_person_data[filtered_person_data['shift'] == 'Morning']
        if not morning_shifts_df.empty:
            morning_shifts_df['date_dt'] = pd.to_datetime(morning_shifts_df['date'], format='%d-%m-%Y', errors='coerce')
            morning_shifts_df = morning_shifts_df.sort_values(by='date_dt')
            grouped_dates_morning = morning_shifts_df.groupby(morning_shifts_df['date_dt'].dt.to_period('M'))['date_dt'].apply(lambda x: sorted(x.dt.day.tolist()))
            date_parts = []
            for period, days in grouped_dates_morning.items():
                month_name = period.strftime('%b')
                days_str = ", ".join(map(str, days))
                date_parts.append(f"{month_name} - {days_str}")
            if date_parts:
                duty_dates_morning_str = ", ".join(date_parts) + f" {morning_shifts_df['date_dt'].min().year}"

        duty_dates_evening_str = ""
        evening_shifts_df = filtered_person_data[filtered_person_data['shift'] == 'Evening']
        if not evening_shifts_df.empty:
            evening_shifts_df['date_dt'] = pd.to_datetime(evening_shifts_df['date'], format='%d-%m-%Y', errors='coerce')
            evening_shifts_df = evening_shifts_df.sort_values(by='date_dt')
            grouped_dates_evening = evening_shifts_df.groupby(evening_shifts_df['date_dt'].dt.to_period('M'))['date_dt'].apply(lambda x: sorted(x.dt.day.tolist()))
            date_parts = []
            for period, days in grouped_dates_evening.items():
                month_name = period.strftime('%b')
                days_str = ", ".join(map(str, days))
                date_parts.append(f"{month_name} - {days_str}")
            if date_parts:
                duty_dates_evening_str = ", ".join(date_parts) + f" {evening_shifts_df['date_dt'].min().year}"
        
        total_morning_shifts = len(morning_shifts_df)
        total_evening_shifts = len(evening_shifts_df)
        total_shifts = total_morning_shifts + total_evening_shifts
        rate_in_rs = remuneration_rules[role_key]['rate'] if role_key in remuneration_rules else 0

        total_base_remuneration = 0
        if role_key == 'senior_center_superintendent':
            unique_dates = filtered_person_data['date'].nunique()
            total_base_remuneration = unique_dates * rate_in_rs
        else:
            total_base_remuneration = filtered_person_data['Base_Remuneration_Per_shift_Unfiltered'].sum()
        
        total_conveyance = filtered_person_data['Conveyance'].sum() 
        
        total_prep_remuneration = 0
        total_closing_remuneration = 0
        total_holiday_conveyance = 0
        
        if remuneration_rules[role_key]['eligible_prep_close']:
            person_assignments = prep_closing_assignments.get(name, {})
            assigned_role = person_assignments.get('role')
            
            if assigned_role == role_key:
                prep_days = person_assignments.get('prep_days', [])
                total_prep_remuneration = len(prep_days) * rate_in_rs
                
                closing_days = person_assignments.get('closing_days', [])
                total_closing_remuneration = len(closing_days) * rate_in_rs
                
                all_assigned_days = prep_days + closing_days
                holiday_assigned_days = [day for day in all_assigned_days if day in holiday_dates]
                total_holiday_conveyance = len(holiday_assigned_days) * manual_rates['holiday_conveyance_allowance_rate']

        grand_total_amount = total_base_remuneration + total_conveyance + total_prep_remuneration + total_closing_remuneration + total_holiday_conveyance

        if grand_total_amount > 0:
            individual_bills.append({
                'SN': len(individual_bills) + 1,
                'Name (with role)': f"{name} ({role_display})",
                'Duty dates of selected class exam shift (morning)': duty_dates_morning_str,
                'Duty dates of selected class exam shift (evening)': duty_dates_evening_str,
                'Total shifts of selected class exams (morning/evening)': total_shifts,
                'Rate in Rs': rate_in_rs,
                'Total Remuneration in Rs': total_base_remuneration,
                'Total Conveyance (in evening shift)': total_conveyance,
                'Preparation Day Remuneration': total_prep_remuneration,
                'Closing Day Remuneration': total_closing_remuneration,
                'Total Holiday Conveyance Added': total_holiday_conveyance,
                'Total amount in Rs': grand_total_amount,
                'Signature': ''
            })

    df_individual_bills = pd.DataFrame(individual_bills)

    if not df_individual_bills.empty:
        total_row = {
            'SN': '',
            'Name (with role)': 'TOTAL',
            'Duty dates of selected class exam shift (morning)': '',
            'Duty dates of selected class exam shift (evening)': '',
            'Total shifts of selected class exams (morning/evening)': df_individual_bills['Total shifts of selected class exams (morning/evening)'].sum(),
            'Rate in Rs': '',
            'Total Remuneration in Rs': df_individual_bills['Total Remuneration in Rs'].sum(),
            'Total Conveyance (in evening shift)': df_individual_bills['Total Conveyance (in evening shift)'].sum(),
            'Preparation Day Remuneration': df_individual_bills['Preparation Day Remuneration'].sum(),
            'Closing Day Remuneration': df_individual_bills['Closing Day Remuneration'].sum(),
            'Total Holiday Conveyance Added': df_individual_bills['Total Holiday Conveyance Added'].sum(),
            'Total amount in Rs': df_individual_bills['Total amount in Rs'].sum(),
            'Signature': ''
        }
        df_individual_bills = pd.concat([df_individual_bills, pd.DataFrame([total_row])], ignore_index=True)

    df_role_summary_matrix = generate_role_summary_matrix_by_date(df_detailed_remuneration, remuneration_rules, prep_closing_assignments, holiday_dates, manual_rates, selected_classes_for_bill, timetable_df, assigned_seats_df)
    
    # --- Corrected LOGIC FOR CLASS 3 AND 4 BILLS ---
    class_3_4_final_bills = []
    
    # Ensure key columns exist and convert to string for robust merging
    required_cols_timetable = ['Paper Code', 'date', 'shift', 'Class']
    required_cols_assigned_seats = ['Paper Code', 'date', 'shift', 'Roll Number']

    if all(col in timetable_df.columns for col in required_cols_timetable) and \
       all(col in assigned_seats_df.columns for col in required_cols_assigned_seats):

        timetable_df_temp = timetable_df.copy()
        assigned_seats_df_temp = assigned_seats_df.copy()
        for col in ['Paper Code', 'date', 'shift']:
            timetable_df_temp[col] = timetable_df_temp[col].astype(str)
            assigned_seats_df_temp[col] = assigned_seats_df_temp[col].astype(str)

        if selected_classes_for_bill:
            # Filter the timetable by the selected classes first
            filtered_timetable = timetable_df_temp[timetable_df_temp['Class'].isin(selected_classes_for_bill)]
            
            # Merge with assigned seats to get students for only those exams
            merged_df = pd.merge(assigned_seats_df_temp, filtered_timetable, on=['Paper Code', 'date', 'shift'], how='inner')
            total_students_for_class_workers = merged_df['Roll Number'].nunique()
        else:
            # If no classes are selected, count all unique students
            total_students_for_class_workers = assigned_seats_df['Roll Number'].nunique()
    else:
        print("Error: Missing required columns in either timetable_df or assigned_seats_df. Cannot calculate class worker bills.")
        total_students_for_class_workers = 0


    total_class_3_remuneration = 0
    total_class_4_remuneration = 0

    # Class 3 Workers Section
    if unique_class_3_workers:
        num_class_3_workers = len(unique_class_3_workers)
        class_3_total_remuneration = total_students_for_class_workers * class_worker_rates['class_3_worker']['rate_per_student']
        rem_per_class_3_worker = class_3_total_remuneration / num_class_3_workers if num_class_3_workers > 0 else 0
        total_class_3_remuneration = class_3_total_remuneration

        # Summary row for Class 3
        class_3_4_final_bills.append({
            'S.N.': 'Class 3 Workers',
            'Name': '',
            'Role': '',
            'Total Students (Center-wide)': total_students_for_class_workers,
            'Rate per Student (for category)': class_worker_rates['class_3_worker']['rate_per_student'],
            'Total Remuneration for Category (Rs.)': class_3_total_remuneration,
            'Number of Workers in Category': num_class_3_workers,
            'Remuneration per Worker in Rs.': '',
            'Signature of Receiver': ''
        })

        # Individual rows for Class 3
        for sn, worker_name in enumerate(sorted(list(unique_class_3_workers))):
            class_3_4_final_bills.append({
                'S.N.': sn + 1,
                'Name': worker_name,
                'Role': class_worker_rates['class_3_worker']['role_display'],
                'Total Students (Center-wide)': '',
                'Rate per Student (for category)': '',
                'Total Remuneration for Category (Rs.)': '',
                'Number of Workers in Category': '',
                'Remuneration per Worker in Rs.': rem_per_class_3_worker,
                'Signature of Receiver': ''
            })

    # Add a blank row for spacing
    if unique_class_3_workers and unique_class_4_workers:
        class_3_4_final_bills.append({k: '' for k in class_3_4_final_bills[0].keys()})

    # Class 4 Workers Section
    if unique_class_4_workers:
        num_class_4_workers = len(unique_class_4_workers)
        class_4_total_remuneration = total_students_for_class_workers * class_worker_rates['class_4_worker']['rate_per_student']
        rem_per_class_4_worker = class_4_total_remuneration / num_class_4_workers if num_class_4_workers > 0 else 0
        total_class_4_remuneration = class_4_total_remuneration
        
        # Summary row for Class 4
        class_3_4_final_bills.append({
            'S.N.': 'Class 4 Workers',
            'Name': '',
            'Role': '',
            'Total Students (Center-wide)': total_students_for_class_workers,
            'Rate per Student (for category)': class_worker_rates['class_4_worker']['rate_per_student'],
            'Total Remuneration for Category (Rs.)': class_4_total_remuneration,
            'Number of Workers in Category': num_class_4_workers,
            'Remuneration per Worker in Rs.': '',
            'Signature of Receiver': ''
        })

        # Individual rows for Class 4
        for sn, worker_name in enumerate(sorted(list(unique_class_4_workers))):
            class_3_4_final_bills.append({
                'S.N.': sn + 1,
                'Name': worker_name,
                'Role': class_worker_rates['class_4_worker']['role_display'],
                'Total Students (Center-wide)': '',
                'Rate per Student (for category)': '',
                'Total Remuneration for Category (Rs.)': '',
                'Number of Workers in Category': '',
                'Remuneration per Worker in Rs.': rem_per_class_4_worker,
                'Signature of Receiver': ''
            })

    # Final Total Row
    if class_3_4_final_bills:
        total_students_for_all = total_students_for_class_workers
        total_rate_per_student = class_worker_rates['class_3_worker']['rate_per_student'] + class_worker_rates['class_4_worker']['rate_per_student']
        total_remuneration_for_category = total_class_3_remuneration + total_class_4_remuneration
        total_num_workers = len(unique_class_3_workers) + len(unique_class_4_workers)
        
        class_3_4_final_bills.append({
            'S.N.': 'TOTAL',
            'Name': '',
            'Role': '',
            'Total Students (Center-wide)': total_students_for_all,
            'Rate per Student (for category)': total_rate_per_student,
            'Total Remuneration for Category (Rs.)': total_remuneration_for_category,
            'Number of Workers in Category': total_num_workers,
            'Remuneration per Worker in Rs.': total_remuneration_for_category,
            'Signature': ''
        })
        
    df_class_3_4_final_bills = pd.DataFrame(class_3_4_final_bills)
    
    return df_individual_bills, df_role_summary_matrix, df_class_3_4_final_bills


def generate_role_summary_matrix_by_date(df_detailed_remuneration, remuneration_rules, prep_closing_assignments, holiday_dates, manual_rates, selected_classes_for_bill, timetable_df, assigned_seats_df):
    """
    Generates a daily summary matrix of remuneration for all roles, including
    prep and closing days, in the requested format with 'Count (Remuneration)'.
    This version corrects the duplication and NaN issues and includes debugging prints.
    """
    summary_data = []

    # Debug print statements
    print("--- Debugging Role Summary Matrix ---")
    print(f"Selected classes for bill (for summary matrix): {selected_classes_for_bill}")
    
    if selected_classes_for_bill:
        papers_for_selected_classes = timetable_df[timetable_df['Class'].isin(selected_classes_for_bill)]['Paper Code'].unique()
        filtered_assigned_seats = assigned_seats_df[assigned_seats_df['Paper Code'].isin(papers_for_selected_classes)]
        total_students_for_class_workers = filtered_assigned_seats['Roll Number'].nunique()
        print(f"Students counted for selected classes: {total_students_for_class_workers}")
    else:
        total_students_for_class_workers = assigned_seats_df['Roll Number'].nunique()
        papers_for_selected_classes = timetable_df['Paper Code'].unique()
        filtered_assigned_seats = assigned_seats_df[assigned_seats_df['Paper Code'].isin(papers_for_selected_classes)]
        print(f"No selected classes, total students for summary: {total_students_for_class_workers}")
    
    prep_closing_remuneration_aggregated = {}
    for name, assignments in prep_closing_assignments.items():
        role = assignments.get('role')
        if role and role in remuneration_rules:
            rate = remuneration_rules[role]['rate']
            prep_days = assignments.get('prep_days', [])
            closing_days = assignments.get('closing_days', [])

            for day in prep_days + closing_days:
                if day not in prep_closing_remuneration_aggregated:
                    prep_closing_remuneration_aggregated[day] = {
                        'type': 'Pre Exam Preparation' if day in prep_days else 'Post Exam Closing',
                        'roles': {}
                    }
                if role not in prep_closing_remuneration_aggregated[day]['roles']:
                    prep_closing_remuneration_aggregated[day]['roles'][role] = {'count': 0, 'rem': 0}
                prep_closing_remuneration_aggregated[day]['roles'][role]['count'] += 1
                prep_closing_remuneration_aggregated[day]['roles'][role]['rem'] += rate

    for day, data in prep_closing_remuneration_aggregated.items():
        row_data = {
            'date & shift': f"{day} ({data['type']})",
            'Paper': data['type'],
            'Number of students': total_students_for_class_workers,
            'Conveyance': 0,
            'Daily Total': 0,
            'SCS': '0 (0)',
            'CS': '0 (0)',
            'ACS': '0 (0)',
            'PI/API': '0 (0)',
            'Invigilators': '0 (0)'
        }
        
        daily_total_rem = 0
        for role, values in data['roles'].items():
            if role == 'permanent_invigilator' or role == 'assistant_permanent_invigilator':
                row_data['PI/API'] = f"{values['count']} ({int(values['rem'])})"
            elif role == 'senior_center_superintendent':
                 row_data['SCS'] = f"{values['count']} ({int(values['rem'])})"
            elif role == 'center_superintendent':
                 row_data['CS'] = f"{values['count']} ({int(values['rem'])})"
            elif role == 'assistant_center_superintendent':
                 row_data['ACS'] = f"{values['count']} ({int(values['rem'])})"
            daily_total_rem += values['rem']
        
        if day in holiday_dates:
            total_workers = sum(d['count'] for d in data['roles'].values())
            holiday_conveyance = total_workers * manual_rates.get('holiday_conveyance_allowance_rate', 0)
            row_data['Conveyance'] = int(holiday_conveyance)
            daily_total_rem += holiday_conveyance

        row_data['Daily Total'] = int(daily_total_rem)
        summary_data.append(row_data)

    exam_dates_from_df = set(df_detailed_remuneration['date'].unique())
    prep_closing_dates_from_df = set(prep_closing_remuneration_aggregated.keys())
    exam_dates_to_process = sorted(list(exam_dates_from_df - prep_closing_dates_from_df))

    for date_str in exam_dates_to_process:
        date_assignments = df_detailed_remuneration[df_detailed_remuneration['date'] == date_str]
        
        if selected_classes_for_bill:
            date_assignments = date_assignments[date_assignments['Is_Selected_Exam'] == True]

        if date_assignments.empty:
            continue

        for shift in ['Morning', 'Evening']:
            shift_data = date_assignments[date_assignments['shift'] == shift]
            
            if shift_data.empty:
                continue

            session_papers = timetable_df[(timetable_df['date'] == date_str) & (timetable_df['shift'] == shift)]
            paper_list = []
            for _, paper_row in session_papers.iterrows():
                if not selected_classes_for_bill or paper_row['Class'] in selected_classes_for_bill:
                    paper_list.append(f"{paper_row['Paper Name']} ({paper_row['Paper Code']})")
            papers_string = ", ".join(paper_list) if paper_list else 'N/A'

            current_session_papers_codes = session_papers['Paper Code'].unique()
            students_count = assigned_seats_df[assigned_seats_df['Paper Code'].isin(current_session_papers_codes)]['Roll Number'].nunique()

            remuneration_summary = {
                'SCS': {'count': 0, 'total_rem': 0},
                'CS': {'count': 0, 'total_rem': 0},
                'ACS': {'count': 0, 'total_rem': 0},
                'PI/API': {'count': 0, 'total_rem': 0},
                'Invigilators': {'count': 0, 'total_rem': 0}
            }
            daily_total_rem = 0
            total_conveyance = 0
            
            for _, person_row in shift_data.iterrows():
                role_key = person_row['Role_Key']
                remuneration = person_row['Base_Remuneration_Per_shift_Unfiltered']
                total_conveyance += person_row['Conveyance']

                if role_key == 'permanent_invigilator' or role_key == 'assistant_permanent_invigilator':
                    remuneration_summary['PI/API']['count'] += 1
                    remuneration_summary['PI/API']['total_rem'] += remuneration
                elif role_key == 'senior_center_superintendent':
                    # SCS is paid per day, not per shift, so we only add their remuneration once per date
                    if date_str not in [s['date & shift'].split(' ')[0] for s in summary_data if 'SCS' in s]: # Check if SCS already added for this day
                        remuneration_summary['SCS']['count'] += 1
                        remuneration_summary['SCS']['total_rem'] += remuneration_rules['senior_center_superintendent']['rate']
                elif role_key == 'center_superintendent':
                    remuneration_summary['CS']['count'] += 1
                    remuneration_summary['CS']['total_rem'] += remuneration
                elif role_key == 'assistant_center_superintendent':
                    remuneration_summary['ACS']['count'] += 1
                    remuneration_summary['ACS']['total_rem'] += remuneration
                elif role_key == 'invigilator':
                    remuneration_summary['Invigilators']['count'] += 1
                    remuneration_summary['Invigilators']['total_rem'] += remuneration
            
            daily_total_rem = sum(r['total_rem'] for r in remuneration_summary.values()) + total_conveyance
            
            row_data = {
                'date & shift': f"{date_str} ({shift})",
                'Paper': papers_string,
                'Number of students': students_count,
                'SCS': f"{remuneration_summary['SCS']['count']} ({int(remuneration_summary['SCS']['total_rem'])})",
                'CS': f"{remuneration_summary['CS']['count']} ({int(remuneration_summary['CS']['total_rem'])})",
                'ACS': f"{remuneration_summary['ACS']['count']} ({int(remuneration_summary['ACS']['total_rem'])})",
                'PI/API': f"{remuneration_summary['PI/API']['count']} ({int(remuneration_summary['PI/API']['total_rem'])})",
                'Invigilators': f"{remuneration_summary['Invigilators']['count']} ({int(remuneration_summary['Invigilators']['total_rem'])})",
                'Conveyance': int(total_conveyance),
                'Daily Total': int(daily_total_rem)
            }
            summary_data.append(row_data)

    df_summary = pd.DataFrame(summary_data)
    
    if df_summary.empty:
        return pd.DataFrame()

    # CORRECTED LINE: Using regex to extract date and handling errors
    df_summary['date_sort'] = pd.to_datetime(df_summary['date & shift'].str.extract(r'(\d{2}-\d{2}-\d{4})')[0], format='%d-%m-%Y', errors='coerce')
    df_summary = df_summary.sort_values('date_sort').drop(columns='date_sort').reset_index(drop=True)

    total_students = df_summary['Number of students'].sum()
    total_conveyance = df_summary['Conveyance'].sum()
    total_daily_total = df_summary['Daily Total'].sum()

    role_totals = {role: {'count': 0, 'rem': 0} for role in ['SCS', 'CS', 'ACS', 'PI/API', 'Invigilators']}

    for _, row in df_summary.iterrows():
        for role in role_totals.keys():
            if isinstance(row[role], str) and '(' in row[role]:
                try:
                    count_str, rem_str = row[role].strip().split(' ')
                    role_totals[role]['count'] += int(count_str)
                    role_totals[role]['rem'] += int(rem_str.strip('()'))
                except (ValueError, IndexError):
                    # Handle cases where the format is unexpected
                    pass

    total_row = {
        'date & shift': 'Total',
        'Paper': '',
        'Number of students': total_students,
        'SCS': f"{role_totals['SCS']['count']} ({int(role_totals['SCS']['rem'])})",
        'CS': f"{role_totals['CS']['count']} ({int(role_totals['CS']['rem'])})",
        'ACS': f"{role_totals['ACS']['count']} ({int(role_totals['ACS']['rem'])})",
        'PI/API': f"{role_totals['PI/API']['count']} ({int(role_totals['PI/API']['rem'])})",
        'Invigilators': f"{role_totals['Invigilators']['count']} ({int(role_totals['Invigilators']['rem'])})",
        'Conveyance': int(total_conveyance),
        'Daily Total': int(total_daily_total)
    }

    df_summary = pd.concat([df_summary, pd.DataFrame([total_row])], ignore_index=True)

    final_cols = ['date & shift', 'Paper', 'Number of students', 'SCS', 'CS', 'ACS', 'PI/API', 'Invigilators', 'Conveyance', 'Daily Total']
    return df_summary[final_cols]
def add_total_row(df):
    """Add a total row to the dataframe"""
    if df.empty:
        return df
    
    total_row = {}
    for col in df.columns:
        if col in ['SN', 'S.N.']:
            total_row[col] = 'TOTAL'
        elif col in ['Name (with role)', 'Name', 'Role', 'Duty dates', 'shift (morning/evening)', 'Signature', 'Signature of Receiver',
                      'Duty dates of selected class exam shift (morning)', 'Duty dates of selected class exam shift (evening)', 'date & shift']:
            total_row[col] = ''
        elif df[col].dtype in ['int64', 'float64']:
            total_row[col] = df[col].sum()
        else:
            total_row[col] = ''
    
    total_df = pd.DataFrame([total_row])
    return pd.concat([df, total_df], ignore_index=True)

def save_bills_to_excel(individual_bills_df, role_summary_df, class_workers_df, filename="remuneration_bills.xlsx"):
    """
    Saves the three remuneration dataframes into a single Excel file with multiple sheets.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not individual_bills_df.empty:
            individual_bills_df.to_excel(writer, sheet_name='Individual Bills', index=False)
            # Auto-adjust column width for individual bills
            worksheet = writer.sheets['Individual Bills']
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter # Get the column name
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_name].width = adjusted_width

        if not role_summary_df.empty:
            role_summary_df.to_excel(writer, sheet_name='Role Summary', index=False)
            # Auto-adjust column width for role summary
            worksheet = writer.sheets['Role Summary']
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_name].width = adjusted_width

        if not class_workers_df.empty:
            class_workers_df.to_excel(writer, sheet_name='Class Workers', index=False)
            # Auto-adjust column width for class workers
            worksheet = writer.sheets['Class Workers']
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_name].width = adjusted_width
    
    output.seek(0)
    return output, filename



# Main app
st.title("Government Law College, Morena (M.P.) Examination Management System")

menu = st.radio("Select Module", ["Student View", "Admin Panel", "Centre Superintendent Panel"])

if menu == "Student View":
    sitting_plan, timetable, assigned_seats_df, attestation_df = load_data()

    # Check if dataframes are empty, indicating files were not loaded
    if sitting_plan.empty or timetable.empty:
        st.warning("Sitting plan or timetable data not found. Please upload them via the Admin Panel for full functionality.")
    
    option = st.radio("Choose Search Option:", [
        "Search by Roll Number and date",
        "Get Full Exam Schedule by Roll Number",
        "View Full Timetable"
    ])

    #
# REPLACE your Streamlit code block with this one
#
    if option == "Search by Roll Number and date":
        roll = st.text_input("Enter Roll Number", max_chars=9)
        date_input = st.date_input("Enter Exam date", value=datetime.date.today())

        if st.button("Search"):
            # --- FIX: Load the CORRECT DataFrames from session state ---
            # We need 'assigned_seats_df' and 'timetable', NOT 'sitting_plan'
            assigned_seats_df = st.session_state.get('assigned_seats_df', pd.DataFrame())
            timetable = st.session_state.get('timetable', pd.DataFrame())

            if assigned_seats_df.empty or timetable.empty:
                st.warning("Assigned seats or timetable data is missing. Please upload them via the Admin Panel.")
            
            else:
                # --- FIX: Call the NEW function with the CORRECT arguments ---
                results = get_student_exam_details(
                    roll, 
                    date_input.strftime('%d-%m-%Y'), 
                    assigned_seats_df, 
                    timetable
                )
                
                if results:
                    st.success(f"Found {len(results)} exam(s) for Roll Number {roll} on {date_input.strftime('%d-%m-%Y')}:")
                    for i, result in enumerate(results):
                        st.markdown(f"---")
                        
                        
                        # This will now display the correct data, not 'nan'
                        st.write(f"**Room Number:** {result['Room Number']}")
                        st.write(f"**🪑 Seat Number:** {result['Seat Number']}")
                        
                        # Format paper string
                        paper_display = f"{result['Paper Name']} ({result['Paper Code']})"
                        if result['Paper']:
                            paper_display = f"{result['Paper']} - {paper_display}"
                        
                        st.write(f"**📚 Paper:** {paper_display}")
                        
                        st.write(f"**🎓 Student type:** {result['Mode']} - {result['Type']}")
                        st.write(f"**🕐 shift:** {result['shift']}, **📅 date:** {result['date']}")
                else:
                    st.warning("No data found for the given inputs. Check that the Roll Number and date are correct and data is loaded.")
    elif option == "Get Full Exam Schedule by Roll Number":
        roll = st.text_input("Enter Roll Number")
        if st.button("Get Schedule"):
            if sitting_plan.empty or timetable.empty:
                st.warning("Sitting plan or timetable data is missing. Please upload them via the Admin Panel to get schedule.")
            else:
                schedule = pd.DataFrame(get_all_exams(roll, sitting_plan, timetable))
                if not schedule.empty:
                    schedule['date_dt'] = pd.to_datetime(schedule['date'], format='%d-%m-%Y', errors='coerce')
                    schedule = schedule.sort_values(by="date_dt").drop(columns=['date_dt'])
                    st.write(schedule)
                else:
                    st.warning("No exam records found for this roll number.")
    
    elif option == "View Full Timetable":
        st.subheader("Full Examination Timetable")
        if timetable.empty:
            st.warning("Timetable data is missing. Please upload it via the Admin Panel.")
        else:
            st.dataframe(timetable)

elif menu == "Admin Panel":
    st.subheader("🔐 Admin Login")
    if admin_login():
        st.success("Login successful!")
        
        # Load data here, inside the successful login block
        sitting_plan, timetable, assigned_seats_df, attestation_df = load_data()
        
        st.markdown("---")
        st.subheader("Current Data Previews")
        col_sp, col_tt, col_assigned, col_attestation = st.columns(4) # Added a column for assigned_seats and attestation
        with col_sp:
            st.write(f"**{SITTING_PLAN_FILE}**")
            if not sitting_plan.empty:
                st.dataframe(sitting_plan)
            else:
                st.info("No sitting plan data loaded.")
        with col_tt:
            st.write(f"**{TIMETABLE_FILE}**")
            if not timetable.empty:
                st.dataframe(timetable)
            else:
                st.info("No timetable data loaded.")
        with col_assigned: # Display assigned_seats.csv
            st.write(f"**{ASSIGNED_SEATS_FILE}**")
            if not assigned_seats_df.empty:
                st.dataframe(assigned_seats_df)
            else:
                st.info("No assigned seats data loaded.")
        with col_attestation: # Display attestation_data_combined.csv
            st.write(f"**{ATTESTATION_DATA_FILE}**")
            if not attestation_df.empty:
                st.dataframe(attestation_df)
            else:
                st.info("No attestation data loaded.")

        st.markdown("---") # Separator

        # Admin Panel Options
        admin_option = st.radio("Select Admin Task:", [
            "Get All Students for date & shift (Room Wise)",
            "Get All Students for date & shift (Roll Number Wise)",
            "Update Timetable Details",
            "Assign Rooms & Seats to Students",
            "Room Occupancy Report",
            "Room Chart Report",
            "Data Processing & Reports",
            "Remuneration Bill Generation",
            "Report Panel",
            "Download Attestation Data" # Added new option for direct download
        ])

        if admin_option == "Download Attestation Data":
            st.subheader("⬇️ Download Attestation Data")
            st.info(f"Click the button below to download '{ATTESTATION_DATA_FILE}' from Supabase to the parent folder of this application.")
            if st.button("Download Attestation Data"):
                download_attestation_data_to_parent_folder()
        
        elif admin_option == "Get All Students for date & shift (Room Wise)":
            st.subheader("List All Students for a date and shift (Room Wise)")
            if assigned_seats_df.empty or timetable.empty: # Changed from sitting_plan to assigned_seats_df
                st.info("Please ensure seats are assigned and 'timetable.csv' is uploaded to use this feature.")
            else:
                list_date_input = st.date_input("Select date", value=datetime.date.today())
                list_shift_options = ["Morning", "Evening"]
                list_shift = st.selectbox("Select shift", list_shift_options)
                
                if st.button("Get Student List (Room Wise)"):
                    formatted_student_list_text, error_message, excel_data_for_students_list = get_all_students_for_date_shift_formatted(
                        list_date_input.strftime('%d-%m-%Y'),
                        list_shift,
                        assigned_seats_df, # Pass assigned_seats_df
                        timetable
                    )
                    if formatted_student_list_text:
                        st.success(f"Generated list for {list_date_input.strftime('%d-%m-%Y')} ({list_shift} shift):")
                        st.text_area("Student List (Text Format)", formatted_student_list_text, height=500)
                        
                        # Download button for TXT
                        file_name_txt = (
                            f"all_students_list_room_wise_{list_date_input.strftime('%Y%m%d')}_"
                            f"{list_shift.lower()}.txt"
                        )
                        st.download_button(
                            label="Download Student List (Room Wise) as TXT",
                            data=formatted_student_list_text,
                            file_name=file_name_txt,
                            mime="text/plain"
                        )

                        # Download button for Excel
                        if excel_data_for_students_list:
                            output = io.BytesIO()
                            workbook = Workbook()
                            sheet = workbook.active
                            sheet.title = "Student List (Room Wise)"

                            for row_data in excel_data_for_students_list:
                                sheet.append(row_data)

                            for col_idx, col_cells in enumerate(sheet.columns):
                                max_length = 0
                                for cell in col_cells:
                                    try:
                                        if cell.value is not None:
                                            cell_value_str = str(cell.value)
                                            current_length = max(len(line) for line in cell_value_str.split('\n'))
                                            if current_length > max_length:
                                                max_length = current_length
                                    except Exception as e:
                                        st.error(f"Error processing cell: {e}")
                                        pass
                            adjusted_width = (max_length + 2)
                            sheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width

                            workbook.save(output)
                            processed_data = output.getvalue()

                            file_name_excel = (
                                f"all_students_list_room_wise_{list_date_input.strftime('%Y%m%d')}_"
                                f"{list_shift.lower()}.xlsx"
                            )
                            st.download_button(
                                label="Download Student List (Room Wise) as Excel",
                                data=processed_data,
                                file_name=file_name_excel,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    else:
                        st.warning(f"No students found: {error_message}")

        elif admin_option == "Get All Students for date & shift (Roll Number Wise)":
            st.subheader("List All Students for a date and shift (Roll Number Wise)")
            if assigned_seats_df.empty or timetable.empty: # Changed from sitting_plan to assigned_seats_df
                st.info("Please ensure seats are assigned and 'timetable.csv' is uploaded to use this feature.")
            else:
                list_date_input = st.date_input("Select date", value=datetime.date.today(), key="roll_num_wise_date")
                list_shift_options = ["Morning", "Evening"]
                list_shift = st.selectbox("Select shift", list_shift_options, key="roll_num_wise_shift")
                
                if st.button("Get Student List (Roll Number Wise)"):
                    formatted_student_list_text, error_message, excel_data_for_students_list = get_all_students_roll_number_wise_formatted(
                        list_date_input.strftime('%d-%m-%Y'),
                        list_shift,
                        assigned_seats_df, # Pass assigned_seats_df
                        timetable
                    )
                    if formatted_student_list_text:
                        st.success(f"Generated list for {list_date_input.strftime('%d-%m-%Y')} ({list_shift} shift):")
                        st.text_area("Student List (Text Format)", formatted_student_list_text, height=500)
                        
                        # Download button for TXT
                        file_name_txt = (
                            f"all_students_list_roll_wise_{list_date_input.strftime('%Y%m%d')}_"
                            f"{list_shift.lower()}.txt"
                        )
                        st.download_button(
                            label="Download Student List (Roll Number Wise) as TXT",
                            data=formatted_student_list_text,
                            file_name=file_name_txt,
                            mime="text/plain"
                        )

                        # Download button for Excel
                        if excel_data_for_students_list:
                            output = io.BytesIO()
                            workbook = Workbook()
                            sheet = workbook.active
                            sheet.title = "Student List (Roll Wise)"

                            for row_data in excel_data_for_students_list:
                                sheet.append(row_data)

                            for col_idx, col_cells in enumerate(sheet.columns):
                                max_length = 0
                                for cell in col_cells:
                                    try:
                                        if cell.value is not None:
                                            cell_value_str = str(cell.value)
                                            current_length = max(len(line) for line in cell_value_str.split('\n'))
                                            if current_length > max_length:
                                                    max_length = current_length
                                    except Exception as e:
                                        st.error(f"Error processing cell: {e}")
                                        pass
                            adjusted_width = (max_length + 2)
                            sheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width

                            workbook.save(output)
                            processed_data = output.getvalue()

                            file_name_excel = (
                                f"all_students_list_roll_wise_{list_date_input.strftime('%Y%m%d')}_"
                                f"{list_shift.lower()}.xlsx"
                            )
                            st.download_button(
                                label="Download Student List (Roll Number Wise) as Excel",
                                data=processed_data,
                                file_name=file_name_excel,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    else:
                        st.warning(f"No students found: {error_message}")

        elif admin_option == "Update Timetable Details":
            st.subheader("✏️ Update Timetable Details")
            if timetable.empty:
                st.info("No timetable data loaded. Please upload 'timetable.csv' first using the 'Upload Data Files' section.")
            else:
                st.write("Current Timetable Preview:")
                st.dataframe(timetable)

                st.markdown("---")
                st.write("Select filters to specify which entries to update:")
                
                # Filters for selecting entries to update
                unique_dates_tt = sorted(timetable['date'].astype(str).unique().tolist())
                unique_shifts_tt = sorted(timetable['shift'].astype(str).unique().tolist())
                unique_classes_tt = sorted(timetable['Class'].astype(str).unique().tolist())
                unique_paper_codes_tt = sorted(timetable['Paper Code'].astype(str).unique().tolist())
                unique_paper_tt = sorted(timetable['Paper'].astype(str).unique().tolist())
                unique_paper_names_tt = sorted(timetable['Paper Name'].astype(str).unique().tolist())

                filter_date_tt_update = st.selectbox("Filter by date", ["All"] + unique_dates_tt, key="filter_date_tt_update")
                filter_shift_tt_update = st.selectbox("Filter by shift", ["All"] + unique_shifts_tt, key="filter_shift_tt_update")
                filter_class_tt_update = st.selectbox("Filter by Class", ["All"] + unique_classes_tt, key="filter_class_tt_update")
                filter_paper_code_tt_update = st.selectbox("Filter by Paper Code", ["All"] + unique_paper_codes_tt, key="filter_paper_code_tt_update")
                filter_paper_tt_update = st.selectbox("Filter by Paper", ["All"] + unique_paper_tt, key="filter_paper_tt_update")
                filter_paper_name_tt_update = st.selectbox("Filter by Paper Name", ["All"] + unique_paper_names_tt, key="filter_paper_name_tt_update")

                st.markdown("---")
                st.write("Entries that will be updated based on your filters:")
                
                temp_filtered_tt = timetable.copy()
                if filter_date_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['date'].astype(str) == filter_date_tt_update]
                if filter_shift_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['shift'].astype(str) == filter_shift_tt_update]
                if filter_class_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Class'].astype(str) == filter_class_tt_update]
                if filter_paper_code_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Paper Code'].astype(str) == filter_paper_code_tt_update]
                if filter_paper_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Paper'].astype(str) == filter_paper_tt_update]
                if filter_paper_name_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Paper Name'].astype(str) == filter_paper_name_tt_update]
                
                if temp_filtered_tt.empty:
                    st.info("No entries match the selected filters. No updates will be applied.")
                else:
                    st.dataframe(temp_filtered_tt)

                st.markdown("---")
                st.write("Enter new values for 'date', 'shift', and 'Time' for the filtered entries:")
                
                # Provide default values from the first row of the *filtered* timetable if available, otherwise from the full timetable or current date/time
                default_date_update_input = datetime.date.today()
                if not temp_filtered_tt.empty and 'date' in temp_filtered_tt.columns and pd.notna(temp_filtered_tt['date'].iloc[0]):
                    try:
                        default_date_update_input = datetime.datetime.strptime(str(temp_filtered_tt['date'].iloc[0]).strip(), '%d-%m-%Y').date()
                    except ValueError:
                        pass
                elif 'date' in timetable.columns and not timetable['date'].empty and pd.notna(timetable['date'].iloc[0]):
                    try:
                        default_date_update_input = datetime.datetime.strptime(str(timetable['date'].iloc[0]).strip(), '%d-%m-%Y').date()
                    except ValueError:
                        pass


                default_shift_update_input = "Morning"
                if not temp_filtered_tt.empty and 'shift' in temp_filtered_tt.columns and pd.notna(temp_filtered_tt['shift'].iloc[0]):
                    default_shift_update_input = str(temp_filtered_tt['shift'].iloc[0]).strip()
                elif 'shift' in timetable.columns and not timetable['shift'].empty and pd.notna(timetable['shift'].iloc[0]):
                    default_shift_update_input = str(timetable['shift'].iloc[0]).strip()


                default_time_update_input = "09:00 AM - 12:00 PM"
                if not temp_filtered_tt.empty and 'Time' in temp_filtered_tt.columns and pd.notna(temp_filtered_tt['Time'].iloc[0]):
                    default_time_update_input = str(temp_filtered_tt['Time'].iloc[0]).strip()
                elif 'Time' in timetable.columns and not timetable['Time'].empty and pd.notna(timetable['Time'].iloc[0]):
                    default_time_update_input = str(timetable['Time'].iloc[0]).strip()


                update_date = st.date_input("New date", value=default_date_update_input, key="update_tt_date")
                update_shift = st.selectbox("New shift", ["Morning", "Evening"], index=["Morning", "Evening"].index(default_shift_update_input) if default_shift_update_input in ["Morning", "Evening"] else 0, key="update_tt_shift")
                update_time = st.text_input("New Time (e.g., 09:00 AM - 12:00 PM)", value=default_time_update_input, key="update_tt_time")

                if st.button("Apply Updates and Save Timetable"):
                    if temp_filtered_tt.empty:
                        st.warning("No entries matched your filters, so no updates were applied.")
                    else:
                        timetable_modified = timetable.copy()
                        
                        # Identify indices to update in the original DataFrame
                        indices_to_update = timetable_modified[
                            (timetable_modified['date'].astype(str) == filter_date_tt_update if filter_date_tt_update != "All" else True) &
                            (timetable_modified['shift'].astype(str) == filter_shift_tt_update if filter_shift_tt_update != "All" else True) &
                            (timetable_modified['Class'].astype(str) == filter_class_tt_update if filter_class_tt_update != "All" else True) &
                            (timetable_modified['Paper Code'].astype(str) == filter_paper_code_tt_update if filter_paper_code_tt_update != "All" else True) &                           (timetable_modified['Paper'].astype(str) == filter_paper_tt_update if filter_paper_tt_update != "All" else True) &
                            (timetable_modified['Paper Name'].astype(str) == filter_paper_name_tt_update if filter_paper_name_tt_update != "All" else True)
                        ].index

                        # Apply updates only to the identified rows
                        if not indices_to_update.empty:
                            timetable_modified.loc[indices_to_update, 'date'] = update_date.strftime('%d-%m-%Y')
                            timetable_modified.loc[indices_to_update, 'shift'] = update_shift
                            timetable_modified.loc[indices_to_update, 'Time'] = update_time

                            success, msg = save_uploaded_file(timetable_modified, TIMETABLE_FILE)
                            if success:
                                st.success(f"Timetable details updated for {len(indices_to_update)} entries and saved successfully.")
                                # Reload data to reflect changes in the app
                                sitting_plan, timetable, assigned_seats_data, attestation_data = load_data()
                                st.rerun()
                            else:
                                st.error(msg)
                        else:
                            st.warning("No entries matched your filters to apply updates.")

        elif admin_option == "Assign Rooms & Seats to Students": # Replaced with assign_seats_app.py logic
            st.subheader("📘 Room & Seat Assignment Tool")
            st.markdown("""
            This tool helps manage seat assignments for exams, offering real-time status updates,
            capacity warnings, and clear error messages based on your selected seat format.
            """)

            if sitting_plan.empty or timetable.empty:
                st.error(f"Error: `{SITTING_PLAN_FILE}` or `{TIMETABLE_FILE}` not found. Please upload these files to run the assignment tool.")
                st.stop() # Stop execution if critical files are missing

            # --- Session State for consistent UI updates ---
            if 'current_room_status_a_rem' not in st.session_state:
                st.session_state.current_room_status_a_rem = None
            if 'current_room_status_b_rem' not in st.session_state:
                st.session_state.current_room_status_b_rem = None
            if 'current_room_status_total_rem' not in st.session_state:
                st.session_state.current_room_status_total_rem = None

            # --- Input Widgets ---
            st.subheader("Exam Details")
            # Ensure date and shift options are available from timetable
            date_options = sorted(timetable["date"].dropna().unique())
            shift_options = sorted(timetable["shift"].dropna().unique())

            if not date_options or not shift_options:
                st.warning("Timetable is empty or missing date/shift information. Please upload a complete timetable.")
            else:
                date = st.selectbox("Select Exam date", date_options, key="assign_date_select")
                shift = st.selectbox("Select shift", shift_options, key="assign_shift_select")

                # --- NEW: Paper-wise summary chart for the selected session ---
                st.markdown("---")
                st.subheader("Session Student Summary (Assigned vs. Unassigned)")
                session_paper_summary_df = get_session_paper_summary(date, shift, sitting_plan, assigned_seats_df, timetable)
                
                if not session_paper_summary_df.empty:
                    st.dataframe(session_paper_summary_df)
                else:
                    st.info("No student data found for the selected date and shift.")
                st.markdown("---")
                # --- END NEW ---

                # Filter relevant papers based on selected date and shift
                filtered_papers = timetable[(timetable["date"] == date) & (timetable["shift"] == shift)]
                # Ensure paper codes are formatted for display
                paper_options = filtered_papers[["Paper Code", "Paper Name"]].drop_duplicates().values.tolist()
                paper_display = [f"{_format_paper_code(code)} - {name}" for code, name in paper_options]

                selected_paper = st.selectbox("Select Paper Code and Name", paper_display, key="assign_paper_select")

                # Only proceed if a paper is selected
                if selected_paper:
                    paper_code = _format_paper_code(selected_paper.split(" - ")[0]) # Format the extracted code
                    paper_name = selected_paper.split(" - ", 1)[1].strip()

                    st.subheader("Room & Seat Configuration")
                    # Added .strip() to handle potential leading/trailing spaces in room input
                    room = st.text_input("Enter Room Number (e.g., 1, G230)", key="room_input").strip()

                    # Enhanced capacity input
                    col1, col2 = st.columns(2)
                    with col1:
                        total_capacity = st.number_input("Enter Total Room Capacity (for '1 to N' format)", min_value=1, max_value=2000, value=2000, key="total_capacity_input")
                    with col2:
                        capacity_per_format = st.number_input("Capacity per Format (for 'A/B' formats)", min_value=1, max_value=100, value=30, key="capacity_per_format_input")

                    seat_format = st.radio("Select Seat Assignment Format:", ["1 to N", "1A to NA", "1B to NB"], key="seat_format_radio")

                    # --- Show current room status BEFORE assignment ---
                    if room:
                        # Get all assigned seats for the current room, date, and shift
                        room_assigned_seats_current = assigned_seats_df[
                            (assigned_seats_df["Room Number"] == room) &
                            (assigned_seats_df["date"] == date) &
                            (assigned_seats_df["shift"] == shift)
                        ]["Seat Number"].tolist()

                        # Calculate used seats for A, B, and no-suffix formats
                        a_seats_used_current = len([s for s in room_assigned_seats_current if str(s).endswith("A") and s])
                        b_seats_used_current = len([s for s in room_assigned_seats_current if str(s).endswith("B") and s])
                        no_suffix_seats_used_current = len([s for s in room_assigned_seats_current if not str(s).endswith("A") and not str(s).endswith("B")])

                        st.subheader("📊 Current Room Status")
                        if seat_format in ["1A to NA", "1B to NB"]:
                            a_remaining_current = capacity_per_format - a_seats_used_current
                            b_remaining_current = capacity_per_format - b_seats_used_current
                            st.info(f"A-format: **{a_remaining_current}** remaining ({a_seats_used_current}/{capacity_per_format} used)")
                            st.info(f"B-format: **{b_remaining_current}** remaining ({b_seats_used_current}/{capacity_per_format} used)")
                            st.session_state.current_room_status_a_rem = a_remaining_current
                            st.session_state.current_room_status_b_rem = b_remaining_current
                            st.session_state.current_room_status_total_rem = None # Clear total if A/B is selected
                        else: # 1 to N format
                            remaining_current = total_capacity - no_suffix_seats_used_current
                            st.info(f"Total: **{remaining_current}** seats remaining ({no_suffix_seats_used_current}/{total_capacity} used)")
                            st.session_state.current_room_status_total_rem = remaining_current
                            st.session_state.current_room_status_a_rem = None # Clear A/B if total is selected
                            st.session_state.current_room_status_b_rem = None


                    st.markdown("---")

                    # --- Assign Seats Button ---
                    if st.button("✅ Assign Seats", key="assign_button"):
                        if not room:
                            st.error("Please enter a Room Number before assigning seats.")
                            st.stop()

                        # Extract roll numbers for the selected paper from sitting_plan
                        roll_cols = [col for col in sitting_plan.columns if col.lower().startswith("roll number")]
                        # Ensure Paper Code is treated as string for comparison
                        paper_rows = sitting_plan[sitting_plan["Paper Code"].astype(str) == paper_code] # Use formatted paper code
                        all_rolls = paper_rows[roll_cols].values.flatten()
                        all_rolls = [str(r).strip() for r in all_rolls if str(r).strip() and str(r).lower() != 'nan']

                        # Remove previously assigned roll numbers for this paper/date/shift
                        already_assigned_rolls = assigned_seats_df[
                            (assigned_seats_df["Paper Code"].astype(str) == paper_code) & # Use formatted paper code
                            (assigned_seats_df["date"] == date) &
                            (assigned_seats_df["shift"] == shift)
                        ]["Roll Number"].astype(str).tolist()

                        unassigned_rolls = [r for r in all_rolls if r not in already_assigned_rolls]

                        if not unassigned_rolls:
                            st.warning("⚠️ All students for this paper are already assigned for this date/shift!")
                            st.stop()

                        # Determine seat format and capacity for the assignment logic
                        suffix = ""
                        format_capacity_for_assignment = 0 # Initialize

                        if seat_format == "1 to N":
                            suffix = ""
                            format_capacity_for_assignment = total_capacity
                        elif seat_format == "1A to NA":
                            suffix = "A"
                            format_capacity_for_assignment = capacity_per_format
                        elif seat_format == "1B to NB":
                            suffix = "B"
                            format_capacity_for_assignment = capacity_per_format

                        # Get a set of all *physically occupied seat keys* for the current room, date, and shift
                        occupied_physical_seat_keys = set(
                            (str(x[0]), str(x[1]), str(x[2]), str(x[3]))
                            for x in assigned_seats_df[
                                (assigned_seats_df["Room Number"] == room) &
                                (assigned_seats_df["date"] == date) &
                                (assigned_seats_df["shift"] == shift)
                            ][['Room Number', 'Seat Number', 'date', 'shift']].values
                        )

                        # Find truly available seat numbers for the selected format.
                        available_seat_numbers = []
                        for i in range(1, format_capacity_for_assignment + 1):
                            prospective_seat_string = f"{i}{suffix}"
                            prospective_seat_key = (str(room), prospective_seat_string, str(date), str(shift)) # Ensure consistency

                            # A seat is available if its specific key (Room, Seat String, date, shift) is NOT already taken
                            if prospective_seat_key not in occupied_physical_seat_keys:
                                available_seat_numbers.append(i)

                        # --- Clear Error Messages & No Automatic Format Switching ---
                        if not available_seat_numbers:
                            st.error(f"❌ ERROR: No seats available in **{seat_format}** format for Room {room}! Please manually change to a different format (e.g., '1A to NA' or '1B to NB') or room.")
                            st.stop() # Stop execution after displaying error

                        # --- Capacity Warnings ---
                        if len(available_seat_numbers) < len(unassigned_rolls):
                            st.warning(f"⚠️ Capacity Warning: Only **{len(available_seat_numbers)}** seats available in **{seat_format}** format, but **{len(unassigned_rolls)}** students need assignment.")
                            st.warning(f"💡 This will assign the first **{len(available_seat_numbers)}** students. Remaining students will need assignment in a different format or room.")

                        # Generate actual seat strings with the selected suffix
                        seats_to_assign_count = min(len(available_seat_numbers), len(unassigned_rolls))
                        assigned_seat_strings = [f"{available_seat_numbers[i]}{suffix}" for i in range(seats_to_assign_count)]

                        # Assign seats to students
                        students_to_assign = unassigned_rolls[:seats_to_assign_count]
                        assigned_rows = []

                        for i, roll in enumerate(students_to_assign):
                            seat_num_str = assigned_seat_strings[i]
                            current_assignment_key = (str(room), seat_num_str, str(date), str(shift)) # Ensure consistency

                            # Check if this specific physical seat key is already taken
                            if current_assignment_key in occupied_physical_seat_keys:
                                st.warning(f"⚠️ Conflict: Seat **{seat_num_str}** in Room **{room}** is already assigned for this date/shift. Skipping assignment for Roll Number **{roll}**.")
                            else:
                                assigned_rows.append({
                                    "Roll Number": roll,
                                    "Paper Code": paper_code, # Keep as string
                                    "Paper Name": paper_name,
                                    "Room Number": room,
                                    "Seat Number": seat_num_str,
                                    "date": date,
                                    "shift": shift
                                })
                                # Add this new assignment's physical seat key to our occupied set for this batch
                                occupied_physical_seat_keys.add(current_assignment_key) # Update the set for subsequent assignments in this batch

                        new_assignments_df = pd.DataFrame(assigned_rows)

                        if new_assignments_df.empty:
                            st.warning("No new unique seats could be assigned in this attempt, possibly due to conflicts with existing assignments.")
                            # st.stop() # Removed st.stop() to allow further interaction
                        else:
                            # Merge new assignments with existing ones and save
                            assigned_seats_df = pd.concat([assigned_seats_df, new_assignments_df], ignore_index=True)
                            # Re-add drop_duplicates on Roll Number/Paper Code/date/shift to prevent a student
                            # from being assigned the *same paper* multiple times if the button is clicked repeatedly.
                            assigned_seats_df.drop_duplicates(subset=["Roll Number", "Paper Code", "date", "shift"], inplace=True)
                            
                            success, msg = save_uploaded_file(assigned_seats_df, ASSIGNED_SEATS_FILE)
                            if success:
                                st.success(f"✅ Successfully assigned **{len(new_assignments_df)}** students to Room **{room}** using **{seat_format}** format.")
                                st.dataframe(new_assignments_df) # Display only the newly assigned students
                            else:
                                st.error(f"Error saving assigned seats: {msg}")

                            # --- Display Updated Room Status AFTER assignment ---
                            st.subheader("📊 Updated Room Status")
                            updated_room_assigned_seats = assigned_seats_df[
                                (assigned_seats_df["Room Number"] == room) &
                                (assigned_seats_df["date"] == date) &
                                (assigned_seats_df["shift"] == shift)
                            ]["Seat Number"].tolist()

                            updated_a_seats_used = len([s for s in updated_room_assigned_seats if s.endswith("A")])
                            updated_b_seats_used = len([s for s in updated_room_assigned_seats if s.endswith("B")])
                            updated_no_suffix_seats_used = len([s for s in updated_room_assigned_seats if not s.endswith("A") and not s.endswith("B")])

                            if seat_format in ["1A to NA", "1B to NB"]:
                                updated_a_remaining = capacity_per_format - updated_a_seats_used
                                updated_b_remaining = capacity_per_format - updated_b_seats_used
                                st.info(f"A-format: **{updated_a_remaining}** remaining ({updated_a_seats_used}/{capacity_per_format} used)")
                                st.info(f"B-format: **{updated_b_remaining}** remaining ({updated_b_seats_used}/{capacity_per_format} used)")
                            else: # 1 to N format
                                updated_remaining = total_capacity - updated_no_suffix_seats_used
                                st.info(f"Total: **{updated_remaining}** seats remaining ({updated_no_suffix_seats_used}/{total_capacity} used)")

                            if len(new_assignments_df) < len(unassigned_rolls):
                                remaining_students_after_assignment = len(unassigned_rolls) - len(new_assignments_df)
                                st.warning(f"⚠️ **{remaining_students_after_assignment}** students from this paper still need assignment. Please run assignment again, potentially with a different format or room.")
                            
                            st.rerun() # Rerun to refresh the dataframes and status

                    st.markdown("---")

                    # --- Display all assignments for the selected room/date/shift ---
                    if room:
                        with st.expander(f"📄 View all current assignments for Room {room} on {date} ({shift})"):
                            room_assignments_display = assigned_seats_df[
                                (assigned_seats_df["Room Number"] == room) &
                                (assigned_seats_df["date"] == date) &
                                (assigned_seats_df["shift"] == shift)
                            ].copy() # Use .copy() to avoid SettingWithCopyWarning

                            if room_assignments_display.empty:
                                st.info("No assignments yet for this room, date, and shift.")
                            else:
                                # Proper sorting for seat numbers (e.g., 1A, 2A, ..., 10A, 1B, 2B)
                                def sort_seat_number(seat):
                                    if isinstance(seat, str):
                                        if seat.endswith('A'):
                                            return (0, int(seat[:-1])) # Group A seats first
                                        elif seat.endswith('B'):
                                            return (1, int(seat[:-1])) # Group B seats second
                                        elif seat.isdigit():
                                            return (2, int(seat)) # Group 1 to N seats last
                                    return (3, seat) # For any unexpected format, put at the end

                                room_assignments_display['sort_key'] = room_assignments_display['Seat Number'].apply(sort_seat_number)
                                room_assignments_sorted = room_assignments_display.sort_values(by='sort_key').drop('sort_key', axis=1)
                                st.dataframe(room_assignments_sorted, use_container_width=True)

            # --- Reset Button (outside the paper selection block for broader access) ---
            st.markdown("---")
            st.subheader("Maintenance")
            if st.button("🔄 Reset All Assigned Seats (Clear assigned_seats.csv)", key="reset_button"):
                if os.path.exists(ASSIGNED_SEATS_FILE):
                    os.remove(ASSIGNED_SEATS_FILE)
                    st.success("`assigned_seats.csv` has been deleted. All assignments reset.")
                else:
                    st.info("No `assigned_seats.csv` found to reset.")
                st.rerun() # Rerun the app to reflect the changes

        elif admin_option == "Room Occupancy Report": 
            display_room_occupancy_report(sitting_plan, assigned_seats_df, timetable)
            
            st.markdown("---")
            st.subheader("💾 Database Backup & Restore")
            
            if st.button("🚀 Start (Download All Tables as CSVs)"):
                with st.spinner("Downloading all Supabase tables to CSV files..."):
                    # Table to CSV filename mapping
                    # UPDATED: Added 'prep_closing_assignments' and 'global_settings'
                    table_csv_mapping = {
                        "timetable": "timetable.csv",
                        "sitting_plan": "sitting_plan.csv",
                        "assigned_seats": "assigned_seats.csv", 
                        "exam_team_members": "exam_team_members.csv",
                        "shift_assignments": "shift_assignments.csv",
                        "room_invigilator_assignments": "room_invigilator_assignments.csv",
                        "cs_reports": "cs_reports.csv",
                        "attestation_data_combined": "attestation_data_combined.csv",
                        "prep_closing_assignments": "prep_closing_assignments.csv",
                        "global_settings": "global_settings.csv"
                    }
                    
                    st.markdown("### 📥 Downloading all Supabase tables to CSV files...")
                    download_success = True
                    
                    for table_name, csv_filename in table_csv_mapping.items():
                        # For attestation, we need to handle the parent folder path
                        if table_name == "attestation_data_combined":
                            current_script_dir = os.path.dirname(os.path.abspath(__file__))
                            parent_dir = os.path.abspath(os.path.join(current_script_dir, os.pardir))
                            full_path_attestation = os.path.join(parent_dir, csv_filename)
                            # We use the helper but target the specific file path logic if needed, 
                            # usually download_supabase_to_csv handles local dir. 
                            # For this button, we usually just save to current dir for backup.
                            success, msg = download_supabase_to_csv(table_name, csv_filename)
                        else:
                            success, msg = download_supabase_to_csv(table_name, csv_filename)
                        
                        if success:
                            st.success(msg)
                        else:
                            st.warning(msg)
                            download_success = False
                    
                    if download_success:
                        st.success("🎉 All tables successfully downloaded as CSV files!")
                    else:
                        st.warning("⚠️ Some tables could not be downloaded. Check the messages above.")

            if st.button("🛑 Stop (Reset and Re-upload All CSVs)"):
                with st.spinner("Deleting all Supabase table rows..."):
                    # UPDATED: Added 'prep_closing_assignments' and 'global_settings' to delete order
                    table_order = [
                        "cs_reports",
                        "room_invigilator_assignments", 
                        "shift_assignments",
                        "exam_team_members",
                        "assigned_seats",
                        "sitting_plan",
                        "timetable",
                        "attestation_data_combined",
                        "prep_closing_assignments",
                        "global_settings"
                    ]

                    delete_errors = []
                    for table in table_order:
                        try:
                            supabase.table(table).delete().neq("id", 0).execute()  # delete all rows
                        except Exception as e:
                            delete_errors.append(f"❌ Error deleting from `{table}`: {str(e)}")

                if delete_errors:
                    st.error("\n".join(delete_errors))
                else:
                    st.success("✅ All existing Supabase table data deleted.")

                    # UPDATED: Added 'prep_closing_assignments' and 'global_settings' to upload mapping
                    csv_table_mapping = {
                        "timetable.csv": ("timetable", None),
                        "sitting_plan.csv": ("sitting_plan", None), 
                        "assigned_seats.csv": ("assigned_seats", None),
                        "exam_team_members.csv": ("exam_team_members", None),
                        "shift_assignments.csv": ("shift_assignments", None),
                        "room_invigilator_assignments.csv": ("room_invigilator_assignments", None),
                        "cs_reports.csv": ("cs_reports", None),
                        "attestation_data_combined.csv": ("attestation_data_combined", None),
                        "prep_closing_assignments.csv": ("prep_closing_assignments", None),
                        "global_settings.csv": ("global_settings", None)
                    }

                    st.markdown("### 📤 Uploading all CSVs to Supabase...")
                    for file, (table, keys) in csv_table_mapping.items():
                        current_script_dir = os.path.dirname(os.path.abspath(__file__))
                        full_path = os.path.join(current_script_dir, file)
                        
                        if os.path.exists(full_path):
                            success, msg = upload_csv_to_supabase(table, full_path, unique_cols=keys)
                        else:
                            # Warning only (some files like global_settings might not exist locally yet)
                            success, msg = False, f"File not found for upload: {full_path}"
                        
                        if success:
                            st.success(msg)
                        else:
                            st.warning(msg)

        elif admin_option == "Remuneration Bill Generation":
            st.subheader("💰 Remuneration Bill Generation")
            st.info("Calculate remuneration for exam team members based on their assignments.")

            # Load necessary dataframes
            shift_assignments_df = load_shift_assignments()
            room_invigilator_assignments_df = load_room_invigilator_assignments()
            
            _, timetable_df_for_remuneration, assigned_seats_df_for_remuneration, _ = load_data()

            if shift_assignments_df.empty and room_invigilator_assignments_df.empty:
                st.warning("No shift or room invigilator assignments found. Please make assignments first.")
                st.stop()
            
            # --- STEP 1: Calculate Eligible Members FIRST ---
            all_eligible_members = []
            for _, row in shift_assignments_df.iterrows():
                for role_col in ['senior_center_superintendent', 'center_superintendent', 'assistant_center_superintendent', 'permanent_invigilator', 'assistant_permanent_invigilator', 'invigilator']:
                    if role_col in row and isinstance(row[role_col], list):
                        # Normalize names here immediately (strip spaces)
                        cleaned_names = [str(n).strip() for n in row[role_col] if n]
                        all_eligible_members.extend(cleaned_names)
            all_eligible_members = sorted(list(set(all_eligible_members))) # Unique names

            # Get unique classes from the timetable for multi-selection
            all_classes_in_timetable = sorted(timetable_df_for_remuneration['Class'].dropna().astype(str).str.strip().unique().tolist())
            
            st.markdown("---")
            st.subheader("Select Classes for Bill Generation")
            st.info("Select specific classes to load saved prep/closing days for that batch.")
            
            if 'previous_selected_classes' not in st.session_state:
                st.session_state.previous_selected_classes = st.session_state.get('selected_classes_for_bill_state', [])

            selected_classes_for_bill = st.multiselect(
                "Select Classes (leave empty for all classes)",
                options=all_classes_in_timetable,
                default=st.session_state.get('selected_classes_for_bill_state', []),
                key="selected_classes_for_bill_multiselect"
            )

            # --- STEP 2: Load Data with Robust Matching ---
            # Check if selection changed OR if we haven't loaded data yet (empty input dict)
            data_needs_loading = (selected_classes_for_bill != st.session_state.previous_selected_classes) or \
                                 (not st.session_state.get('current_prep_closing_input'))

            if data_needs_loading:
                st.session_state.previous_selected_classes = selected_classes_for_bill
                st.session_state.selected_classes_for_bill_state = selected_classes_for_bill
                
                if 'current_prep_closing_input' not in st.session_state:
                    st.session_state.current_prep_closing_input = {}

                # 1. Fetch ALL rows
                response = supabase.table("prep_closing_assignments").select("*").execute()
                
                # 2. Filter data for CURRENT class selection
                matched_data_by_name = {}
                current_selected_set = set(selected_classes_for_bill)
                
                # Debug lists
                debug_found_in_db = []

                if response.data:
                    for row in response.data:
                        # --- FIX START: Handle both string (JSON) and list formats ---
                        raw_classes = row.get('selected_classes')
                        if isinstance(raw_classes, str):
                            row_classes = json.loads(raw_classes)
                        elif isinstance(raw_classes, list):
                            row_classes = raw_classes
                        else:
                            row_classes = []
                        # --- FIX END ---
                        
                        # Match sets (handling empty vs None)
                        if set(row_classes) == current_selected_set:
                            # Normalize DB name too
                            db_name = str(row.get('name')).strip()
                            
                            # --- FIX START: Safely parse prep/closing days ---
                            raw_prep = row.get('prep_days')
                            if isinstance(raw_prep, str):
                                prep_days = json.loads(raw_prep)
                            elif isinstance(raw_prep, list):
                                prep_days = raw_prep
                            else:
                                prep_days = []

                            raw_closing = row.get('closing_days')
                            if isinstance(raw_closing, str):
                                closing_days = json.loads(raw_closing)
                            elif isinstance(raw_closing, list):
                                closing_days = raw_closing
                            else:
                                closing_days = []
                            # --- FIX END ---

                            matched_data_by_name[db_name] = {
                                'role': row.get('role', 'senior_center_superintendent'),
                                'prep_days': prep_days,
                                'closing_days': closing_days
                            }
                            debug_found_in_db.append(db_name)

                # 3. Map to Inputs
                for member in all_eligible_members:
                    # member is already stripped/normalized in Step 1
                    if member in matched_data_by_name:
                        data = matched_data_by_name[member]
                        st.session_state.current_prep_closing_input[member] = {
                            'role': data['role'],
                            'prep_days': ", ".join(data['prep_days']),
                            'closing_days': ", ".join(data['closing_days'])
                        }
                    else:
                        st.session_state.current_prep_closing_input[member] = {
                            'role': 'senior_center_superintendent',
                            'prep_days': '',
                            'closing_days': ''
                        }
                
                # Store debug info in session state to persist after rerun
                st.session_state.debug_db_names = debug_found_in_db
                st.session_state.debug_eligible_members = all_eligible_members
                
                st.rerun()

            # --- Troubleshooting Section ---
            with st.expander("🛠️ Troubleshooting: Data Loading Status"):
                st.write(f"**Eligible Members (from assignments):** {st.session_state.get('debug_eligible_members', [])}")
                st.write(f"**Found in Database for this Class Selection:** {st.session_state.get('debug_db_names', [])}")
                st.info("If a name is in the 'Database' list but not loading below, check for spelling differences.")

            st.markdown("---")
            st.subheader("Manual Remuneration Rates (per shift/day)")

            manual_rates = {
                'senior_center_superintendent_rate': st.number_input("Senior Center Superintendent Rate (Rs./day - no conveyance on exam days)", min_value=0, value=200, key="scs_rate"),
                'center_superintendent_rate': st.number_input("Center Superintendent Rate (Rs.)", min_value=0, value=175, key="cs_rate"),
                'assistant_center_superintendent_rate': st.number_input("Assistant Center Superintendent Rate (Rs.)", min_value=0, value=150, key="acs_rate"),
                'permanent_invigilator_rate': st.number_input("Permanent Invigilator Rate (Rs.)", min_value=0, value=100, key="pi_rate"),
                'assistant_permanent_invigilator_rate': st.number_input("Assistant Permanent Invigilator Rate (Rs.)", min_value=0, value=100, key="api_rate"),
                'invigilator_rate': st.number_input("Invigilator Rate (Rs.)", min_value=0, value=100, key="inv_rate"),
                'conveyance_rate': st.number_input("Conveyance Rate (Evening shift - both shifts worked) (Rs.)", min_value=0, value=100, key="conveyance_rate"),
                'class_3_worker_rate_per_student': st.number_input("Class 3 Worker Rate (per student) (Rs.)", min_value=0.0, value=4.0, key="c3_rate"),
                'class_4_worker_rate_per_student': st.number_input("Class 4 Worker Rate (per student) (Rs.)", min_value=0.0, value=3.0, key="c4_rate"),
                'holiday_conveyance_allowance_rate': st.number_input("Holiday Conveyance Allowance (Rs.)", min_value=0, value=100, key="holiday_conveyance_rate")
            }

            st.markdown("---")
            st.subheader("Preparation and Closing Day Assignments")
            
            role_options = [
                'senior_center_superintendent', 'center_superintendent', 'assistant_center_superintendent',
                'permanent_invigilator', 'assistant_permanent_invigilator', 'invigilator'
            ]
            
            role_display_names = {
                'senior_center_superintendent': 'Senior Center Superintendent',
                'center_superintendent': 'Center Superintendent',
                'assistant_center_superintendent': 'Assistant Center Superintendent', 
                'permanent_invigilator': 'Permanent Invigilator',
                'assistant_permanent_invigilator': 'Assistant Permanent Invigilator',
                'invigilator': 'Invigilator'
            }

            if all_eligible_members:
                # Ensure session state is populated if we bypassed the loader (e.g. on simple re-renders)
                if 'current_prep_closing_input' not in st.session_state:
                     st.session_state.current_prep_closing_input = {}
                
                for member in all_eligible_members:
                    if member not in st.session_state.current_prep_closing_input:
                         st.session_state.current_prep_closing_input[member] = {'role': 'senior_center_superintendent', 'prep_days': '', 'closing_days': ''}

                for member in all_eligible_members:
                    st.markdown(f"**{member}**")
                    
                    current_role = st.session_state.current_prep_closing_input[member]['role']
                    role_index = role_options.index(current_role) if current_role in role_options else 0

                    selected_role = st.selectbox(
                        f"Select Role for {member}",
                        options=role_options,
                        format_func=lambda x: role_display_names[x],
                        index=role_index,
                        key=f"{member}_role_selection"
                    )
                    st.session_state.current_prep_closing_input[member]['role'] = selected_role

                    prep_days_input = st.text_input(
                        f"Preparation Days for {member} as {role_display_names[selected_role]} (comma-separated DD-MM-YYYY dates)",
                        value=st.session_state.current_prep_closing_input[member]['prep_days'],
                        key=f"{member}_prep_days"
                    )
                    st.session_state.current_prep_closing_input[member]['prep_days'] = prep_days_input

                    closing_days_input = st.text_input(
                        f"Closing Days for {member} as {role_display_names[selected_role]} (comma-separated DD-MM-YYYY dates)",
                        value=st.session_state.current_prep_closing_input[member]['closing_days'],
                        key=f"{member}_closing_days"
                    )
                    st.session_state.current_prep_closing_input[member]['closing_days'] = closing_days_input
            else:
                st.info("No eligible team members found for preparation/closing day assignments.")

            st.markdown("---")
            # Added unique key to prevent DuplicateElementId error
            if st.button("💾 Save Prep/Closing Assignments", key="save_prep_btn"):
                data_to_save_to_db = []
                for member, inputs in st.session_state.current_prep_closing_input.items():
                    prep_days_list = [d.strip() for d in inputs['prep_days'].split(',') if d.strip()]
                    closing_days_list = [d.strip() for d in inputs['closing_days'].split(',') if d.strip()]

                    valid_prep_days = []
                    for d in prep_days_list:
                        try:
                            datetime.datetime.strptime(d, '%d-%m-%Y')
                            valid_prep_days.append(d)
                        except ValueError:
                            st.warning(f"Invalid date format for {d} in prep days. Skipping.")
                    
                    valid_closing_days = []
                    for d in closing_days_list:
                        try:
                            datetime.datetime.strptime(d, '%d-%m-%Y')
                            valid_closing_days.append(d)
                        except ValueError:
                            st.warning(f"Invalid date format for {d} in closing days. Skipping.")

                    # Save stripped name
                    data_to_save_to_db.append({
                        'name': member.strip(),
                        'role': inputs['role'],
                        'prep_days': valid_prep_days,
                        'closing_days': valid_closing_days,
                        'selected_classes': selected_classes_for_bill
                    })
                
                success, message = save_prep_closing_assignments_to_supabase(data_to_save_to_db)
                if success:
                    st.success(message)
                else:
                    st.error(message)
                st.rerun()

            st.markdown("---")
            st.subheader("Holiday dates for Conveyance Allowance")
            
            loaded_holiday_dates = load_global_setting_from_supabase('holiday_dates')
            if loaded_holiday_dates and not st.session_state.get('holiday_dates_input_state'):
                st.session_state.holiday_dates_input_state = ", ".join(loaded_holiday_dates)

            holiday_dates_input = st.text_input(
                "Enter Holiday dates (comma-separated DD-MM-YYYY dates)",
                value=st.session_state.holiday_dates_input_state,
                key="holiday_dates_input"
            )
            st.session_state.holiday_dates_input_state = holiday_dates_input

            holiday_dates = [d.strip() for d in holiday_dates_input.split(',') if d.strip()]
            valid_holiday_dates = []
            for d in holiday_dates:
                try:
                    datetime.datetime.strptime(d, '%d-%m-%Y')
                    valid_holiday_dates.append(d)
                except ValueError:
                    pass
            holiday_dates = valid_holiday_dates

            # Added unique key here too
            if st.button("💾 Save Holiday dates", key="save_holiday_btn"):
                success, message = save_global_setting_to_supabase('holiday_dates', holiday_dates)
                if success:
                    st.success(message)
                else:
                    st.error(message)
                st.rerun()

            st.markdown("---")
            st.subheader("📋 Conveyance Rules Summary")
            st.info("Rules: SCS gets Rs 200/day (no exam day conveyance). Others Rs 100 evening conveyance. Holiday conveyance applies to prep/closing days.")

            # Added unique key here too
            if st.button("Generate Remuneration Bills", key="gen_bills_btn"):
                if shift_assignments_df.empty:
                    st.warning("shift assignments data is required.")
                elif assigned_seats_df_for_remuneration.empty:
                    st.warning("Assigned seats data is required.")
                else:
                    with st.spinner("Calculating remuneration..."):
                        dynamic_prep_closing_assignments_for_calc = {}
                        for member, inputs in st.session_state.current_prep_closing_input.items():
                            prep_days_list_calc = [d.strip() for d in inputs['prep_days'].split(',') if d.strip()]
                            closing_days_list_calc = [d.strip() for d in inputs['closing_days'].split(',') if d.strip()]
                            dynamic_prep_closing_assignments_for_calc[member] = {
                                'role': inputs['role'],
                                'prep_days': prep_days_list_calc,
                                'closing_days': closing_days_list_calc,
                                'selected_classes': selected_classes_for_bill
                            }

                        df_individual_bills, df_role_summary_matrix, df_class_3_4_final_bills = calculate_remuneration(
                            shift_assignments_df,
                            room_invigilator_assignments_df,
                            timetable_df_for_remuneration, 
                            assigned_seats_df_for_remuneration,
                            manual_rates,
                            dynamic_prep_closing_assignments_for_calc, 
                            holiday_dates, 
                            selected_classes_for_bill
                        )

                        st.markdown("### Individual Remuneration Bills")
                        if not df_individual_bills.empty:
                            df_individual_bills_with_total = add_total_row(df_individual_bills)
                            st.dataframe(df_individual_bills_with_total, use_container_width=True)
                        else:
                            st.info("No individual bills generated.")

                        st.markdown("### Role-wise Summary Matrix")
                        if not df_role_summary_matrix.empty:
                            df_role_summary_matrix_with_total = add_total_row(df_role_summary_matrix)
                            st.dataframe(df_role_summary_matrix_with_total, use_container_width=True)
                        else:
                            st.info("No role-wise summary generated.")

                        st.markdown("### Class 3 & Class 4 Worker Bills")
                        if not df_class_3_4_final_bills.empty:
                            df_class_3_4_final_bills_with_total = add_total_row(df_class_3_4_final_bills)
                            st.dataframe(df_class_3_4_final_bills_with_total, use_container_width=True)
                        else:
                            st.info("No Class 3 & 4 worker bills generated.")
                        
                        if not df_individual_bills.empty or not df_role_summary_matrix.empty or not df_class_3_4_final_bills.empty:
                            excel_file_buffer, excel_filename = save_bills_to_excel(
                                df_individual_bills_with_total, 
                                df_role_summary_matrix_with_total, 
                                df_class_3_4_final_bills_with_total
                            )
                            st.download_button(
                                label="Download All Remuneration Bills as Excel",
                                data=excel_file_buffer,
                                file_name=excel_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )
                            st.success(f"Remuneration bills generated and ready for download as '{excel_filename}'.")
                        else:
                            st.warning("No bills were generated to save.")



        elif admin_option == "Room Chart Report": # New Room Chart Report section
            st.subheader("📄 Room Chart Report")
            st.info("Generate a detailed room chart showing student seating arrangements for a specific exam session.")

            if sitting_plan.empty or timetable.empty or assigned_seats_df.empty:
                st.warning("Please upload 'sitting_plan.csv', 'timetable.csv', and ensure seats are assigned via 'Assign Rooms & Seats to Students' (Admin Panel) to generate this report.")
                st.stop() # Use st.stop() to halt execution if critical data is missing
            
            # date and shift filters for the room chart
            chart_date_options = sorted(timetable["date"].dropna().unique())
            chart_shift_options = sorted(timetable["shift"].dropna().unique())

            if not chart_date_options or not chart_shift_options:
                st.info("No exam dates or shifts found in the timetable to generate a room chart.")
                st.stop() # Use st.stop() to halt execution if no options

            selected_chart_date = st.selectbox("Select date", chart_date_options, key="room_chart_date")
            selected_chart_shift = st.selectbox("Select shift", chart_shift_options, key="room_chart_shift")

            if st.button("Generate Room Chart"):
                with st.spinner("Generating room chart..."):
                    # The generate_room_chart_report function now returns a string message if there's an error
                    room_chart_output = generate_room_chart_report(selected_chart_date, selected_chart_shift, sitting_plan, assigned_seats_df, timetable)
                    
                    # Check if the output is an error message (string) or the actual chart data
                    if room_chart_output and "Error:" in room_chart_output:
                        st.error(room_chart_output) # Display the error message
                    elif room_chart_output:
                        st.text_area("Generated Room Chart", room_chart_output, height=600)
                        
                        # Download button
                        file_name = f"room_chart_{selected_chart_date}_{selected_chart_shift}.csv"
                        st.download_button(
                            label="Download Room Chart as CSV",
                            data=room_chart_output.encode('utf-8'),
                            file_name=file_name,
                            mime="text/csv",
                        )
                    else:
                        st.warning("Could not generate room chart. Please ensure data is complete and assignments are made.")


        elif admin_option == "Data Processing & Reports":
            st.subheader("⚙️ Data Processing & Report Generation")

            st.markdown("---")
            st.subheader("Upload PDFs for Sitting Plan (pdf_folder.zip)")
            st.info(f"Upload a ZIP file containing folders of PDFs (e.g., 'pdf_folder/Zoology'). Each folder name will be used as the 'Paper' name. This will generate/update '{SITTING_PLAN_FILE}' and an initial '{TIMETABLE_FILE}'.")
            uploaded_sitting_plan_zip = st.file_uploader("Upload Sitting Plan PDFs (ZIP)", type=["zip"], key="upload_sitting_plan_zip")
            if uploaded_sitting_plan_zip:
                with st.spinner("Processing sitting plan PDFs and generating initial timetable... This may take a while."):
                    success, message = process_sitting_plan_pdfs(uploaded_sitting_plan_zip, SITTING_PLAN_FILE, TIMETABLE_FILE)
                    if success:
                        st.success(message)
                        # Reload data after processing
                        sitting_plan, timetable, assigned_seats_data, attestation_data = load_data()
                    else:
                        st.error(message)

            st.markdown("---")
            st.subheader("Upload Attestation PDFs (rasa_pdf.zip)")
            st.info(f"Upload a ZIP file containing attestation PDFs. These will be parsed to create a combined attestation data CSV ('{ATTESTATION_DATA_FILE}') and then automatically generate college statistics ('{COLLEGE_STATISTICS_FILE}').")
            uploaded_attestation_zip = st.file_uploader("Upload Attestation PDFs (ZIP)", type=["zip"], key="upload_attestation_zip")
            if uploaded_attestation_zip:
                with st.spinner("Processing attestation PDFs and generating college statistics... This may take a while."):
                    success, message = process_attestation_pdfs(uploaded_attestation_zip, ATTESTATION_DATA_FILE)
                    if success:
                        st.success(message)
                        # Automatically generate college statistics after attestation PDFs are processed
                        st.info("Generating college statistics...")
                        stats_success, stats_message = generate_college_statistics(ATTESTATION_DATA_FILE, COLLEGE_STATISTICS_FILE)
                        if stats_success:
                            st.success(stats_message)
                            if os.path.exists(COLLEGE_STATISTICS_FILE):
                                with open(COLLEGE_STATISTICS_FILE, "rb") as f:
                                    st.download_button(
                                        label="Download College Statistics CSV",
                                        data=f,
                                        file_name=COLLEGE_STATISTICS_FILE,
                                        mime="text/csv",
                                        key="download_college_stats_auto" # Unique key added
                                    )
                        else:
                            st.error(stats_message)
                    else:
                        st.error(message)

            st.markdown("---")
            st.subheader("Generate College Statistics (Manual Trigger)")
            st.info(f"This will generate college-wise statistics from '{ATTESTATION_DATA_FILE}' and save it to '{COLLEGE_STATISTICS_FILE}'. Only use if attestation data is already processed.")
            if st.button("Generate College Statistics (Manual)"):
                success, message = generate_college_statistics(ATTESTATION_DATA_FILE, COLLEGE_STATISTICS_FILE)
                if success:
                    st.success(message)
                    if os.path.exists(COLLEGE_STATISTICS_FILE):
                        with open(COLLEGE_STATISTICS_FILE, "rb") as f:
                            st.download_button(
                                label="Download College Statistics CSV",
                                data=f,
                                file_name=COLLEGE_STATISTICS_FILE,
                                mime="text/csv",
                                key="download_college_stats_manual" # Unique key added
                            )
                else:
                    st.error(message)

        elif admin_option == "Report Panel":
            display_report_panel()

    else:
        st.warning("Enter valid admin credentials.")

elif menu == "Centre Superintendent Panel":
    st.subheader("🔐 Centre Superintendent Login")
    if cs_login():
        st.success("Login successful!")

        # Load data for CS panel
        sitting_plan, timetable, assigned_seats_df, attestation_df = load_data() # Load assigned_seats_df here
        
        cs_panel_option = st.radio("Select CS Task:", ["Report Exam Session", "Manage Exam Team & shift Assignments", "View Full Timetable", "Room Chart Report", "Generate UFM Print Form"])

        if cs_panel_option == "Manage Exam Team & shift Assignments":
            st.subheader("👥 Manage Exam Team Members")
            
            current_members = load_exam_team_members()
            new_member_name = st.text_input("Add New Team Member Name")
            if st.button("Add Member"):
                if new_member_name and new_member_name not in current_members:
                    current_members.append(new_member_name)
                    success, msg = save_exam_team_members(current_members)
                    if success:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)
                elif new_member_name:
                    st.warning("Member already exists.")
                else:
                    st.warning("Please enter a name.")

            if current_members:
                st.write("Current Team Members:")
                st.write(current_members)
                
                member_to_remove = st.selectbox("Select Member to Remove", [""] + current_members)
                if st.button("Remove Selected Member"):
                    if member_to_remove:
                        current_members.remove(member_to_remove)
                        success, msg = save_exam_team_members(current_members)
                        if success:
                            st.success(msg)
                            st.rerun()
                        else:
                            st.error(msg)
                    else:
                        st.warning("Please select a member to remove.")
            else:
                st.info("No team members added yet.")

            st.markdown("---")
            st.subheader("🗓️ Assign Roles for Exam shift")

            assignment_date = st.date_input("Select date for Assignment", value=datetime.date.today(), key="assignment_date")
            assignment_shift = st.selectbox("Select shift for Assignment", ["Morning", "Evening"], key="assignment_shift")

            all_team_members = load_exam_team_members()
            if not all_team_members:
                st.warning("Please add exam team members first in the 'Manage Exam Team Members' section.")
            else:
                current_assignments_df = load_shift_assignments()
                current_assignment_for_shift = current_assignments_df[
                    (current_assignments_df['date'] == assignment_date.strftime('%d-%m-%Y')) &
                    (current_assignments_df['shift'] == assignment_shift)
                ]
                
                loaded_senior_cs = []
                loaded_cs = []
                loaded_assist_cs = []
                loaded_perm_inv = []
                loaded_assist_perm_inv = []
                loaded_class_3 = []
                loaded_class_4 = []


                if not current_assignment_for_shift.empty:
                    loaded_senior_cs = current_assignment_for_shift.iloc[0].get('senior_center_superintendent', [])
                    loaded_cs = current_assignment_for_shift.iloc[0].get('center_superintendent', [])
                    loaded_assist_cs = current_assignment_for_shift.iloc[0].get('assistant_center_superintendent', [])
                    loaded_perm_inv = current_assignment_for_shift.iloc[0].get('permanent_invigilator', [])
                    loaded_assist_perm_inv = current_assignment_for_shift.iloc[0].get('assistant_permanent_invigilator', [])
                    loaded_class_3 = current_assignment_for_shift.iloc[0].get('class_3_worker', [])
                    loaded_class_4 = current_assignment_for_shift.iloc[0].get('class_4_worker', [])


                selected_senior_cs = st.multiselect("Senior Center Superintendent (Max 1)", all_team_members, default=loaded_senior_cs, max_selections=1)
                selected_cs = st.multiselect("Center Superintendent (Max 1)", all_team_members, default=loaded_cs, max_selections=1)
                selected_assist_cs = st.multiselect("Assistant Center Superintendent (Max 3)", all_team_members, default=loaded_assist_cs, max_selections=3)
                selected_perm_inv = st.multiselect("Permanent Invigilator (Max 3)", all_team_members, default=loaded_perm_inv, max_selections=3)
                selected_assist_perm_inv = st.multiselect("Assistant Permanent Invigilator (Max 5)", all_team_members, default=loaded_assist_perm_inv, max_selections=5)
                selected_class_3 = st.multiselect("Class 3 Worker (Max 10)", all_team_members, default=loaded_class_3, max_selections=10)
                selected_class_4 = st.multiselect("Class 4 Worker (Max 10)", all_team_members, default=loaded_class_4, max_selections=10)


                if st.button("Save shift Assignments"):
                    all_selected_members = (
                        selected_senior_cs + selected_cs + selected_assist_cs +
                        selected_perm_inv + selected_assist_perm_inv + selected_class_3 + selected_class_4
                    )
                    if len(all_selected_members) != len(set(all_selected_members)):
                        st.error("Error: A team member cannot be assigned to multiple roles for the same shift.")
                    else:
                        assignments_to_save = {
                            'senior_center_superintendent': selected_senior_cs,
                            'center_superintendent': selected_cs,
                            'assistant_center_superintendent': selected_assist_cs,
                            'permanent_invigilator': selected_perm_inv,
                            'assistant_permanent_invigilator': selected_assist_perm_inv,
                            'class_3_worker': selected_class_3,
                            'class_4_worker': selected_class_4
                        }
                        success, msg = save_shift_assignment(assignment_date.strftime('%d-%m-%Y'), assignment_shift, assignments_to_save)
                        if success:
                            st.success(msg)
                            st.rerun()
                        else:
                            st.error(msg)
                
                st.markdown("---")
                st.subheader("Current shift Assignments")
                display_assignments_df = load_shift_assignments()
                if not display_assignments_df.empty:
                    st.dataframe(display_assignments_df)
                else:
                    st.info("No shift assignments saved yet.")

            st.markdown("---")

            st.subheader("Assign Invigilators to Rooms")
            if assigned_seats_df.empty: # Check assigned_seats_df instead of sitting_plan/timetable for rooms
                st.info("Please assign seats to students via the Admin Panel's 'Assign Rooms & Seats to Students' section first to see available rooms for invigilator assignment.")
            else:
                room_inv_date = st.date_input("Select date for Room Invigilators", value=datetime.date.today(), key="room_inv_date")
                room_inv_shift = st.selectbox("Select shift for Room Invigilators", ["Morning", "Evening"], key="room_inv_shift")
                
                # MODIFIED: Get unique rooms for the selected date and shift from assigned_seats_df
                relevant_rooms_assigned = assigned_seats_df[
                    (assigned_seats_df["date"].astype(str).str.strip() == room_inv_date.strftime('%d-%m-%Y')) &
                    (assigned_seats_df["shift"].astype(str).str.strip().str.lower() == room_inv_shift.lower())
                ]
                
                unique_relevant_rooms = sorted(list(relevant_rooms_assigned['Room Number'].dropna().astype(str).str.strip().unique()))

                selected_room_for_inv = st.selectbox("Select Room to Assign Invigilators", [""] + unique_relevant_rooms, key="selected_room_for_inv")

                if selected_room_for_inv:
                    current_room_inv_df = load_room_invigilator_assignments()
                    loaded_invigilators = []
                    
                    filtered_inv_for_room = current_room_inv_df[
                        (current_room_inv_df['date'] == room_inv_date.strftime('%d-%m-%Y')) &
                        (current_room_inv_df['shift'] == room_inv_shift) &
                        (current_room_inv_df['room_num'] == selected_room_for_inv)
                    ]
                    
                    if not filtered_inv_for_room.empty:
                        loaded_invigilators = filtered_inv_for_room.iloc[0].get('invigilators', [])
                    
                    invigilators_for_room = st.multiselect(
                        "Invigilators for this Room",
                        options=all_team_members, # Use the same team members list
                        default=loaded_invigilators,
                        key="invigilators_for_room_multiselect"
                    )

                    if st.button("Save Room Invigilators"):
                        success, msg = save_room_invigilator_assignment(
                            room_inv_date.strftime('%d-%m-%Y'),
                            room_inv_shift,
                            selected_room_for_inv,
                            invigilators_for_room
                        )
                        if success:
                            st.success(msg)
                            st.rerun()
                        else:
                            st.error(msg)
                else:
                    st.info("Select a date, shift, and room to assign invigilators.")
                
                st.markdown("---")
                st.subheader("Current Room Invigilator Assignments")
                display_room_inv_df = load_room_invigilator_assignments()
                if not display_room_inv_df.empty:
                    st.dataframe(display_room_inv_df)
                else:
                    st.info("No room invigilator assignments saved yet.")
                    
                    


        elif cs_panel_option == "Report Exam Session":
            st.subheader("📝 Report Exam Session")
            if assigned_seats_df.empty or timetable.empty: # Check assigned_seats_df instead of sitting_plan
                st.info("Please ensure seats are assigned and 'timetable.csv' is uploaded via the Admin Panel to report exam sessions.")
            else:
                # date and shift selection
                report_date = st.date_input("Select date", value=datetime.date.today(), key="cs_report_date")
                report_shift = st.selectbox("Select shift", ["Morning", "Evening"], key="cs_report_shift")

                # Filter assigned_seats_df for selected date and shift to get available exam sessions
                available_sessions_assigned = assigned_seats_df[
                    (assigned_seats_df["date"].astype(str).str.strip() == report_date.strftime('%d-%m-%Y')) &
                    (assigned_seats_df["shift"].astype(str).str.strip().str.lower() == report_shift.lower())
                ].copy()

                if available_sessions_assigned.empty:
                    st.warning("No assigned seats found for the selected date and shift. Please assign seats via the Admin Panel first.")
                else:
                    # Create a unique identifier for each exam session (Room - Paper Code (Paper Name))
                    # Ensure Paper Code and Paper Name are strings before combining
                    available_sessions_assigned['Paper Code'] = available_sessions_assigned['Paper Code'].astype(str)
                    available_sessions_assigned['Paper Name'] = available_sessions_assigned['Paper Name'].astype(str)

                    available_sessions_assigned['exam_session_id'] = \
                        available_sessions_assigned['Room Number'].astype(str).str.strip() + " - " + \
                        available_sessions_assigned['Paper Code'].apply(_format_paper_code) + " (" + \
                        available_sessions_assigned['Paper Name'].str.strip() + ")"
                    
                    unique_exam_sessions = available_sessions_assigned[['Room Number', 'Paper Code', 'Paper Name', 'exam_session_id']].drop_duplicates().sort_values(by='exam_session_id')
                    
                    if unique_exam_sessions.empty:
                        st.warning("No unique exam sessions found for the selected date and shift in assigned seats.")
                    else:
                        selected_exam_session_option = st.selectbox(
                            "Select Exam Session (Room - Paper Code (Paper Name))",
                            [""] + unique_exam_sessions['exam_session_id'].tolist(),
                            key="cs_exam_session_select"
                        )

                        if selected_exam_session_option:
                            # Extract room_number, paper_code, paper_name from the selected option
                            selected_room_num = selected_exam_session_option.split(" - ")[0].strip()
                            selected_paper_code_with_name = selected_exam_session_option.split(" - ", 1)[1].strip()
                            selected_paper_code = _format_paper_code(selected_paper_code_with_name.split(" (")[0]) # Format the extracted code
                            selected_paper_name = selected_paper_code_with_name.split(" (")[1].replace(")", "").strip()

                            # Find the corresponding class for the selected session from timetable
                            # This assumes a paper code/name maps to a consistent class in the timetable
                            matching_class_info = timetable[
                                (timetable['Paper Code'].astype(str).str.strip() == selected_paper_code) & # Use formatted paper code
                                (timetable['Paper Name'].astype(str).str.strip() == selected_paper_name)
                            ]
                            selected_class = ""
                            if not matching_class_info.empty:
                                selected_class = str(matching_class_info.iloc[0]['Class']).strip()

                            # Create a unique key for CSV row ID
                            report_key = f"{report_date.strftime('%Y%m%d')}_{report_shift.lower()}_{selected_room_num}_{selected_paper_code}"

                            # Load existing report from CSV
                            loaded_success, loaded_report = load_single_cs_report_csv(report_key)
                            if loaded_success:
                                st.info("Existing report loaded.")
                            else:
                                st.info("No existing report found for this session. Starting new.")
                                loaded_report = {} # Ensure it's an empty dict if not found

                            # MODIFIED: Get all *assigned* roll numbers for this specific session from assigned_seats_df
                            expected_students_for_session = assigned_seats_df[
                                (assigned_seats_df['Room Number'].astype(str).str.strip() == selected_room_num) &
                                (assigned_seats_df['Paper Code'].astype(str).str.strip() == selected_paper_code) & # Use formatted paper code
                                (assigned_seats_df['Paper Name'].astype(str).str.strip() == selected_paper_name) &
                                (assigned_seats_df['date'].astype(str).str.strip() == report_date.strftime('%d-%m-%Y')) &
                                (assigned_seats_df['shift'].astype(str).str.strip().str.lower() == report_shift.lower())
                            ]['Roll Number'].astype(str).tolist()
                            
                            expected_students_for_session = sorted(list(set(expected_students_for_session))) # Remove duplicates and sort

                            st.write(f"**Reporting for:** Room {selected_room_num}, Paper: {selected_paper_name} ({selected_paper_code})")

                            # Multiselect for Absent Roll Numbers
                            absent_roll_numbers_selected = st.multiselect(
                                "Absent Roll Numbers", 
                                options=expected_students_for_session, 
                                default=loaded_report.get('absent_roll_numbers', []),
                                key="absent_roll_numbers_multiselect"
                            )

                            # Multiselect for UFM Roll Numbers
                            ufm_roll_numbers_selected = st.multiselect(
                                "UFM (Unfair Means) Roll Numbers", 
                                options=expected_students_for_session, 
                                default=loaded_report.get('ufm_roll_numbers', []),
                                key="ufm_roll_numbers_multiselect"
                            )

                            col1, col2 = st.columns(2)
                            with col1:
                                if st.button("Save Report", key="save_cs_report"):
                                    # --- Validation Logic ---
                                    expected_set = set(expected_students_for_session)
                                    absent_set = set(absent_roll_numbers_selected)
                                    ufm_set = set(ufm_roll_numbers_selected)

                                    validation_errors = []

                                    # 1. All reported absent students must be in the expected list
                                    if not absent_set.issubset(expected_set):
                                        invalid_absent = list(absent_set.difference(expected_set))
                                        validation_errors.append(f"Error: Absent roll numbers {invalid_absent} are not in the expected student list for this session.")

                                    # 2. All reported UFM students must be in the expected list
                                    if not ufm_set.issubset(expected_set):
                                        invalid_ufm = list(ufm_set.difference(expected_set))
                                        validation_errors.append(f"Error: UFM roll numbers {invalid_ufm} are not in the expected student list for this session.")

                                    # 3. No student can be both absent and UFM
                                    if not absent_set.isdisjoint(ufm_set):
                                        overlap = list(absent_set.intersection(ufm_set))
                                        validation_errors.append(f"Error: Roll numbers {overlap} are marked as both Absent and UFM. A student cannot be both.")
                                    
                                    if validation_errors:
                                        for err in validation_errors:
                                            st.error(err)
                                    else:
                                        report_data = {
                                            'report_key': report_key, # Add report_key to data
                                            'date': report_date.strftime('%d-%m-%Y'),
                                            'shift': report_shift,
                                            'room_num': selected_room_num,
                                            'paper_code': selected_paper_code,
                                            'paper_name': selected_paper_name,
                                            'class': selected_class, # Added 'class' here
                                            'absent_roll_numbers': absent_roll_numbers_selected, # Store as list
                                            'ufm_roll_numbers': ufm_roll_numbers_selected # Store as list
                                        }
                                        success, message = save_cs_report_csv(report_key, report_data)
                                        if success:
                                            st.success(message)
                                        else:
                                            st.error(message)
                                        st.rerun() # Rerun to refresh the UI with saved data

                                st.markdown("---")
                                st.subheader("All Saved Reports (for debugging/review)")
                                
                                # Fetch all reports for the current CS user from CSV
                                all_reports_df_display = load_cs_reports_csv()
                                room_invigilators_df_display = load_room_invigilator_assignments()

                                # CORRECT: Standardize column names immediately after loading the DataFrame
                                room_invigilators_df_display.columns = room_invigilators_df_display.columns.str.lower()

                                if not all_reports_df_display.empty:
                                    # Merge with room invigilators for display
                                    if not room_invigilators_df_display.empty:
                                        all_reports_df_display = pd.merge(
                                            all_reports_df_display,
                                            room_invigilators_df_display[['date', 'shift', 'room_num', 'invigilators']],
                                            on=['date', 'shift', 'room_num'],
                                            how='left',
                                            suffixes=('', '_room_inv_display')
                                        )

                                        all_reports_df_display['invigilators'] = all_reports_df_display['invigilators'].apply(lambda x: x if isinstance(x, list) else [])
                                    else:
                                        all_reports_df_display['invigilators'] = [[]] * len(all_reports_df_display)

                                    # Reorder columns for better readability
                                    display_cols = [
                                        "date", "shift", "room_num", "paper_code", "paper_name", "class", 
                                        "invigilators", "absent_roll_numbers", "ufm_roll_numbers", "report_key"
                                    ]
                                    
                                    # This line is no longer needed here as it was moved up
                                    # room_invigilators_df_display.columns = room_invigilators_df_display.columns.str.lower()

                                    # Map internal keys to display keys
                                    df_all_reports_display = all_reports_df_display.rename(columns={
                                        'date': 'date', 'shift': 'shift', 'room_num': 'Room',
                                        'paper_code': 'Paper Code', 'paper_name': 'Paper Name', 'class': 'Class', 
                                        'invigilators': 'Invigilators',
                                        'absent_roll_numbers': 'Absent Roll Numbers',
                                        'ufm_roll_numbers': 'UFM Roll Numbers',
                                        'report_key': 'Report Key'
                                    })
                                    
                                    # Ensure all display_cols exist, fill missing with empty string
                                    for col in display_cols:
                                        if col not in df_all_reports_display.columns:
                                            df_all_reports_display[col] = ""
                                    
                                    st.dataframe(df_all_reports_display[
                                        ['date', 'shift', 'Room', 'Paper Code', 'Paper Name', 'Class', 
                                            'Invigilators', 'Absent Roll Numbers', 'UFM Roll Numbers', 'Report Key']
                                    ])
                                else:
                                    st.info("No reports saved yet.")


        # ... (previous cs_panel_option elif blocks) ...

        elif cs_panel_option == "Generate UFM Print Form":
            st.subheader("🖨️ Generate UFM Print Form")
            st.info("Select a session date and shift to view reported UFM cases and generate their print forms.")

            # Load all data, including attestation_df
            sitting_plan, timetable, assigned_seats_data, attestation_data = load_data()
            all_cs_reports_df = load_cs_reports_csv()

            if attestation_df.empty:
                st.warning("Attestation data ('attestation_data_combined.csv') is missing. Please upload it via the Admin Panel to generate UFM forms.")
                st.stop()
            if all_cs_reports_df.empty:
                st.info("No CS reports available yet. UFM cases must be reported first in 'Report Exam Session'.")
                st.stop()

            # Get unique dates and shifts from reports that have UFM cases
            reports_with_ufm = all_cs_reports_df[all_cs_reports_df['ufm_roll_numbers'].apply(lambda x: len(x) > 0)]

            if reports_with_ufm.empty:
                st.info("No UFM cases have been reported yet in any session.")
                st.stop()

            unique_report_dates = sorted(reports_with_ufm['date'].unique())
            unique_report_shifts = sorted(reports_with_ufm['shift'].unique())

            selected_ufm_report_date = st.selectbox("Select date of UFM Report", unique_report_dates, key="ufm_report_date_select")
            selected_ufm_report_shift = st.selectbox("Select shift of UFM Report", unique_report_shifts, key="ufm_report_shift_select")

            # Filter reports for the selected date and shift
            filtered_ufm_reports = reports_with_ufm[
                (reports_with_ufm['date'] == selected_ufm_report_date) &
                (reports_with_ufm['shift'] == selected_ufm_report_shift)
            ]

            # Extract all unique UFM roll numbers for the selected date/shift and paper info
            ufm_roll_numbers_details = [] # Stores (roll_num, paper_code, paper_name, room_num)
            for _, row in filtered_ufm_reports.iterrows():
                room = str(row['room_num']).strip()
                paper_code = str(row['paper_code']).strip()
                paper_name = str(row['paper_name']).strip()
                for ufm_roll in row['ufm_roll_numbers']:
                    ufm_roll_numbers_details.append({
                        "roll_number": ufm_roll,
                        "paper_code": paper_code,
                        "paper_name": paper_name,
                        "room_num": room,
                        "display": f"{ufm_roll} - {room} - {paper_code} ({paper_name})"
                    })

            if not ufm_roll_numbers_details:
                st.info("No UFM cases found for the selected date and shift.")
            else:
                ufm_options = [d["display"] for d in ufm_roll_numbers_details]
                selected_ufm_display = st.multiselect(
                    "Select UFM Roll Number(s) to Generate Form",
                    options=ufm_options,
                    key="select_ufm_roll_for_form"
                )

                if st.button("Generate UFM Form(s)"):
                    if not selected_ufm_display:
                        st.warning("Please select at least one UFM roll number.")
                    else:
                        all_generated_forms = []
                        for display_string in selected_ufm_display:
                            # Find the corresponding detail dictionary
                            selected_detail = next(item for item in ufm_roll_numbers_details if item["display"] == display_string)

                            ufm_roll = selected_detail["roll_number"]
                            ufm_paper_code = selected_detail["paper_code"]
                            ufm_paper_name = selected_detail["paper_name"]

                            # Get the actual UFM form generation function
                            generate_ufm_form_func = generate_ufm_print_form(
                                ufm_roll, 
                                attestation_df, 
                                assigned_seats_df, 
                                timetable,
                                # Pass context for accurate data retrieval
                                selected_ufm_report_date,
                                selected_ufm_report_shift,
                                ufm_paper_code,
                                ufm_paper_name
                            )

                            if "Error:" in generate_ufm_form_func:
                                st.error(generate_ufm_form_func)
                            else:
                                all_generated_forms.append(generate_ufm_form_func)
                                st.subheader(f"UFM Form for Roll Number: {ufm_roll}")
                                st.text_area(f"Form for {ufm_roll}", generate_ufm_form_func, height=600)

                                # Download button for individual UFM form
                                ufm_file_name = f"UFM_Form_{ufm_roll}_{selected_ufm_report_date.replace('-', '')}_{ufm_paper_code}.txt"
                                st.download_button(
                                    label=f"Download Form for {ufm_roll}",
                                    data=generate_ufm_form_func.encode('utf-8'),
                                    file_name=ufm_file_name,
                                    mime="text/plain",
                                    key=f"download_ufm_{ufm_roll}"
                                )
                                st.markdown("---") # Separator between forms

                        if len(all_generated_forms) > 1:
                            combined_forms_text = "\n\n" + "-"*50 + "\n\n".join(all_generated_forms)
                            combined_file_name = f"Combined_UFM_Forms_{selected_ufm_report_date.replace('-', '')}_{selected_ufm_report_shift}.txt"
                            st.download_button(
                                label="Download All Selected UFM Forms (Combined)",
                                data=combined_forms_text.encode('utf-8'),
                                file_name=combined_file_name,
                                mime="text/plain",
                                key="download_all_ufm_forms"
                            )


        elif cs_panel_option == "View Full Timetable": # New section for CS timetable view
            st.subheader("Full Examination Timetable")
            if timetable.empty:
                st.warning("Timetable data is missing. Please upload it via the Admin Panel.")
            else:
                st.dataframe(timetable)

        elif cs_panel_option == "Room Chart Report": # New Room Chart Report section for CS
            st.subheader("📄 Room Chart Report")
            st.info("Generate a detailed room chart showing student seating arrangements for a specific exam session.")

            if sitting_plan.empty or timetable.empty or assigned_seats_df.empty:
                st.warning("Please upload 'sitting_plan.csv', 'timetable.csv', and ensure seats are assigned via 'Assign Rooms & Seats to Students' (Admin Panel) to generate this report.")
                st.stop() # Use st.stop() to halt execution if critical data is missing
            
            # date and shift filters for the room chart
            chart_date_options = sorted(timetable["date"].dropna().unique())
            chart_shift_options = sorted(timetable["shift"].dropna().unique())

            if not chart_date_options or not chart_shift_options:
                st.info("No exam dates or shifts found in the timetable to generate a room chart.")
                st.stop() # Use st.stop() to halt execution if no options

            selected_chart_date = st.selectbox("Select date", chart_date_options, key="cs_room_chart_date")
            selected_chart_shift = st.selectbox("Select shift", chart_shift_options, key="cs_room_chart_shift")

            if st.button("Generate Room Chart"):
                with st.spinner("Generating room chart..."):
                    # The generate_room_chart_report function now returns a string message if there's an error
                    room_chart_output = generate_room_chart_report(selected_chart_date, selected_chart_shift, sitting_plan, assigned_seats_df, timetable)
                    
                    # Check if the output is an error message (string) or the actual chart data
                    if room_chart_output and "Error:" in room_chart_output:
                        st.error(room_chart_output) # Display the error message
                    elif room_chart_output:
                        st.text_area("Generated Room Chart", room_chart_output, height=600)
                        
                        # Download button
                        file_name = f"room_chart_{selected_chart_date}_{selected_chart_shift}.csv"
                        st.download_button(
                            label="Download Room Chart as CSV",
                            data=room_chart_output.encode('utf-8'),
                            file_name=file_name,
                            mime="text/csv",
                        )
                    else:
                        st.warning("Could not generate room chart. Please ensure data is complete and assignments are made.")

    else:
        st.warning("Enter valid Centre Superintendent credentials.")
